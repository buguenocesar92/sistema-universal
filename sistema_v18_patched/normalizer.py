"""
KraftDo — normalizer.py
Detecta y normaliza patrones raros de Excel:

  PATRÓN 1 — Tabla vertical estándar (el caso feliz)
  PATRÓN 2 — Tabla horizontal/pivot (días, meses como columnas)
  PATRÓN 3 — Multi-header (headers en 2+ filas)
  PATRÓN 4 — Formulario vertical (campo:valor en pares)
  PATRÓN 5 — Múltiples tablas en la misma hoja
  PATRÓN 6 — Totales/subtotales mezclados con datos
  PATRÓN 7 — Tabla con columnas variables (ej: meses que crecen)
  PATRÓN 8 — Tabla sparse (muchas celdas vacías, datos dispersos)

Para cada patrón retorna:
  - diagnóstico: qué detectó
  - confianza: alta / media / baja
  - transformacion: cómo normalizarlo
  - requiere_humano: True si necesita confirmación manual
  - preview: muestra de los datos normalizados
"""

import re
from datetime import datetime, date


# ── Helpers ───────────────────────────────────────────────────────────────────

def _val(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return str(v)
    s = str(v).strip()
    return s if s else None

def _es_header(v) -> bool:
    """Heurística para detectar si un valor parece header."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s:
        return False
    # Headers suelen ser cortos, sin números puros, con letras
    if len(s) > 60:
        return False
    if re.match(r'^\d+[\.,]?\d*$', s):
        return False  # número puro
    return bool(re.search(r'[a-záéíóúñA-ZÁÉÍÓÚÑ]', s))

def _es_total(v) -> bool:
    """Detecta filas de totales/subtotales."""
    if not isinstance(v, str):
        return False
    palabras = ['total', 'subtotal', 'suma', 'promedio', 'average',
                'grand total', 'totales', '=sum', 'sum(']
    return any(p in v.lower() for p in palabras)

def _densidad_fila(ws, row_idx: int, max_col: int) -> float:
    """Proporción de celdas no vacías en una fila."""
    if max_col == 0:
        return 0
    llenas = sum(1 for c in range(1, max_col + 1)
                 if ws.cell(row_idx, c).value is not None)
    return llenas / max_col

def _densidad_col(ws, col_idx: int, max_row: int) -> float:
    """Proporción de celdas no vacías en una columna."""
    if max_row == 0:
        return 0
    llenas = sum(1 for r in range(1, max_row + 1)
                 if ws.cell(r, col_idx).value is not None)
    return llenas / max_row

def _es_fecha_o_mes(v) -> bool:
    """Detecta si un valor parece fecha, mes o día de la semana."""
    if v is None:
        return False
    if isinstance(v, (datetime, date)):
        return True
    s = str(v).lower().strip()
    meses = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic',
             'enero','febrero','marzo','abril','mayo','junio','julio','agosto',
             'septiembre','octubre','noviembre','diciembre']
    dias  = ['lun','mar','mié','jue','vie','sáb','dom',
             'lunes','martes','miércoles','jueves','viernes','sábado','domingo']
    return any(s.startswith(m) for m in meses + dias) or re.match(r'\d{1,2}/\d{1,2}', s)


# ═══════════════════════════════════════════════════════════════════════════════
# DETECTOR DE PATRONES
# ═══════════════════════════════════════════════════════════════════════════════

def detectar_patron(ws, fila_ini_hint: int = 1) -> dict:
    """
    Analiza una hoja de Excel y detecta su patrón estructural.

    Retorna:
    {
      "patron":         "vertical" | "horizontal" | "multi_header" |
                        "formulario" | "multi_tabla" | "con_totales" |
                        "columnas_variables" | "sparse",
      "confianza":      "alta" | "media" | "baja",
      "descripcion":    "Texto legible del diagnóstico",
      "fila_headers":   int,    # fila donde están los headers
      "fila_datos":     int,    # primera fila de datos
      "ultima_fila":    int,
      "max_col":        int,
      "filas_totales":  list,   # índices de filas de totales detectadas
      "bloques":        list,   # para multi_tabla: lista de (fila_ini, fila_fin)
      "eje_pivot":      str,    # para horizontal: "filas" o "columnas"
      "headers_fila1":  list,
      "headers_fila2":  list,   # para multi_header
      "requiere_humano": bool,
      "transformacion": str,    # descripción de qué hacer
      "preview_normal": list,   # primeras filas normalizadas
    }
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return _resultado("vacia", "baja", "Hoja vacía o sin datos detectables",
                          1, 2, 0, 0, True, "Ignorar esta hoja")

    # ── Buscar fila de headers ─────────────────────────────────────────────
    fila_headers = None
    for r in range(1, min(8, max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        headers_en_fila = sum(1 for v in row_vals if _es_header(v))
        if headers_en_fila >= 3:
            fila_headers = r
            break

    if fila_headers is None:
        fila_headers = 1

    fila_datos = fila_headers + 1
    headers_f1 = [ws.cell(fila_headers, c).value for c in range(1, max_col + 1)]

    # ── PATRÓN: Formulario vertical (campo:valor) ──────────────────────────
    # Detectar si la primera columna tiene labels y la segunda tiene valores
    if max_col <= 3:
        pares_label_valor = 0
        for r in range(fila_headers, min(fila_headers + 15, max_row + 1)):
            col1 = ws.cell(r, 1).value
            col2 = ws.cell(r, 2).value
            if _es_header(col1) and col2 is not None:
                pares_label_valor += 1
        if pares_label_valor >= 3:
            return _resultado(
                "formulario", "alta",
                f"Formulario vertical con {pares_label_valor} pares campo:valor. "
                "Representa una sola entidad (una fila en BD).",
                fila_headers, fila_datos, max_row, max_col,
                False,
                "Convertir a tabla config con columnas = campos. Una sola fila en la BD.",
                _preview_formulario(ws, fila_headers, max_row)
            )

    # ── PATRÓN: Tabla horizontal / pivot ──────────────────────────────────
    # Si la primera fila tiene muchas fechas/meses/días
    headers_son_fechas = sum(1 for h in headers_f1 if _es_fecha_o_mes(h))
    if headers_son_fechas >= 3:
        return _resultado(
            "horizontal", "alta",
            f"Tabla horizontal/pivot: {headers_son_fechas} columnas son fechas/meses/días. "
            "Necesita unpivot para normalizar.",
            fila_headers, fila_datos, max_row, max_col,
            True,
            "Unpivot: convertir columnas de fechas a filas. "
            "Resultado: tabla(entidad, fecha, valor).",
            _preview_unpivot(ws, fila_headers, fila_datos, max_row, max_col)
        )

    # Alternativa: muchas columnas numéricas con header de período
    if max_col > 10:
        cols_con_fechas = sum(1 for c in range(2, max_col + 1)
                              if _es_fecha_o_mes(ws.cell(fila_headers, c).value))
        if cols_con_fechas > max_col * 0.5:
            return _resultado(
                "horizontal", "media",
                f"Posible tabla horizontal: más del 50% de columnas ({cols_con_fechas}/{max_col}) "
                "parecen períodos de tiempo.",
                fila_headers, fila_datos, max_row, max_col,
                True,
                "Unpivot: convertir columnas de períodos a filas.",
                _preview_unpivot(ws, fila_headers, fila_datos, max_row, max_col)
            )

    # ── PATRÓN: Multi-header (2 filas de headers) ─────────────────────────
    if fila_headers < max_row:
        headers_f2 = [ws.cell(fila_headers + 1, c).value
                      for c in range(1, max_col + 1)]
        headers_en_f2 = sum(1 for v in headers_f2 if _es_header(v))
        headers_en_f1 = sum(1 for v in headers_f1 if _es_header(v))

        # Si la segunda fila también tiene headers (y no son datos)
        if headers_en_f2 >= 3 and headers_en_f2 >= headers_en_f1 * 0.5:
            # Verificar que no sean datos numéricos
            datos_f2 = sum(1 for v in headers_f2
                           if v is not None and re.match(r'^\d', str(v)))
            if datos_f2 < headers_en_f2:
                return _resultado(
                    "multi_header", "alta",
                    f"Headers en 2 filas: fila {fila_headers} y fila {fila_headers+1}. "
                    "Columnas tienen categoría + subcategoría.",
                    fila_headers, fila_headers + 2, max_row, max_col,
                    True,
                    "Combinar las 2 filas de headers: 'Categoría_Subcategoría'. "
                    "Ejemplo: 'Ventas_Enero', 'Ventas_Febrero'.",
                    _preview_multi_header(ws, fila_headers, max_row, max_col)
                )

    # ── PATRÓN: Múltiples tablas en la misma hoja ─────────────────────────
    bloques = _detectar_bloques(ws, fila_datos, max_row, max_col)
    if len(bloques) > 1:
        return _resultado(
            "multi_tabla", "media",
            f"Se detectaron {len(bloques)} bloques de datos separados en la misma hoja. "
            f"Bloques: {[(b[0], b[1]) for b in bloques]}",
            fila_headers, fila_datos, max_row, max_col,
            True,
            f"Separar en {len(bloques)} hojas/tablas distintas. "
            "Cada bloque puede ser una entidad diferente.",
            [], bloques=bloques
        )

    # ── PATRÓN: Totales mezclados ──────────────────────────────────────────
    filas_total = []
    for r in range(fila_datos, max_row + 1):
        v = ws.cell(r, 1).value or ws.cell(r, 2).value
        if v and _es_total(str(v)):
            filas_total.append(r)

    if filas_total:
        return _resultado(
            "con_totales", "alta",
            f"Tabla vertical con {len(filas_total)} fila(s) de totales mezcladas "
            f"(filas {filas_total}). Pueden contaminar importaciones.",
            fila_headers, fila_datos, max_row, max_col,
            False,
            f"Excluir filas de totales automáticamente al importar: {filas_total}.",
            [], filas_totales=filas_total
        )

    # ── PATRÓN: Tabla sparse (muchas celdas vacías) ────────────────────────
    # Calcular densidad solo en filas donde ya hay datos reales
    # (saltar filas de header en blanco)
    densidades = [_densidad_fila(ws, r, max_col)
                  for r in range(fila_datos, min(fila_datos + 15, max_row + 1))
                  if any(ws.cell(r, c).value is not None for c in range(1, max_col + 1))]
    densidad_promedio = sum(densidades) / len(densidades) if densidades else 1

    if densidad_promedio < 0.25:
        return _resultado(
            "sparse", "media",
            f"Tabla con muchas celdas vacías (densidad {densidad_promedio:.0%}). "
            "Puede ser datos incompletos o estructura irregular.",
            fila_headers, fila_datos, max_row, max_col,
            True,
            "Revisar manualmente qué columnas son relevantes. "
            "Considerar si los campos vacíos son opcionales o errores.",
        )

    # ── PATRÓN: Vertical estándar (el caso feliz) ──────────────────────────
    n_datos = max_row - fila_datos + 1
    return _resultado(
        "vertical", "alta",
        f"Tabla vertical estándar: {len([h for h in headers_f1 if h])} columnas, "
        f"~{n_datos} filas de datos. Lista para importar directamente.",
        fila_headers, fila_datos, max_row, max_col,
        False,
        "Importar directamente. No requiere transformación.",
    )


# ── Resultado estándar ────────────────────────────────────────────────────────
def _resultado(patron, confianza, descripcion, fila_headers, fila_datos,
               max_row, max_col, requiere_humano, transformacion,
               preview=None, bloques=None, filas_totales=None) -> dict:
    return {
        "patron":          patron,
        "confianza":       confianza,
        "descripcion":     descripcion,
        "fila_headers":    fila_headers,
        "fila_datos":      fila_datos,
        "ultima_fila":     max_row,
        "max_col":         max_col,
        "filas_totales":   filas_totales or [],
        "bloques":         bloques or [],
        "requiere_humano": requiere_humano,
        "transformacion":  transformacion,
        "preview_normal":  preview or [],
    }


# ── Previews ──────────────────────────────────────────────────────────────────
def _preview_formulario(ws, fila_ini: int, max_row: int) -> list:
    """Preview de formulario: lista de {campo: valor}."""
    resultado = {}
    for r in range(fila_ini, min(fila_ini + 20, max_row + 1)):
        k = _val(ws.cell(r, 1).value)
        v = _val(ws.cell(r, 2).value)
        if k and v:
            resultado[k] = v
    return [resultado]  # una sola fila


def _preview_unpivot(ws, fila_headers: int, fila_datos: int,
                     max_row: int, max_col: int) -> list:
    """Preview de unpivot: muestra cómo quedarían las primeras filas normalizadas."""
    filas = []
    entidad_col = 1
    # Primera columna = entidad (nombre, trabajador, etc.)
    entidad_header = _val(ws.cell(fila_headers, entidad_col).value) or "entidad"

    for r in range(fila_datos, min(fila_datos + 3, max_row + 1)):
        entidad = _val(ws.cell(r, entidad_col).value)
        if not entidad:
            continue
        for c in range(2, min(max_col + 1, 6)):  # solo primeras cols para preview
            periodo = _val(ws.cell(fila_headers, c).value)
            valor   = _val(ws.cell(r, c).value)
            if periodo and valor:
                filas.append({
                    entidad_header: entidad,
                    "periodo":      periodo,
                    "valor":        valor,
                })
    return filas[:6]


def _preview_multi_header(ws, fila_headers: int,
                           max_row: int, max_col: int) -> list:
    """Preview combinando las 2 filas de headers."""
    headers_combinados = []
    for c in range(1, max_col + 1):
        h1 = _val(ws.cell(fila_headers,     c).value) or ""
        h2 = _val(ws.cell(fila_headers + 1, c).value) or ""
        if h1 and h2:
            headers_combinados.append(f"{h1}_{h2}")
        elif h1:
            headers_combinados.append(h1)
        elif h2:
            headers_combinados.append(h2)
        else:
            headers_combinados.append(f"col_{c}")

    # Primera fila de datos
    fila_datos = fila_headers + 2
    if fila_datos <= max_row:
        vals = [_val(ws.cell(fila_datos, c).value) for c in range(1, max_col + 1)]
        return [dict(zip(headers_combinados, vals))]
    return []


def _detectar_bloques(ws, fila_ini: int, max_row: int, max_col: int) -> list:
    """Detecta bloques separados por filas vacías."""
    bloques = []
    en_bloque = False
    inicio = None

    for r in range(fila_ini, max_row + 2):
        fila_vacia = r > max_row or all(
            ws.cell(r, c).value is None for c in range(1, max_col + 1)
        )
        if not fila_vacia and not en_bloque:
            en_bloque = True
            inicio = r
        elif fila_vacia and en_bloque:
            en_bloque = False
            if r - inicio >= 2:  # al menos 2 filas para contar como bloque
                bloques.append((inicio, r - 1))

    return bloques


# ═══════════════════════════════════════════════════════════════════════════════
# NORMALIZADOR — transforma datos según el patrón detectado
# ═══════════════════════════════════════════════════════════════════════════════

def normalizar(ws, diagnostico: dict, columnas_cfg: dict = None) -> dict:
    """
    Aplica la transformación correspondiente al patrón detectado.
    Retorna datos normalizados listos para importar.

    {
      "filas":    [dict, ...],   # datos normalizados
      "columnas": [str, ...],    # nombres de columnas del resultado
      "excluidas": [int, ...],   # filas excluidas (totales, etc.)
      "patron":   str,
    }
    """
    patron = diagnostico["patron"]

    if patron == "vertical":
        return _norm_vertical(ws, diagnostico)
    elif patron == "horizontal":
        return _norm_unpivot(ws, diagnostico)
    elif patron == "multi_header":
        return _norm_multi_header(ws, diagnostico)
    elif patron == "formulario":
        return _norm_formulario(ws, diagnostico)
    elif patron == "con_totales":
        return _norm_excluir_totales(ws, diagnostico)
    elif patron == "multi_tabla":
        return _norm_multi_tabla(ws, diagnostico)
    else:
        return _norm_vertical(ws, diagnostico)  # fallback


def _norm_vertical(ws, d: dict) -> dict:
    """Normalizador estándar: tabla vertical limpia."""
    max_col = d["max_col"]
    fila_h  = d["fila_headers"]
    fila_d  = d["fila_datos"]
    max_row = d["ultima_fila"]

    columnas = [_val(ws.cell(fila_h, c).value) or f"col_{c}"
                for c in range(1, max_col + 1)]
    # Limpiar nombres de columnas
    columnas = [re.sub(r'\s+', '_', re.sub(r'[^\w\s]', '', str(c).lower()))
                for c in columnas]

    filas = []
    for r in range(fila_d, max_row + 1):
        vals = [_val(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        if any(v is not None for v in vals):
            filas.append(dict(zip(columnas, vals)))

    return {"filas": filas, "columnas": columnas, "excluidas": [], "patron": "vertical"}


def _norm_unpivot(ws, d: dict) -> dict:
    """Unpivot: tabla horizontal → vertical."""
    max_col  = d["max_col"]
    fila_h   = d["fila_headers"]
    fila_d   = d["fila_datos"]
    max_row  = d["ultima_fila"]

    # Primera columna = identificador de la entidad
    entidad_col = _val(ws.cell(fila_h, 1).value) or "entidad"
    entidad_col = re.sub(r'[^\w]', '_', str(entidad_col).lower())

    # El resto de columnas son períodos/valores
    periodos = [_val(ws.cell(fila_h, c).value)
                for c in range(2, max_col + 1)]

    filas = []
    for r in range(fila_d, max_row + 1):
        entidad = _val(ws.cell(r, 1).value)
        if not entidad:
            continue
        for i, periodo in enumerate(periodos):
            if periodo is None:
                continue
            valor = _val(ws.cell(r, i + 2).value)
            if valor is not None:
                filas.append({
                    entidad_col: entidad,
                    "periodo":   str(periodo),
                    "valor":     valor,
                })

    return {
        "filas":    filas,
        "columnas": [entidad_col, "periodo", "valor"],
        "excluidas": [],
        "patron":   "horizontal_unpivot",
    }


def _norm_multi_header(ws, d: dict) -> dict:
    """Multi-header: combinar 2 filas de headers."""
    max_col = d["max_col"]
    fila_h  = d["fila_headers"]
    max_row = d["ultima_fila"]

    columnas = []
    for c in range(1, max_col + 1):
        h1 = _val(ws.cell(fila_h,     c).value) or ""
        h2 = _val(ws.cell(fila_h + 1, c).value) or ""
        nombre = f"{h1}_{h2}" if h1 and h2 else (h1 or h2 or f"col_{c}")
        nombre = re.sub(r'[^\w]', '_', nombre.lower())
        columnas.append(nombre)

    fila_d = fila_h + 2
    filas  = []
    for r in range(fila_d, max_row + 1):
        vals = [_val(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        if any(v is not None for v in vals):
            filas.append(dict(zip(columnas, vals)))

    return {"filas": filas, "columnas": columnas, "excluidas": [], "patron": "multi_header"}


def _norm_formulario(ws, d: dict) -> dict:
    """Formulario vertical: convierte pares campo:valor en una sola fila."""
    fila_h  = d["fila_headers"]
    max_row = d["ultima_fila"]

    fila = {}
    for r in range(fila_h, max_row + 1):
        k = _val(ws.cell(r, 1).value)
        v = _val(ws.cell(r, 2).value)
        if k:
            k_norm = re.sub(r'[^\w]', '_', k.lower())
            fila[k_norm] = v

    return {
        "filas":    [fila],
        "columnas": list(fila.keys()),
        "excluidas": [],
        "patron":   "formulario",
    }


def _norm_excluir_totales(ws, d: dict) -> dict:
    """Tabla con totales: normaliza excluyendo filas de totales."""
    resultado = _norm_vertical(ws, d)
    filas_excluidas = set(d["filas_totales"])

    fila_d   = d["fila_datos"]
    filas_ok = []
    for i, fila in enumerate(resultado["filas"]):
        fila_real = fila_d + i
        if fila_real not in filas_excluidas:
            filas_ok.append(fila)

    resultado["filas"]    = filas_ok
    resultado["excluidas"] = d["filas_totales"]
    resultado["patron"]   = "con_totales_excluidos"
    return resultado


def _norm_multi_tabla(ws, d: dict) -> dict:
    """Múltiples tablas: retorna el primer bloque (el más grande)."""
    bloques = d.get("bloques", [])
    if not bloques:
        return _norm_vertical(ws, d)

    # Tomar el bloque más grande
    bloque = max(bloques, key=lambda b: b[1] - b[0])
    d_bloque = {**d, "fila_headers": bloque[0], "fila_datos": bloque[0] + 1,
                "ultima_fila": bloque[1]}
    resultado = _norm_vertical(ws, d_bloque)
    resultado["patron"] = "multi_tabla_primer_bloque"
    resultado["nota"]   = f"Se usó el bloque más grande ({bloque[0]}-{bloque[1]}). " \
                           f"Hay {len(bloques)} bloques en total."
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS COMPLETO DE UN EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def analizar_excel_completo(path: str) -> dict:
    """
    Analiza todas las hojas de un Excel local y retorna diagnóstico completo.
    Para Google Sheets usa analizar_sheets_completo().
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    resultado = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws.max_row or ws.max_row < 2:
            continue
        diag = detectar_patron(ws)
        resultado[sheet_name] = {**diag, "hoja": sheet_name}
    return resultado


def analizar_sheets_completo(gspread_doc) -> dict:
    """
    Analiza todas las hojas de un Google Sheets y retorna diagnóstico.
    Recibe el objeto documento de gspread (ya autenticado).

    Para Google Sheets no tenemos acceso directo a la estructura celda por celda
    de forma eficiente, así que usamos una heurística más simple basada en
    get_all_values() de cada hoja.
    """
    resultado = {}

    for ws in gspread_doc.worksheets():
        sheet_name = ws.title
        try:
            # Obtener los primeros 20 valores para análisis
            todos = ws.get_all_values()
            if not todos or len(todos) < 2:
                resultado[sheet_name] = _resultado(
                    "vacia", "baja", "Hoja vacía", 1, 2, 0, 0, False, "Ignorar"
                )
                continue

            max_row = len(todos)
            max_col = max(len(r) for r in todos) if todos else 0

            # Detectar headers en primera fila no vacía
            fila_headers = None
            for i, fila in enumerate(todos[:7]):
                headers_en_fila = sum(1 for v in fila if _es_header(v))
                if headers_en_fila >= 3:
                    fila_headers = i + 1
                    break

            if not fila_headers:
                fila_headers = 1

            headers_f1 = todos[fila_headers - 1] if fila_headers <= len(todos) else []

            # Detectar patrón basado en headers
            headers_son_fechas = sum(1 for h in headers_f1 if _es_fecha_o_mes(h))
            es_formulario = max_col <= 3 and sum(
                1 for fila in todos[fila_headers:fila_headers+10]
                if len(fila) >= 2 and _es_header(fila[0]) and fila[1]
            ) >= 3

            if es_formulario:
                patron = "formulario"
                req_humano = False
                desc = "Formulario campo:valor — representa una sola entidad"
                transform = "Convertir a tabla config con una sola fila"
            elif headers_son_fechas >= 3:
                patron = "horizontal"
                req_humano = True
                desc = f"Tabla horizontal: {headers_son_fechas} columnas son fechas/períodos"
                transform = "Unpivot: convertir columnas de períodos a filas"
            else:
                patron = "vertical"
                req_humano = False
                n_datos = max_row - fila_headers
                desc = f"Tabla vertical estándar: {len([h for h in headers_f1 if h])} columnas, ~{n_datos} filas"
                transform = "Importar directamente"

            resultado[sheet_name] = _resultado(
                patron, "media", desc,
                fila_headers, fila_headers + 1, max_row, max_col,
                req_humano, transform
            )
            resultado[sheet_name]["hoja"] = sheet_name

        except Exception as e:
            resultado[sheet_name] = _resultado(
                "error", "baja", f"Error analizando: {e}",
                1, 2, 0, 0, True, "Revisar manualmente"
            )

    return resultado


def analizar_fuente(sistema) -> dict:
    """
    Analiza automáticamente según el modo del sistema (local o sheets).
    Punto de entrada unificado para el classifier.
    """
    if sistema.modo == "sheets":
        return analizar_sheets_completo(sistema._db._doc)
    else:
        return analizar_excel_completo(sistema._db._path)


# ── Test ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

    path = "KraftDo_BD_Maestra_v5.xlsx"
    if not os.path.exists(path):
        print(f"Excel no encontrado: {path}")
        sys.exit(1)

    diagnosticos = analizar_excel_completo(path)
    for hoja, diag in diagnosticos.items():
        emoji = {
            "vertical":    "✅",
            "horizontal":  "↔️ ",
            "multi_header":"📊",
            "formulario":  "📝",
            "multi_tabla": "📋",
            "con_totales": "⚠️ ",
            "sparse":      "🕳️ ",
            "vacia":       "⬜",
        }.get(diag["patron"], "❓")

        print(f"\n{emoji} {hoja}")
        print(f"   Patrón:    {diag['patron']} [{diag['confianza']}]")
        print(f"   Diagnóstico: {diag['descripcion']}")
        print(f"   Acción:    {diag['transformacion']}")
        if diag["requiere_humano"]:
            print(f"   ⚠️  Requiere revisión manual")
        if diag.get("preview_normal"):
            print(f"   Preview:   {diag['preview_normal'][0]}")
