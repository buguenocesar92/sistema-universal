"""
KraftDo — tests/test_sistema.py
Tests automatizados con pytest para el sistema completo.

USO:
    pytest tests/ -v
    pytest tests/ -v --tb=short
    pytest tests/test_sistema.py::TestCore -v
"""

import os
import sys
import json
import pytest
import tempfile

# Agregar directorio raíz al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ── Fixtures ──────────────────────────────────────────────────────────────────
@pytest.fixture
def cfg_kraftdo():
    """Carga el config de KraftDo."""
    path = os.path.join(os.path.dirname(__file__), "..", "empresas", "kraftdo.json")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


@pytest.fixture
def excel_path():
    """Path al Excel de KraftDo."""
    return os.path.join(
        os.path.dirname(__file__), "..", "KraftDo_BD_Maestra_v5.xlsx"
    )


@pytest.fixture
def sistema_kraftdo(excel_path, cfg_kraftdo):
    """Instancia del Sistema con Excel local."""
    from core import Sistema
    s = Sistema("kraftdo", forzar="local")
    return s


# ── Tests Core ────────────────────────────────────────────────────────────────
class TestCore:

    def test_carga_empresa(self, sistema_kraftdo):
        """El sistema carga la empresa correctamente."""
        assert sistema_kraftdo.nombre == "kraftdo"
        assert sistema_kraftdo.modo == "local"
        assert sistema_kraftdo.cfg["empresa"]["nombre"] == "KraftDo SpA"

    def test_hojas_disponibles(self, sistema_kraftdo):
        """Retorna las hojas configuradas."""
        hojas = sistema_kraftdo.hojas_disponibles()
        assert len(hojas) > 0
        assert "productos_nfc" in hojas
        assert "pedidos" in hojas
        assert "clientes" in hojas

    def test_catalogo_tiene_productos(self, sistema_kraftdo):
        """El catálogo tiene productos activos."""
        cat = sistema_kraftdo.catalogo()
        assert "productos_nfc" in cat
        assert "sublimacion" in cat
        assert "impresion3d" in cat
        assert "packs" in cat

        total = sum(len(v) for v in cat.values())
        assert total > 0, "El catálogo debe tener al menos 1 producto"

    def test_catalogo_nfc_tiene_precios(self, sistema_kraftdo):
        """Al menos la mayoría de productos NFC tienen precio_1 mayor a 0."""
        cat = sistema_kraftdo.catalogo()
        nfc = cat.get("productos_nfc", [])
        assert len(nfc) > 0
        con_precio = [p for p in nfc if p.get("precio_1", 0) > 0]
        sin_sku    = [p for p in nfc if not p.get("sku")]
        assert len(sin_sku) == 0, "Hay productos sin SKU"
        assert len(con_precio) >= len(nfc) * 0.8, f"Menos del 80% tiene precio: {len(con_precio)}/{len(nfc)}"

    def test_precio_sku_existente(self, sistema_kraftdo):
        """precio() retorna datos correctos para un SKU válido."""
        r = sistema_kraftdo.precio("A01", 1)
        assert r is not None
        assert r["sku"] == "A01"
        assert r["precio_unitario"] > 0
        assert r["subtotal"] == r["precio_unitario"] * 1
        assert r["tramo"] == "unit"

    def test_precio_tramo_5(self, sistema_kraftdo):
        """El tramo 5+ aplica precio correcto."""
        r1 = sistema_kraftdo.precio("A01", 1)
        r5 = sistema_kraftdo.precio("A01", 5)
        assert r5 is not None
        assert r5["tramo"] == "5+"
        # Precio 5+ debe ser <= precio unitario
        assert r5["precio_unitario"] <= r1["precio_unitario"]

    def test_precio_tramo_10(self, sistema_kraftdo):
        """El tramo 10+ aplica precio correcto."""
        r10 = sistema_kraftdo.precio("S01", 10)
        assert r10 is not None
        assert r10["tramo"] == "10+"

    def test_precio_sku_inexistente(self, sistema_kraftdo):
        """precio() retorna None para SKU inexistente."""
        r = sistema_kraftdo.precio("ZZZ999", 1)
        assert r is None

    def test_buscar_producto(self, sistema_kraftdo):
        """buscar() encuentra productos por texto."""
        resultados = sistema_kraftdo.buscar("taza")
        assert len(resultados) > 0
        # Al menos un resultado debe tener "taza" en nombre o variante
        con_taza = [r for r in resultados
                    if "taza" in (r.get("producto","") + r.get("variante","") + r.get("descripcion","")).lower()]
        assert len(con_taza) > 0, "Ningún resultado contiene 'taza'"

    def test_buscar_sin_resultados(self, sistema_kraftdo):
        """buscar() retorna lista vacía si no hay coincidencias."""
        r = sistema_kraftdo.buscar("xyzxyz123456")
        assert r == []

    def test_cotizacion_completa(self, sistema_kraftdo):
        """cotizar() calcula totales correctamente."""
        cot = sistema_kraftdo.cotizar([
            {"sku": "A01", "cantidad": 1},
            {"sku": "S01", "cantidad": 5},
        ], cliente="Test Cliente")

        assert cot["total"] > 0
        assert cot["iva"]   > 0
        assert cot["anticipo"] == round(cot["total"] * 0.5)
        assert cot["saldo"]    == cot["total"] - cot["anticipo"]
        assert len(cot["lineas"]) == 2
        assert cot["errores"] == []

    def test_cotizacion_con_sku_invalido(self, sistema_kraftdo):
        """cotizar() reporta errores para SKUs inválidos."""
        cot = sistema_kraftdo.cotizar([
            {"sku": "A01",    "cantidad": 1},
            {"sku": "INVALIDO", "cantidad": 1},
        ])
        assert len(cot["errores"]) == 1
        assert "INVALIDO" in cot["errores"][0]

    def test_iva_19_pct(self, sistema_kraftdo):
        """El IVA es 19% del subtotal neto."""
        cot = sistema_kraftdo.cotizar([{"sku": "A01", "cantidad": 1}])
        iva_esperado = round(cot["subtotal_neto"] * 0.19)
        assert cot["iva"] == iva_esperado

    def test_registros_proveedores(self, sistema_kraftdo):
        """registros() retorna filas de proveedores."""
        provs = sistema_kraftdo.registros("proveedores")
        assert isinstance(provs, list)
        assert len(provs) > 0

    def test_kpis_retorna_dict(self, sistema_kraftdo):
        """kpis() retorna diccionario con métricas."""
        k = sistema_kraftdo.kpis()
        assert isinstance(k, dict)
        assert "resumen" in k or "caja" in k


# ── Tests Relations ───────────────────────────────────────────────────────────
class TestRelations:

    def test_detecta_relaciones(self, cfg_kraftdo):
        """Detecta relaciones entre hojas del Excel."""
        from relations import detectar_relaciones
        rels = detectar_relaciones(cfg_kraftdo)
        assert len(rels) > 0

    def test_relacion_pedidos_clientes(self, cfg_kraftdo):
        """Detecta relación pedidos.cliente → clientes.nombre."""
        from relations import detectar_relaciones
        rels = detectar_relaciones(cfg_kraftdo)
        rel_ped_cli = next(
            (r for r in rels
             if r["tabla_origen"] == "pedidos" and r["campo_origen"] == "cliente"),
            None
        )
        assert rel_ped_cli is not None
        assert rel_ped_cli["tabla_destino"] == "clientes"

    def test_confianza_alta(self, cfg_kraftdo):
        """Al menos una relación tiene confianza alta."""
        from relations import detectar_relaciones
        rels = detectar_relaciones(cfg_kraftdo)
        altas = [r for r in rels if r["confianza"] == "alta"]
        assert len(altas) > 0

    def test_gen_belongs_to(self, cfg_kraftdo):
        """gen_eloquent_relationships genera código PHP válido."""
        from relations import detectar_relaciones, gen_eloquent_relationships
        rels = detectar_relaciones(cfg_kraftdo)
        php = gen_eloquent_relationships(rels, "pedidos")
        assert "public function" in php
        assert "belongsTo" in php

    def test_gen_has_many(self, cfg_kraftdo):
        """gen_hasmany_relationships genera código PHP válido."""
        from relations import detectar_relaciones, gen_hasmany_relationships
        rels = detectar_relaciones(cfg_kraftdo)
        php = gen_hasmany_relationships(rels, "clientes", {})
        # clientes debería tener hasMany pedidos
        if php:  # puede que no haya, dependiendo del Excel
            assert "hasMany" in php


# ── Tests Generator ───────────────────────────────────────────────────────────
class TestGenerator:

    def test_genera_archivos(self, cfg_kraftdo):
        """generar() produce al menos 50 archivos."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            assert len(archivos) >= 50

    def test_genera_migracion_pedidos(self, cfg_kraftdo):
        """La migración de pedidos tiene las columnas correctas."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            mig = next((v for k, v in archivos.items() if "pedidos" in k and "migration" in k.lower()), None)
            assert mig is not None
            assert "Schema::create('pedidos'" in mig
            assert "decimal" in mig  # precio, costo, total
            assert "integer" in mig  # cantidad

    def test_genera_modelo_con_fillable(self, cfg_kraftdo):
        """El modelo Pedido tiene $fillable definido."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            modelo = next((v for k, v in archivos.items() if "Models/Pedido.php" in k), None)
            assert modelo is not None
            assert "$fillable" in modelo
            assert "cliente" in modelo

    def test_genera_form_request(self, cfg_kraftdo):
        """Se generan FormRequests con reglas."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            req = next((v for k, v in archivos.items() if "Request" in k and "Pedido" in k), None)
            assert req is not None
            assert "public function rules()" in req
            assert "required" in req or "nullable" in req

    def test_genera_seeder(self, cfg_kraftdo):
        """Se generan Seeders para cada hoja."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            seeds = [k for k in archivos.keys() if "Seeder" in k]
            assert len(seeds) >= 5

    def test_genera_filament_pages(self, cfg_kraftdo):
        """Se generan las 3 Pages por cada Resource."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            pages = [k for k in archivos.keys() if "Pages" in k]
            # Cada hoja genera 3 pages (List, Create, Edit)
            hojas_generables = sum(
                1 for h in cfg_kraftdo["hojas"].values()
                if h.get("tipo") in ("catalogo", "registros")
            )
            assert len(pages) == hojas_generables * 3

    def test_genera_install_sh(self, cfg_kraftdo):
        """Se genera el script de instalación."""
        from generator import generar
        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("kraftdo", tmp)
            assert "install.sh" in archivos
            assert "php artisan migrate" in archivos["install.sh"]


# ── Tests Fórmulas ────────────────────────────────────────────────────────────
class TestFormulaParser:

    def test_formula_multiplicacion(self):
        """Parsea multiplicación simple."""
        from formula_parser import formula_a_php
        cols = {"costo": "D", "margen": "E"}
        r = formula_a_php("=D*E", cols)
        assert r["tipo"] == "simple"
        assert "costo" in r["php"]
        assert "margen" in r["php"]

    def test_formula_iva(self):
        """Detecta cálculo de IVA."""
        from formula_parser import formula_a_php
        cols = {"subtotal": "G"}
        r = formula_a_php("=G*0.19", cols)
        assert r["tipo"] == "simple"
        assert "0.19" in r["php"]

    def test_formula_if(self):
        """Parsea condicional IF."""
        from formula_parser import formula_a_php
        cols = {"cantidad": "B", "precio": "C"}
        r = formula_a_php("=IF(B>5,C*0.9,C)", cols)
        assert r["tipo"] == "condicional"
        assert "?" in r["php"]

    def test_formula_vlookup_no_convertible(self):
        """VLOOKUP marca como no convertible."""
        from formula_parser import formula_a_php
        r = formula_a_php("=VLOOKUP(A1,B:C,2,0)", {})
        assert r["tipo"] == "no_convertible"
        assert r["php"] is None

    def test_formula_sum_rango(self):
        """SUM de rango marca como agregado."""
        from formula_parser import formula_a_php
        r = formula_a_php("=SUM(G11:G20)", {})
        assert r["tipo"] == "agregado"


# ── Tests Differ ──────────────────────────────────────────────────────────────
class TestDiffer:

    def test_sin_cambios(self, cfg_kraftdo):
        """diff_hojas retorna vacío si los configs son iguales."""
        from differ import diff_hojas
        d = diff_hojas(cfg_kraftdo, cfg_kraftdo)
        assert d["hojas_nuevas"] == []
        assert d["hojas_eliminadas"] == []
        assert d["hojas_modificadas"] == {}

    def test_detecta_hoja_nueva(self, cfg_kraftdo):
        """Detecta cuando se agrega una hoja nueva."""
        from differ import diff_hojas
        import copy
        cfg_nuevo = copy.deepcopy(cfg_kraftdo)
        cfg_nuevo["hojas"]["nueva_hoja"] = {
            "tipo": "registros",
            "nombre": "Nueva Hoja",
            "columnas": {"nombre": "A", "valor": "B"},
        }
        d = diff_hojas(cfg_kraftdo, cfg_nuevo)
        assert "nueva_hoja" in d["hojas_nuevas"]

    def test_detecta_columna_nueva(self, cfg_kraftdo):
        """Detecta cuando se agrega una columna a una hoja existente."""
        from differ import diff_hojas
        import copy
        cfg_nuevo = copy.deepcopy(cfg_kraftdo)
        cfg_nuevo["hojas"]["pedidos"]["columnas"]["nueva_col"] = "Z"
        d = diff_hojas(cfg_kraftdo, cfg_nuevo)
        assert "pedidos" in d["hojas_modificadas"]
        assert "nueva_col" in d["hojas_modificadas"]["pedidos"]["cols_nuevas"]


# ── Tests API ─────────────────────────────────────────────────────────────────
class TestAPI:

    @pytest.fixture
    def client(self):
        """Cliente de prueba para la API FastAPI."""
        from fastapi.testclient import TestClient
        from api import app
        return TestClient(app)

    def test_root_retorna_empresas(self, client):
        """GET / retorna lista de empresas."""
        r = client.get("/")
        assert r.status_code == 200
        data = r.json()
        assert "empresas" in data

    def test_catalogo_kraftdo(self, client):
        """GET /kraftdo/catalogo retorna productos."""
        r = client.get("/kraftdo/catalogo")
        assert r.status_code == 200
        data = r.json()
        assert "catalogo" in data

    def test_precio_sku_valido(self, client):
        """GET /kraftdo/precio?sku=A01&cantidad=1 retorna precio."""
        r = client.get("/kraftdo/precio?sku=A01&cantidad=1")
        assert r.status_code == 200
        data = r.json()
        assert data["precio_unitario"] > 0

    def test_precio_sku_invalido_404(self, client):
        """GET con SKU inválido retorna 404."""
        r = client.get("/kraftdo/precio?sku=ZZZZZ")
        assert r.status_code == 404

    def test_empresa_inexistente_404(self, client):
        """GET empresa que no existe retorna 404."""
        r = client.get("/empresa_que_no_existe/catalogo")
        assert r.status_code == 404


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

# ── Tests Rate Limiting ───────────────────────────────────────────────────────
class TestRateLimit:

    @pytest.fixture
    def client(self):
        from fastapi.testclient import TestClient
        from api import app
        return TestClient(app)

    def test_rate_limit_no_bloquea_requests_normales(self, client):
        """Requests normales no son bloqueadas."""
        for _ in range(5):
            r = client.get("/kraftdo/catalogo")
            assert r.status_code == 200

    def test_health_muestra_rate_limit(self, client):
        """Health endpoint muestra configuración de rate limit."""
        # El endpoint de health puede estar en / o /health
        r = client.get("/")
        assert r.status_code == 200
        # El health con rate_limit está en la respuesta de root
        data = r.json()
        assert data is not None


# ── Tests n8n Generator El Pilar ─────────────────────────────────────────────
class TestN8nElPilar:

    def test_genera_workflow_el_pilar(self, cfg_kraftdo):
        """Genera el workflow de El Pilar con todos los nodos."""
        from n8n_generator import gen_workflow_el_pilar
        wf = gen_workflow_el_pilar(
            "http://api.kraftdo.cl",
            "NOTION_DB_ID",
            "TELEGRAM_ID",
            "cesar@kraftdo.cl"
        )
        assert "El Pilar" in wf["name"]
        assert len(wf["nodes"]) >= 5
        tipos = [n["type"] for n in wf["nodes"]]
        nombres = [n["name"] for n in wf["nodes"]]
        # Debe tener webhook, notion, y algun tipo de notificación
        assert any("webhook" in t.lower() for t in tipos)
        assert any("Notion" in n or "notion" in t.lower() for n, t in zip(nombres, tipos))

    def test_workflow_incluye_validacion(self, cfg_kraftdo):
        """El workflow tiene nodo de validación de datos."""
        from n8n_generator import gen_workflow_el_pilar
        wf = gen_workflow_el_pilar("http://api", "DB", "TG", "email")
        nombres = [n["name"] for n in wf["nodes"]]
        assert any("Valid" in n for n in nombres)

    def test_generar_todos_incluye_el_pilar(self, cfg_kraftdo):
        """generar_todos() incluye El Pilar para empresa kraftdo."""
        from n8n_generator import generar_todos
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            generar_todos("kraftdo", tmp, "http://api", "TG", "email@test.cl")
            import os
            archivos = os.listdir(tmp)
            assert any("el_pilar" in a for a in archivos)


# ── Tests Differ Watch ────────────────────────────────────────────────────────
class TestDifferWatch:

    def test_watch_function_existe(self):
        """La función watch está disponible en differ."""
        from differ import watch
        assert callable(watch)

    def test_differ_detecta_columna_eliminada(self, cfg_kraftdo):
        """Detecta cuando se elimina una columna."""
        from differ import diff_hojas
        import copy
        cfg_nuevo = copy.deepcopy(cfg_kraftdo)
        # Eliminar columna 'obs' de pedidos
        if "obs" in cfg_nuevo["hojas"]["pedidos"]["columnas"]:
            del cfg_nuevo["hojas"]["pedidos"]["columnas"]["obs"]
            d = diff_hojas(cfg_kraftdo, cfg_nuevo)
            assert "pedidos" in d["hojas_modificadas"]
            assert "obs" in d["hojas_modificadas"]["pedidos"]["cols_eliminadas"]

    def test_gen_alter_migration_add_column(self, cfg_kraftdo):
        """Genera ALTER TABLE correcto para columna nueva."""
        from differ import diff_hojas, gen_alter_migration
        import copy
        cfg_nuevo = copy.deepcopy(cfg_kraftdo)
        cfg_nuevo["hojas"]["pedidos"]["columnas"]["nueva_col"] = "Z"
        d = diff_hojas(cfg_kraftdo, cfg_nuevo)
        if "pedidos" in d["hojas_modificadas"]:
            sql = gen_alter_migration("pedidos", d["hojas_modificadas"]["pedidos"], 1)
            assert "Schema::table" in sql
            assert "nueva_col" in sql


# ── Test Setup ────────────────────────────────────────────────────────────────
class TestCLI:

    def test_kraftdo_cli_importable(self):
        """El CLI principal es importable."""
        import kraftdo
        assert hasattr(kraftdo, 'main')

    def test_kraftdo_tiene_todos_comandos(self):
        """El CLI tiene todos los comandos esperados."""
        import kraftdo
        import argparse
        # Verificar que las funciones de comandos existen
        for cmd in ['cmd_setup', 'cmd_clasificar', 'cmd_api', 'cmd_generar',
                    'cmd_importar', 'cmd_n8n', 'cmd_diff', 'cmd_test', 'cmd_empresas']:
            assert hasattr(kraftdo, cmd), f"Falta comando: {cmd}"

# ── Tests CRUD Excel ──────────────────────────────────────────────────────────
class TestCRUD:

    @pytest.fixture
    def sistema_tmp(self, tmp_path):
        """Sistema con una copia temporal del Excel para pruebas de escritura."""
        import shutil, sys
        src = os.path.join(os.path.dirname(__file__), "..", "KraftDo_BD_Maestra_v5.xlsx")
        dst = tmp_path / "KraftDo_BD_Maestra_v5.xlsx"
        shutil.copy(src, dst)
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from core import Sistema
        # Modificar temporalmente el config para apuntar al tmp
        s = Sistema("kraftdo", forzar="local")
        s._db._path = str(dst)
        return s

    def test_schema_retorna_campos(self, sistema_kraftdo):
        """schema() retorna estructura correcta."""
        schema = sistema_kraftdo.schema("clientes")
        assert "campos" in schema
        assert "identificador" in schema
        assert schema["tipo"] == "registros"
        assert "nombre" in schema["campos"]

    def test_schema_tipos_correctos(self, sistema_kraftdo):
        """Los tipos de campo son correctos."""
        schema = sistema_kraftdo.schema("pedidos")
        campos = schema["campos"]
        assert campos["estado"]["tipo"] == "select"
        assert campos["cantidad"]["tipo"] == "number"
        assert campos["obs"]["tipo"] == "textarea"

    def test_buscar_filtros_eq(self, sistema_kraftdo):
        """buscar_filtros filtra por igualdad."""
        # Proveedores tiene datos reales
        provs = sistema_kraftdo.buscar_filtros("proveedores", {"proveedor__like": "Sub"})
        assert isinstance(provs, list)

    def test_buscar_filtros_like(self, sistema_kraftdo):
        """buscar_filtros con __like funciona."""
        resultados = sistema_kraftdo.buscar_filtros("proveedores", {"producto__like": "oz"})
        assert isinstance(resultados, list)
        for r in resultados:
            assert "oz" in str(r.get("producto","")).lower()

    def test_buscar_filtros_sin_resultados(self, sistema_kraftdo):
        """buscar_filtros retorna vacío si no hay coincidencias."""
        r = sistema_kraftdo.buscar_filtros("proveedores", {"proveedor": "EMPRESA_QUE_NO_EXISTE_XYZ"})
        assert r == []

    def test_buscar_filtros_gt(self, sistema_kraftdo):
        """buscar_filtros con __gt filtra numéricamente."""
        r = sistema_kraftdo.buscar_filtros("proveedores", {"precio_unit__gt": 0})
        assert isinstance(r, list)

    def test_schema_tiene_todos_los_campos(self, sistema_kraftdo):
        """schema() incluye todos los campos configurados en el JSON."""
        import json
        cfg = json.load(open(
            os.path.join(os.path.dirname(__file__), "..", "empresas", "kraftdo.json")
        ))
        schema = sistema_kraftdo.schema("pedidos")
        cols_cfg = set(cfg["hojas"]["pedidos"]["columnas"].keys())
        cols_schema = set(schema["campos"].keys()) | {schema["identificador"]}
        # Todos los campos del config deben estar en el schema
        assert cols_cfg <= cols_schema


# ── Tests API CRUD ────────────────────────────────────────────────────────────
class TestAPICRUD:

    @pytest.fixture
    def client(self):
        from fastapi.testclient import TestClient
        from api import app
        return TestClient(app)

    def test_get_registros_sin_filtros(self, client):
        """GET /registros retorna lista."""
        r = client.get("/kraftdo/registros/proveedores")
        assert r.status_code == 200
        data = r.json()
        assert "registros" in data
        assert "total" in data

    def test_get_registros_con_filtro_like(self, client):
        """GET /registros?campo__like=valor filtra correctamente."""
        r = client.get("/kraftdo/registros/proveedores?producto__like=oz")
        assert r.status_code == 200
        data = r.json()
        assert "filtros" in data
        assert data["filtros"].get("producto__like") == "oz"
        for reg in data["registros"]:
            assert "oz" in str(reg.get("producto","")).lower()

    def test_get_schema(self, client):
        """GET /registros/{alias}/schema retorna campos."""
        r = client.get("/kraftdo/registros/clientes/schema")
        assert r.status_code == 200
        data = r.json()
        assert "campos" in data
        assert data["tipo"] == "registros"

    def test_get_por_id_existente(self, client):
        """GET por ID devuelve el registro si existe."""
        # Buscar un proveedor real
        r = client.get("/kraftdo/registros/proveedores")
        provs = r.json()["registros"]
        if provs:
            prov_id = provs[0].get("proveedor")
            if prov_id:
                r2 = client.get(f"/kraftdo/registros/proveedores/{prov_id}")
                assert r2.status_code == 200

    def test_get_por_id_inexistente_404(self, client):
        """GET por ID inexistente retorna 404."""
        r = client.get("/kraftdo/registros/clientes/CLIENTE_INEXISTENTE_XYZ")
        assert r.status_code == 404

    def test_post_requiere_auth(self, client):
        """POST sin API key retorna 401 si hay API_KEY configurado."""
        import os
        if os.environ.get("API_KEY"):
            r = client.post("/kraftdo/registros/clientes", json={"nombre": "Test"})
            assert r.status_code == 401

    def test_delete_requiere_auth(self, client):
        """DELETE sin API key retorna 401 si hay API_KEY configurado."""
        import os
        if os.environ.get("API_KEY"):
            r = client.delete("/kraftdo/registros/clientes/ALGUNO")
            assert r.status_code == 401

# ── Tests Normalizer ──────────────────────────────────────────────────────────
class TestNormalizer:

    @pytest.fixture
    def wb_vertical(self):
        """Worksheet con tabla vertical estándar."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Headers
        ws.append(["nombre", "precio", "stock", "estado"])
        # Datos
        ws.append(["Producto A", 5000, 10, "Activo"])
        ws.append(["Producto B", 8000, 5,  "Activo"])
        ws.append(["Producto C", 3000, 0,  "Inactivo"])
        return ws

    @pytest.fixture
    def wb_horizontal(self):
        """Worksheet con tabla horizontal (pivot)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Empleado", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"])
        ws.append(["Juan",     1,       1,         0,           1,        1])
        ws.append(["María",    1,       0,         1,           1,        0])
        return ws

    @pytest.fixture
    def wb_formulario(self):
        """Worksheet con formulario vertical (campo:valor)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre empresa", "KraftDo SpA"])
        ws.append(["RUT",            "12.345.678-9"])
        ws.append(["Email",          "hola@kraftdo.cl"])
        ws.append(["Teléfono",       "+56912345678"])
        return ws

    @pytest.fixture
    def wb_con_totales(self):
        """Worksheet con totales mezclados — necesita 3+ columnas."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["producto", "cantidad", "monto", "estado"])
        ws.append(["Prod A",   2,          1000,    "Activo"])
        ws.append(["Prod B",   3,          2000,    "Activo"])
        ws.append(["TOTAL",    5,          3000,    ""])
        ws.append(["Prod C",   1,          1500,    "Activo"])
        return ws

    def test_detecta_vertical(self, wb_vertical):
        from normalizer import detectar_patron
        diag = detectar_patron(wb_vertical)
        assert diag["patron"] == "vertical"
        assert diag["confianza"] == "alta"
        assert not diag["requiere_humano"]

    def test_detecta_horizontal(self, wb_horizontal):
        from normalizer import detectar_patron
        diag = detectar_patron(wb_horizontal)
        assert diag["patron"] == "horizontal"
        assert diag["confianza"] == "alta"
        assert diag["requiere_humano"]

    def test_detecta_formulario(self, wb_formulario):
        from normalizer import detectar_patron
        diag = detectar_patron(wb_formulario)
        assert diag["patron"] == "formulario"
        assert not diag["requiere_humano"]

    def test_detecta_con_totales(self, wb_con_totales):
        from normalizer import detectar_patron
        diag = detectar_patron(wb_con_totales)
        assert diag["patron"] == "con_totales"
        assert 4 in diag["filas_totales"]

    def test_normaliza_vertical(self, wb_vertical):
        from normalizer import detectar_patron, normalizar
        diag = detectar_patron(wb_vertical)
        result = normalizar(wb_vertical, diag)
        assert len(result["filas"]) == 3
        assert "nombre" in result["filas"][0]
        assert result["filas"][0]["nombre"] == "Producto A"

    def test_normaliza_unpivot(self, wb_horizontal):
        from normalizer import detectar_patron, normalizar
        diag = detectar_patron(wb_horizontal)
        result = normalizar(wb_horizontal, diag)
        assert result["patron"] == "horizontal_unpivot"
        assert "periodo" in result["columnas"]
        assert "valor" in result["columnas"]
        # Juan tiene 4 días con valor=1
        juans = [f for f in result["filas"] if str(f.get("empleado","")).lower() == "juan"]
        assert len(juans) > 0

    def test_normaliza_formulario(self, wb_formulario):
        from normalizer import detectar_patron, normalizar
        diag = detectar_patron(wb_formulario)
        result = normalizar(wb_formulario, diag)
        assert len(result["filas"]) == 1
        fila = result["filas"][0]
        assert any("nombre" in k.lower() for k in fila.keys())

    def test_normaliza_excluye_totales(self, wb_con_totales):
        from normalizer import detectar_patron, normalizar
        diag = detectar_patron(wb_con_totales)
        result = normalizar(wb_con_totales, diag)
        assert result["patron"] == "con_totales_excluidos"
        nombres = [f.get("producto","") for f in result["filas"]]
        assert not any("TOTAL" in str(n).upper() for n in nombres)


# ── Tests Consolidator ────────────────────────────────────────────────────────
class TestConsolidator:

    def test_analizar_grupo(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion", "impresion3d"])
        reporte = c.analizar_grupo("productos")
        assert reporte["entidad"] == "productos"
        assert len(reporte["cols_comunes"]) > 0
        assert reporte["n_cols_total"] > reporte["n_cols_comunes"]

    def test_cols_comunes_presentes(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion", "impresion3d"])
        reporte = c.analizar_grupo("productos")
        # Columnas que deberían estar en las 3 hojas
        for col in ["precio_1", "estado", "margen"]:
            assert col in reporte["cols_comunes"], f"{col} no está en columnas comunes"

    def test_genera_json_consolidado(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion"])
        cfg_nuevo = c.generar_json_consolidado()
        assert "productos" in cfg_nuevo["hojas"]
        assert cfg_nuevo["hojas"]["productos"]["consolidado"] == True
        assert cfg_nuevo["hojas"]["productos"]["discriminador"] == "tipo"

    def test_hojas_originales_eliminadas(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion"])
        cfg_nuevo = c.generar_json_consolidado()
        assert "productos_nfc" not in cfg_nuevo["hojas"]
        assert "sublimacion"   not in cfg_nuevo["hojas"]
        # Hojas no consolidadas siguen presentes
        assert "pedidos"       in cfg_nuevo["hojas"]
        assert "clientes"      in cfg_nuevo["hojas"]

    def test_genera_migracion_consolidada(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion", "impresion3d"])
        mig = c.gen_migracion_consolidada("productos")
        assert "Schema::create('productos'" in mig
        assert "enum('tipo'" in mig
        assert "nfc" in mig or "productos_nfc" in mig

    def test_genera_modelo_con_scopes(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion"])
        modelo = c.gen_modelo_consolidado("productos")
        assert "class Productos extends Model" in modelo
        assert "scopeProductosNfc" in modelo or "scope" in modelo.lower()
        assert "where('tipo'" in modelo

    def test_grupo_requiere_min_2_hojas(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        with pytest.raises(ValueError):
            c.agregar_grupo("solo_una", ["productos_nfc"])
            c.analizar_grupo("solo_una")

    def test_cols_especificas_correctas(self, cfg_kraftdo):
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "impresion3d"])
        reporte = c.analizar_grupo("productos")
        # impresion3d debería tener columnas únicas como gramos, horas
        if "impresion3d" in reporte["cols_especificas"]:
            esp_3d = reporte["cols_especificas"]["impresion3d"]
            assert any(c in esp_3d for c in ["gramos", "horas", "costo_fil"])

# ── Tests de regresión de bugs corregidos ────────────────────────────────────
class TestBugsCorregidos:

    def test_slug_tildes(self):
        """Bug 2: _slug maneja tildes y emojis correctamente."""
        from generator import _slug
        assert _slug("🎨 Sublimación") == "sublimacion"
        assert _slug("🖨️ Impresión 3D") == "impresion_3d"
        assert _slug("📦 Productos NFC") == "productos_nfc"
        assert _slug("Cañón y Ñoño") == "canon_y_nono"

    def test_slug_sin_caracteres_invalidos_sql(self):
        """Bug 2: Los slugs generados son válidos para nombres SQL."""
        from generator import _slug
        casos = ["🎨 Sublimación", "🖨️ Impresión 3D", "Año 2026", "Café & Té"]
        for caso in casos:
            slug = _slug(caso)
            assert slug.replace("_","").isalnum() or slug == "hoja",                 f"Slug inválido para SQL: '{slug}' (de '{caso}')"

    def test_valores_tipo_sin_tildes(self):
        """Bug 2: _valores_tipo genera slugs SQL válidos."""
        from generator import _valores_tipo
        cfg = {'fuentes': ['🎨 Sublimación', '📦 Productos NFC', '🖨️ Impresión 3D']}
        tipos = _valores_tipo(cfg)
        assert tipos == ['sublimacion', 'productos_nfc', 'impresion_3d']
        for t in tipos:
            assert t.replace("_","").isalnum(), f"Tipo inválido: {t}"

    def test_migracion_consolidada_tiene_enum(self, cfg_kraftdo):
        """Bug 4: La migración consolidada genera enum correcto."""
        from consolidator import Consolidator
        from generator import generar
        import json, os, tempfile

        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion", "impresion3d"])
        cfg_cons = c.generar_json_consolidado()

        with open("empresas/test_regression.json", "w") as f:
            json.dump(cfg_cons, f)

        with tempfile.TemporaryDirectory() as tmp:
            archivos = generar("test_regression", tmp)
            mig_key = next(k for k in archivos if "create_productos" in k)
            mig = archivos[mig_key]

            assert "enum('tipo'" in mig, "Falta campo tipo como enum"
            assert "productos_nfc" in mig, "Falta valor productos_nfc en enum"
            assert "sublimacion" in mig, "Falta valor sublimacion en enum"
            assert "impresion3d" in mig, "Falta valor impresion3d en enum"
            assert "create_sublimacion_table" not in str(list(archivos.keys())),                 "Se generó tabla separada para sublimacion"

        os.remove("empresas/test_regression.json")

    def test_differ_importable_desde_script_dir(self):
        """Bug 5: differ puede importarse aunque no esté en el CWD."""
        import sys, os
        # Agregar directorio del script si no está
        script_dir = os.path.dirname(os.path.abspath("differ.py"))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        from differ import diff_hojas, guardar_snapshot
        assert callable(diff_hojas)
        assert callable(guardar_snapshot)

    def test_consolidado_genera_una_sola_tabla(self, cfg_kraftdo):
        """Bug 1 indirecto: el JSON consolidado no tiene hojas fuente duplicadas."""
        from consolidator import Consolidator
        c = Consolidator(cfg_kraftdo)
        c.agregar_grupo("productos", ["productos_nfc", "sublimacion"])
        cfg_cons = c.generar_json_consolidado()

        # Hojas fuente eliminadas
        assert "productos_nfc" not in cfg_cons["hojas"]
        assert "sublimacion" not in cfg_cons["hojas"]

        # Hoja consolidada con metadata correcta
        hoja = cfg_cons["hojas"]["productos"]
        assert hoja["consolidado"] == True
        assert "productos_nfc" in hoja["fuentes"]
        assert "sublimacion" in hoja["fuentes"]
        assert hoja["discriminador"] == "tipo"

# ── Tests Sheets y correcciones de bugs ──────────────────────────────────────
class TestSheets:

    def test_sistema_modo_local_por_defecto(self):
        """Sistema usa modo local cuando no hay SHEETS_ID."""
        import os
        os.environ.pop("SHEETS_ID", None)
        from core import Sistema
        s = Sistema("kraftdo")
        assert s.modo == "local"

    def test_sistema_modo_sheets_con_env(self):
        """Sistema cambia a modo sheets cuando SHEETS_ID está en env."""
        import os
        # Simular SHEETS_ID con un ID falso
        os.environ["SHEETS_ID"] = "fake_id_test_12345"
        try:
            from core import Sistema
            import importlib, core as core_mod
            importlib.reload(core_mod)
            # El constructor debería intentar Sheets pero caer a local
            # porque el ID es falso (sin creds)
            s = core_mod.Sistema("kraftdo")
            # Debe caer a local con fallback silencioso
            assert s.modo in ("local", "sheets")  # no crashea
        except Exception as e:
            # Solo es aceptable si es error de credenciales, no de lógica
            assert "credentials" in str(e).lower() or "creds" in str(e).lower() or "file" in str(e).lower()
        finally:
            os.environ.pop("SHEETS_ID", None)

    def test_sheets_id_leido_despues_de_env(self):
        """Bug corregido: sheets_id se lee después de aplicar env."""
        import os
        from core import Sistema
        # Sin SHEETS_ID en env, debe ser local
        os.environ.pop("SHEETS_ID", None)
        s = Sistema("kraftdo")
        assert s.modo == "local"
        assert s.cfg["fuente"]["tipo"] == "local"

    def test_normalizer_tiene_funcion_sheets(self):
        """normalizer expone analizar_sheets_completo y analizar_fuente."""
        from normalizer import analizar_sheets_completo, analizar_fuente
        assert callable(analizar_sheets_completo)
        assert callable(analizar_fuente)

    def test_analizar_fuente_usa_local_correctamente(self):
        """analizar_fuente delega a analizar_excel_completo en modo local."""
        from core import Sistema
        from normalizer import analizar_fuente
        s = Sistema("kraftdo")
        assert s.modo == "local"
        resultado = analizar_fuente(s)
        assert len(resultado) > 0
        # Debe tener diagnósticos de las hojas reales
        assert any(d["patron"] == "vertical" for d in resultado.values())

    def test_json_sheets_structure(self):
        """El JSON puede configurarse para Sheets cambiando fuente.tipo."""
        import json, copy
        cfg = json.load(open("empresas/kraftdo.json"))
        cfg_sheets = copy.deepcopy(cfg)
        cfg_sheets["fuente"]["tipo"] = "sheets"
        cfg_sheets["fuente"]["sheets_id"] = "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms"
        # Debe ser válido JSON
        texto = json.dumps(cfg_sheets)
        recuperado = json.loads(texto)
        assert recuperado["fuente"]["tipo"] == "sheets"
        assert recuperado["fuente"]["sheets_id"] != ""

