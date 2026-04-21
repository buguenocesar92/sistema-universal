"""
KraftDo — core.py v9
Lector UNIVERSAL dirigido por JSON.
Lee CUALQUIER Excel o Google Sheet cuya estructura esté descrita en un archivo .json.
No requiere modificar código al agregar hojas, columnas o nuevas empresas.

USO:
    from core import Sistema
    s = Sistema("kraftdo")          # carga empresas/kraftdo.json
    s = Sistema("adille")           # carga empresas/adille.json
    cat = s.catalogo()              # todos los productos activos
    precio = s.precio("A01", 3)     # precio para SKU + cantidad
    filas = s.registros("clientes") # todas las filas de una hoja
    kpis  = s.kpis()                # métricas del dashboard
"""

import os
import json
from datetime import datetime, date

# ── Helpers de formulario ────────────────────────────────────────────────────
def _tipo_form(tipo_laravel: str, campo: str, estados: list) -> str:
    """Convierte tipo Laravel a tipo de campo de formulario HTML."""
    n = campo.lower()
    if campo == "estado" and estados:      return "select"
    if "fecha" in n or "_at" in n:         return "date"
    if tipo_laravel == "text":             return "textarea"
    if tipo_laravel == "boolean":          return "checkbox"
    if tipo_laravel.startswith("decimal"): return "number"
    if tipo_laravel == "integer":          return "number"
    if "correo" in n or "email" in n:      return "email"
    if "telefono" in n or "whatsapp" in n: return "tel"
    return "text"


# ── Letra de columna a índice (A=0, B=1, …, Z=25, AA=26) ───────────────────
def _col_idx(letra: str) -> int:
    letra = letra.upper().strip()
    idx = 0
    for c in letra:
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1

# ── Valor limpio de celda ────────────────────────────────────────────────────
def _limpio(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y") if isinstance(v, date) else str(v)
    if isinstance(v, str):
        return v.strip() or None
    return v

def _num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace("$", "").strip()
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        elif s.count(".") > 1:
            s = s.replace(".", "")
        return float(s)
    except Exception:
        return 0

def _str(v):
    r = _limpio(v)
    return str(r) if r is not None else ""


# ══════════════════════════════════════════════════════════════════════════════
# ADAPTADORES DE FUENTE
# ══════════════════════════════════════════════════════════════════════════════

class AdaptadorLocal:
    """Lee desde archivo .xlsx / .xlsm usando Pandas (motor: openpyxl).
    
    Ventajas sobre openpyxl puro:
    - Manejo eficiente de archivos grandes (100K+ filas)
    - Tipos de dato correctos automáticamente (fechas, decimales)
    - Filtrado y búsqueda mucho más rápidos via DuckDB
    """

    def __init__(self, cfg_fuente: dict, base_dir: str):
        nombre = cfg_fuente["archivo"]
        rutas = [
            os.path.join(base_dir, nombre),
            nombre,
            os.path.join(os.path.expanduser("~"), "Documents", "KraftDo", nombre),
        ]
        path = next((r for r in rutas if os.path.exists(r)), None)
        if not path:
            raise FileNotFoundError(
                f"Excel '{nombre}' no encontrado.\n"
                f"Rutas buscadas: {rutas}"
            )
        self._path  = path
        self._cache_df: dict = {}  # cache de DataFrames por hoja

    def _nombre_hoja_real(self, nombre: str) -> str:
        """Encuentra el nombre real de la hoja (case-insensitive)."""
        import openpyxl
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        for h in wb.sheetnames:
            if nombre.lower() in h.lower():
                wb.close()
                return h
        wb.close()
        raise KeyError(f"Hoja '{nombre}' no encontrada en el Excel.")

    def _df(self, nombre_hoja: str, fila_headers: int = 1) -> "pd.DataFrame":
        """Carga la hoja como DataFrame con cache. Invalida cache si el archivo cambia."""
        import pandas as pd
        cache_key = f"{nombre_hoja}:{fila_headers}"
        mtime = os.path.getmtime(self._path)

        if cache_key not in self._cache_df or self._cache_df[cache_key][0] != mtime:
            nombre_real = self._nombre_hoja_real(nombre_hoja)
            df = pd.read_excel(
                self._path,
                sheet_name=nombre_real,
                header=fila_headers - 1,  # pandas usa 0-based
                engine="openpyxl",
                dtype=str,         # leer todo como string primero
            )
            # Limpiar columnas sin nombre (Unnamed: X)
            df = df.loc[:, ~df.columns.str.startswith("Unnamed:")]
            # Eliminar filas completamente vacías
            df = df.dropna(how="all")
            self._cache_df[cache_key] = (mtime, df)

        return self._cache_df[cache_key][1]

    def filas(self, nombre_hoja: str, fila_inicio: int) -> list[list]:
        """
        Retorna lista de listas de celdas desde fila_inicio (igual que openpyxl).
        fila_inicio es 1-based y apunta a la PRIMERA FILA DE DATOS.
        No hace header detection — eso lo maneja _leer_filas() del Sistema.
        """
        import pandas as pd
        try:
            nombre_real = self._nombre_hoja_real(nombre_hoja)
            mtime       = os.path.getmtime(self._path)
            cache_key   = f"raw:{nombre_hoja}:{fila_inicio}"

            if cache_key not in self._cache_df or self._cache_df[cache_key][0] != mtime:
                # Leer sin header, saltar filas antes de fila_inicio
                df = pd.read_excel(
                    self._path,
                    sheet_name=nombre_real,
                    header=None,           # sin header — datos crudos
                    skiprows=fila_inicio - 1,  # saltar hasta la primera fila de datos
                    engine="openpyxl",
                    dtype=object,          # preservar tipos originales
                )
                self._cache_df[cache_key] = (mtime, df)

            df = self._cache_df[cache_key][1]

            resultado = []
            for _, row in df.iterrows():
                fila = [
                    None if (v is None or (isinstance(v, float) and pd.isna(v))) else v
                    for v in row
                ]
                if any(v is not None for v in fila):
                    resultado.append(fila)
            return resultado

        except Exception:
            # Fallback a openpyxl si Pandas falla
            import openpyxl
            wb  = openpyxl.load_workbook(self._path, data_only=True)
            ws  = self._hoja_wb(wb, nombre_hoja)
            res = [
                [ws.cell(r, c).value for c in range(1, ws.max_column + 2)]
                for r in range(fila_inicio, ws.max_row + 1)
                if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 2))
            ]
            return res

    def celda(self, nombre_hoja: str, ref: str):
        """Lee una celda específica por referencia (ej: A1)."""
        import openpyxl
        wb = openpyxl.load_workbook(self._path, data_only=True)
        ws = self._hoja_wb(wb, nombre_hoja)
        val = ws[ref].value
        wb.close()
        return val

    def _hoja_wb(self, wb, nombre: str):
        for h in wb.sheetnames:
            if nombre.lower() in h.lower():
                return wb[h]
        raise KeyError(f"Hoja '{nombre}' no encontrada en el Excel.")

    # ── Escritura ──────────────────────────────────────────────────────────
    def _abrir_escritura(self):
        import openpyxl
        return openpyxl.load_workbook(self._path)

    def _guardar(self, wb):
        wb.save(self._path)
        self._cache_df.clear()  # invalidar cache tras escritura

    def append_fila(self, nombre_hoja: str, valores: list) -> int:
        wb  = self._abrir_escritura()
        ws  = self._hoja_wb(wb, nombre_hoja)
        ws.append(valores)
        fila_nueva = ws.max_row
        self._guardar(wb)
        return fila_nueva

    def update_fila(self, nombre_hoja: str, fila_idx: int, datos: dict, columnas: dict):
        wb = self._abrir_escritura()
        ws = self._hoja_wb(wb, nombre_hoja)
        for campo, letra in columnas.items():
            if campo in datos:
                col_idx = _col_idx(letra)
                ws.cell(fila_idx, col_idx + 1).value = datos[campo]
        self._guardar(wb)

    def delete_fila(self, nombre_hoja: str, fila_idx: int):
        wb = self._abrir_escritura()
        ws = self._hoja_wb(wb, nombre_hoja)
        ws.delete_rows(fila_idx)
        self._guardar(wb)


class AdaptadorSheets:
    """Lee desde Google Sheets vía gspread"""

    def __init__(self, cfg_fuente: dict, base_dir: str):
        try:
            import gspread
            from google.oauth2.service_account import Credentials
        except ImportError:
            raise ImportError("Instalar: pip install gspread google-auth")

        creds_path = os.path.join(base_dir, cfg_fuente.get("creds_json", "creds.json"))
        if not os.path.exists(creds_path):
            creds_path = cfg_fuente.get("creds_json", "creds.json")

        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(creds_path, scopes=scope)
        gc = gspread.authorize(creds)
        self._doc = gc.open_by_key(cfg_fuente["sheets_id"])

    def filas(self, nombre_hoja: str, fila_inicio: int) -> list[list]:
        ws = self._hoja(nombre_hoja)
        todos = ws.get_all_values()
        return [r for r in todos[fila_inicio - 1:] if any(c.strip() for c in r)]

    def celda(self, nombre_hoja: str, ref: str):
        return self._hoja(nombre_hoja).acell(ref).value

    def _hoja(self, nombre: str):
        for ws in self._doc.worksheets():
            if nombre.lower() in ws.title.lower():
                return ws
        raise KeyError(f"Hoja '{nombre}' no encontrada en el Sheet.")

    # ── Escritura ──────────────────────────────────────────────────────────
    def append_fila(self, nombre_hoja: str, valores: list) -> int:
        """Agrega fila al final del Sheet. Retorna índice (1-based)."""
        ws = self._hoja(nombre_hoja)
        ws.append_row(valores, value_input_option="USER_ENTERED")
        return len(ws.get_all_values())

    def update_fila(self, nombre_hoja: str, fila_idx: int, datos: dict, columnas: dict):
        """Actualiza celdas específicas de una fila."""
        ws = self._hoja(nombre_hoja)
        for campo, letra in columnas.items():
            if campo in datos:
                col_num = _col_idx(letra) + 1
                ws.update_cell(fila_idx, col_num, datos[campo])

    def delete_fila(self, nombre_hoja: str, fila_idx: int):
        """Elimina una fila del Sheet."""
        ws = self._hoja(nombre_hoja)
        ws.delete_rows(fila_idx)


# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA UNIVERSAL
# ══════════════════════════════════════════════════════════════════════════════

class Sistema:
    """
    Punto de entrada universal.
    Recibe el nombre de una empresa y carga su JSON de configuración.
    Desde ahí, puede leer cualquier hoja del Excel/Sheet asociado.
    """

    def __init__(self, empresa: str, forzar: str = None):
        """
        empresa: nombre del archivo JSON en /empresas/ (sin extensión)
                 Ej: "kraftdo" carga empresas/kraftdo.json
        forzar:  "local" | "sheets" — sobreescribe el tipo del JSON
        """
        self.nombre = empresa
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.cfg = self._cargar_config(empresa)

        tipo = forzar or self.cfg["fuente"].get("tipo", "local")

        # Variable de entorno tiene prioridad sobre el JSON
        env_id = os.environ.get("SHEETS_ID", "")
        if env_id:
            self.cfg["fuente"]["sheets_id"] = env_id
            if not forzar:
                tipo = "sheets"

        # Leer sheets_id DESPUÉS de aplicar env (fix bug: lectura prematura)
        sheets_id = self.cfg["fuente"].get("sheets_id", "")

        if tipo == "sheets" and sheets_id:
            try:
                self._db = AdaptadorSheets(self.cfg["fuente"], self.base_dir)
                self.modo = "sheets"
            except Exception as e:
                # Si falla Sheets (sin creds, sin conexión), caer a local
                import warnings
                warnings.warn(f"Sheets no disponible ({e}), usando Excel local")
                self._db = AdaptadorLocal(self.cfg["fuente"], self.base_dir)
                self.modo = "local"
        else:
            self._db = AdaptadorLocal(self.cfg["fuente"], self.base_dir)
            self.modo = "local"

    # ── Config ────────────────────────────────────────────────────────────────
    def _cargar_config(self, empresa: str) -> dict:
        rutas = [
            os.path.join(self.base_dir, "empresas", f"{empresa}.json"),
            os.path.join(self.base_dir, f"{empresa}.json"),
            f"{empresa}.json",
        ]
        for r in rutas:
            if os.path.exists(r):
                with open(r, encoding="utf-8") as f:
                    return json.load(f)
        raise FileNotFoundError(
            f"Config '{empresa}.json' no encontrado.\n"
            f"Rutas buscadas: {rutas}"
        )

    # ── API pública ────────────────────────────────────────────────────────────

    def hojas_disponibles(self) -> dict:
        """Lista todas las hojas configuradas con su tipo."""
        return {
            alias: {
                "nombre":      h.get("nombre", alias),
                "tipo":        h.get("tipo"),
                "descripcion": h.get("descripcion", ""),
            }
            for alias, h in self.cfg["hojas"].items()
        }

    def catalogo(self, solo_activos: bool = True) -> dict:
        """Retorna todos los catálogos agrupados por alias de hoja."""
        resultado = {}
        for alias, cfg_hoja in self.cfg["hojas"].items():
            if cfg_hoja.get("tipo") == "catalogo":
                resultado[alias] = self._leer_catalogo(alias, cfg_hoja, solo_activos)
        return resultado

    def catalogo_plano(self, solo_activos: bool = True) -> list:
        """Lista única de todos los productos de todos los catálogos."""
        todos = []
        for alias, prods in self.catalogo(solo_activos).items():
            todos.extend(prods)
        return todos

    def precio(self, sku: str, cantidad: int = 1) -> dict | None:
        """Precio exacto para un SKU y cantidad, aplicando tramos."""
        for prod in self.catalogo_plano():
            if prod.get("sku", "").upper() == sku.upper():
                return self._aplicar_tramo(prod, cantidad)
        return None

    def buscar(self, texto: str) -> list:
        """Búsqueda libre en producto/variante/descripción de catálogos."""
        texto = texto.lower()
        return [
            p for p in self.catalogo_plano()
            if any(
                texto in str(p.get(campo, "")).lower()
                for campo in ("sku", "producto", "variante", "descripcion", "categoria")
            )
        ]

    def registros(self, alias: str, solo_activos: bool = False) -> list:
        """Todas las filas de una hoja de tipo 'registros'."""
        cfg_hoja = self._cfg_hoja(alias)
        filas = self._leer_filas(cfg_hoja)
        if solo_activos:
            filtro = cfg_hoja.get("filtro_activos")
            if filtro:
                col = filtro["columna"]
                vals = filtro.get("valores", [filtro.get("valor")])
                filas = [f for f in filas if f.get(col) in vals]
        return filas

    def kpis(self, alias: str = None) -> dict:
        """Lee celdas de KPIs de hojas tipo 'kpis'."""
        resultado = {}
        hojas_kpi = {
            a: h for a, h in self.cfg["hojas"].items()
            if h.get("tipo") == "kpis" and (alias is None or a == alias)
        }
        for a, cfg_hoja in hojas_kpi.items():
            kpis_cfg = cfg_hoja.get("kpis", {})
            resultado[a] = {
                nombre: _num(self._db.celda(cfg_hoja["nombre"], ref))
                for nombre, ref in kpis_cfg.items()
            }
            # KPIs extra en hojas de registros (ej: Caja)
            kpis_extra = cfg_hoja.get("kpis_celdas", {})
            if kpis_extra:
                resultado[a].update({
                    nombre: _num(self._db.celda(cfg_hoja["nombre"], ref))
                    for nombre, ref in kpis_extra.items()
                })
        # También buscar kpis_celdas en hojas de registros
        for a, cfg_hoja in self.cfg["hojas"].items():
            if "kpis_celdas" in cfg_hoja and (alias is None or a == alias):
                resultado[a] = {
                    nombre: _num(self._db.celda(cfg_hoja["nombre"], ref))
                    for nombre, ref in cfg_hoja["kpis_celdas"].items()
                }
        return resultado

    def cotizar(self, items: list[dict], cliente: str = "", telefono: str = "") -> dict:
        """
        Calcula una cotización completa.
        items: [{"sku": "A01", "cantidad": 2, "descuento": 0}, ...]
        """
        logica = self.cfg.get("logica_negocios", {})
        iva_pct = logica.get("iva", 0.19)
        anticipo_pct = logica.get("anticipo_pct", 0.5)

        lineas = []
        errores = []
        for item in items:
            sku = item.get("sku", "").upper()
            cantidad = int(item.get("cantidad", 1))
            descuento = float(item.get("descuento", 0))
            obs = item.get("obs", "")

            p = self.precio(sku, cantidad)
            if p is None:
                errores.append(f"SKU '{sku}' no encontrado")
                continue

            subtotal = int(p["precio_unitario"] * cantidad * (1 - descuento))
            lineas.append({**p, "descuento": descuento, "subtotal": subtotal, "obs": obs})

        subtotal_neto = sum(l["subtotal"] for l in lineas)
        iva           = round(subtotal_neto * iva_pct)
        total         = subtotal_neto + iva
        anticipo      = round(total * anticipo_pct)
        saldo         = total - anticipo

        return {
            "empresa":       self.cfg["empresa"]["nombre"],
            "cliente":       cliente,
            "telefono":      telefono,
            "fecha":         datetime.today().strftime("%d/%m/%Y"),
            "lineas":        lineas,
            "subtotal_neto": subtotal_neto,
            "iva":           iva,
            "total":         total,
            "anticipo":      anticipo,
            "saldo":         saldo,
            "errores":       errores,
        }

    def info_empresa(self) -> dict:
        return self.cfg["empresa"]

    # ── CRUD ───────────────────────────────────────────────────────────────

    def crear(self, alias: str, datos: dict) -> dict:
        """
        POST — Agrega una fila nueva a la hoja.
        Calcula campos derivados (total, iva, etc.) si no vienen en datos.
        Retorna la fila creada con su índice.
        """
        cfg_hoja = self._cfg_hoja(alias)
        columnas = cfg_hoja.get("columnas", {})
        fila_ini = cfg_hoja.get("fila_datos", 5)

        # Completar campos calculados básicos
        datos = self._completar_calculados(datos, cfg_hoja)

        # Construir lista de valores en orden de columnas
        max_col = max(_col_idx(l) for l in columnas.values()) + 1
        valores = [None] * max_col
        for campo, letra in columnas.items():
            idx = _col_idx(letra)
            valores[idx] = datos.get(campo)

        fila_idx = self._db.append_fila(cfg_hoja["nombre"], valores)
        # Recargar workbook en modo lectura para reflejar cambios
        self._reload()
        return {"fila": fila_idx, "datos": datos, "ok": True}

    def actualizar(self, alias: str, id_valor: str, datos: dict) -> dict:
        """
        PUT — Modifica campos de una fila identificada por su campo identificador.
        Solo actualiza los campos que vienen en datos.
        """
        cfg_hoja  = self._cfg_hoja(alias)
        columnas  = cfg_hoja.get("columnas", {})
        ident     = cfg_hoja.get("identificador")
        fila_ini  = cfg_hoja.get("fila_datos", 5)

        if not ident:
            raise ValueError(f"La hoja '{alias}' no tiene campo identificador configurado")

        fila_idx = self._buscar_fila(alias, ident, id_valor)
        if fila_idx is None:
            raise KeyError(f"No se encontró '{id_valor}' en {alias}.{ident}")

        self._db.update_fila(cfg_hoja["nombre"], fila_idx, datos, columnas)
        self._reload()
        return {"fila": fila_idx, "actualizado": list(datos.keys()), "ok": True}

    def eliminar(self, alias: str, id_valor: str) -> dict:
        """
        DELETE — Elimina la fila identificada por su campo identificador.
        """
        cfg_hoja = self._cfg_hoja(alias)
        ident    = cfg_hoja.get("identificador")

        if not ident:
            raise ValueError(f"La hoja '{alias}' no tiene campo identificador configurado")

        fila_idx = self._buscar_fila(alias, ident, id_valor)
        if fila_idx is None:
            raise KeyError(f"No se encontró '{id_valor}' en {alias}.{ident}")

        self._db.delete_fila(cfg_hoja["nombre"], fila_idx)
        self._reload()
        return {"eliminado": id_valor, "fila": fila_idx, "ok": True}

    def buscar_filtros(self, alias: str, filtros: dict) -> list:
        """
        SEARCH con DuckDB — SQL completo sobre cualquier hoja Excel/Sheets.

        Operadores soportados:
            campo          → igual (case-insensitive)
            campo__gt      → mayor que
            campo__gte     → mayor o igual
            campo__lt      → menor que
            campo__lte     → menor o igual
            campo__like    → contiene (ILIKE)
            campo__in      → IN (lista separada por coma)
            campo__sql     → expresión SQL raw (avanzado)

        Ejemplos:
            {"estado": "Activo"}
            {"precio_1__gt": "5000", "estado__in": "Activo,Confirmado"}
            {"producto__like": "taza"}
            {"total__sql": "total BETWEEN 5000 AND 50000"}
        """
        try:
            return self._buscar_duckdb(alias, filtros)
        except Exception:
            # Fallback a filtrado en memoria si DuckDB falla
            return self._buscar_memoria(alias, filtros)

    def _buscar_duckdb(self, alias: str, filtros: dict) -> list:
        """Implementación DuckDB — SQL real sobre DataFrame Pandas."""
        import duckdb
        import pandas as pd

        filas = self.registros(alias)
        if not filas:
            return []

        df = pd.DataFrame(filas)
        if df.empty:
            return []

        # Convertir columnas numéricas automáticamente
        for col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors="ignore")
            except Exception:
                pass

        # Construir cláusula WHERE desde filtros
        condiciones = []
        params      = {}

        for clave, valor in filtros.items():
            if "__" in clave:
                campo, op = clave.rsplit("__", 1)
            else:
                campo, op = clave, "eq"

            if campo not in df.columns:
                continue

            param_name = f"p_{campo}_{op}".replace("__","_")

            if op == "sql":
                # SQL raw — el valor ES la condición
                condiciones.append(str(valor))
            elif op == "eq":
                condiciones.append(f"LOWER(CAST({campo} AS VARCHAR)) = LOWER(${param_name})")
                params[param_name] = str(valor)
            elif op == "like":
                condiciones.append(f"LOWER(CAST({campo} AS VARCHAR)) LIKE LOWER(${param_name})")
                params[param_name] = f"%{valor}%"
            elif op == "in":
                opciones = [v.strip() for v in str(valor).split(",")]
                placeholders = ", ".join(f"'{o.lower()}'" for o in opciones)
                condiciones.append(f"LOWER(CAST({campo} AS VARCHAR)) IN ({placeholders})")
            elif op in ("gt", "gte", "lt", "lte"):
                op_sql = {"gt": ">", "gte": ">=", "lt": "<", "lte": "<="}[op]
                condiciones.append(f"TRY_CAST({campo} AS DOUBLE) {op_sql} ${param_name}")
                params[param_name] = float(str(valor).replace(",", "."))

        where_clause = " AND ".join(condiciones) if condiciones else "TRUE"
        sql = f"SELECT * FROM df WHERE {where_clause}"

        con = duckdb.connect()
        resultado_df = con.execute(sql, list(params.values()) if params else []).df()
        con.close()

        return resultado_df.where(resultado_df.notna(), None).to_dict(orient="records")

    def _buscar_memoria(self, alias: str, filtros: dict) -> list:
        """Fallback: filtrado en memoria (original)."""
        filas = self.registros(alias)
        resultado = []
        for fila in filas:
            pasa = True
            for clave, valor in filtros.items():
                campo, op = (clave.rsplit("__", 1) if "__" in clave else (clave, "eq"))
                val_fila = fila.get(campo)
                if val_fila is None:
                    pasa = False; break
                try:
                    vn = float(str(val_fila).replace(",",".")); cn = float(str(valor).replace(",",".")); en = True
                except: en = False
                if op == "eq"   and str(val_fila).lower() != str(valor).lower(): pasa = False
                elif op == "like" and str(valor).lower() not in str(val_fila).lower(): pasa = False
                elif op == "in"  and str(val_fila).lower() not in [v.strip().lower() for v in str(valor).split(",")]: pasa = False
                elif en:
                    if op == "gt"  and not vn >  cn: pasa = False
                    if op == "gte" and not vn >= cn: pasa = False
                    if op == "lt"  and not vn <  cn: pasa = False
                    if op == "lte" and not vn <= cn: pasa = False
                if not pasa: break
            if pasa: resultado.append(fila)
        return resultado

    def query(self, alias: str, sql_where: str) -> list:
        """
        SQL directo sobre cualquier hoja — DuckDB completo.
        
        Ejemplo:
            s.query("pedidos", "estado = 'Confirmado' AND total > 50000")
            s.query("caja", "fecha >= '2026-01-01' ORDER BY fecha DESC LIMIT 10")
        """
        return self._buscar_duckdb(alias, {"__sql": sql_where})

    def schema(self, alias: str) -> dict:
        """
        Retorna el schema de una hoja — campos, tipos y restricciones.
        Usado por el formulario dinámico para saber qué renderizar.
        """
        cfg_hoja = self._cfg_hoja(alias)
        columnas = cfg_hoja.get("columnas", {})
        # Buscar estados en múltiples lugares del config de hoja
        estados = (
            cfg_hoja.get("logica", {}).get("estados") or
            cfg_hoja.get("estados_validos") or
            self.cfg.get("logica_negocios", {}).get("estados_pedido_todos") or
            []
        )
        ident    = cfg_hoja.get("identificador")

        from generator import inferir_tipo
        campos = {}
        for campo, letra in columnas.items():
            tipo_laravel, mod = inferir_tipo(campo)
            tipo_form = _tipo_form(tipo_laravel, campo, estados)
            campos[campo] = {
                "tipo":       tipo_form,
                "requerido":  "nullable" not in mod,
                "columna":    letra,
                "opciones":   estados if tipo_form == "select" and campo == "estado" else [],
                "es_id":      campo == ident,
            }

        return {
            "alias":        alias,
            "nombre":       cfg_hoja.get("nombre", alias),
            "tipo":         cfg_hoja.get("tipo"),
            "identificador": ident,
            "campos":       campos,
        }

    # ── Internos ───────────────────────────────────────────────────────────────

    def _cfg_hoja(self, alias: str) -> dict:
        if alias not in self.cfg["hojas"]:
            disponibles = list(self.cfg["hojas"].keys())
            raise KeyError(f"Hoja '{alias}' no configurada. Disponibles: {disponibles}")
        return self.cfg["hojas"][alias]

    def _reload(self):
        """Recarga el adaptador para reflejar cambios escritos."""
        try:
            if self.modo == "local":
                self._db = AdaptadorLocal(self.cfg["fuente"], self.base_dir)
            # Sheets no necesita reload (API en tiempo real)
        except Exception:
            pass

    def _buscar_fila(self, alias: str, campo_id: str, valor: str) -> int | None:
        """Busca el número de fila (absoluto en el archivo) de un registro."""
        cfg_hoja = self._cfg_hoja(alias)
        columnas = cfg_hoja.get("columnas", {})
        fila_ini = cfg_hoja.get("fila_datos", 5)

        if campo_id not in columnas:
            return None

        col_idx = _col_idx(columnas[campo_id])
        filas_raw = self._db.filas(cfg_hoja["nombre"], fila_ini)

        for i, fila in enumerate(filas_raw):
            v = fila[col_idx] if col_idx < len(fila) else None
            if v is not None and str(v).strip() == str(valor).strip():
                return fila_ini + i  # fila absoluta en el archivo

        return None

    def _completar_calculados(self, datos: dict, cfg_hoja: dict) -> dict:
        """Calcula campos derivados simples si no vienen en datos."""
        d = {**datos}
        logica = self.cfg.get("logica_negocios", {})
        iva_pct      = logica.get("iva", 0.19)
        anticipo_pct = logica.get("anticipo_pct", 0.5)

        # precio × cantidad → total
        if "total" not in d and "precio" in d and "cantidad" in d:
            try:
                d["total"] = round(_num(d["precio"]) * _num(d["cantidad"]), 2)
            except Exception:
                pass

        # total × iva → iva (si existe campo)
        if "iva" not in d and "total" in d:
            d["iva"] = round(_num(d["total"]) * iva_pct, 2)

        # total × anticipo_pct → anticipo
        if "anticipo" not in d and "total" in d:
            total_con_iva = _num(d["total"]) * (1 + iva_pct)
            d["anticipo"] = round(total_con_iva * anticipo_pct, 2)
            d["saldo"]    = round(total_con_iva - d["anticipo"], 2)

        return d

    def _leer_filas(self, cfg_hoja: dict) -> list[dict]:
        """Lee filas de cualquier hoja y mapea columnas según config.
        
        Si la hoja es consolidada (consolidado: true), lee desde cada hoja
        fuente y agrega el campo 'tipo' automáticamente.
        """
        # ── Hoja consolidada: leer desde las fuentes originales ───────────
        if cfg_hoja.get("consolidado"):
            return self._leer_filas_consolidadas(cfg_hoja)

        nombre    = cfg_hoja["nombre"]
        fila_ini  = cfg_hoja.get("fila_datos", 5)
        columnas  = cfg_hoja.get("columnas", {})
        ident     = cfg_hoja.get("identificador")

        raw = self._db.filas(nombre, fila_ini)
        resultado = []
        for fila in raw:
            fila_ext = list(fila) + [None] * 30
            registro = {}
            for campo, letra in columnas.items():
                if letra == "[CALCULADO]":
                    continue
                idx = _col_idx(letra)
                registro[campo] = _limpio(fila_ext[idx])

            if ident and not registro.get(ident):
                continue
            elif not ident and not any(v for v in registro.values()):
                continue

            resultado.append(registro)
        return resultado

    def _inferir_columnas_desde_excel(self, nombre_hoja: str,
                                      fila_header: int, cfg_hoja: dict) -> dict:
        """
        Lee los headers reales del Excel para una hoja y genera el mapeo
        campo → letra basándose en los nombres de columna del JSON consolidado.
        Usado como fallback cuando columnas_por_fuente no está disponible.
        """
        import openpyxl
        letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(self._db._path, data_only=True)
        ws = next((wb[h] for h in wb.sheetnames
                   if nombre_hoja.lower() in h.lower()), None)
        if not ws:
            wb.close()
            return {}

        # Leer la fila de headers
        headers_row = [ws.cell(fila_header, c).value
                       for c in range(1, min(30, ws.max_column + 1))]
        wb.close()

        # Normalizar headers → slug
        import unicodedata, re as _re
        def _slug_h(s):
            if s is None: return ""
            s = unicodedata.normalize("NFKD", str(s))
            s = s.encode("ascii", "ignore").decode("ascii")
            return _re.sub(r"[^a-z0-9]", "_", s.lower()).strip("_")

        # Construir mapeo campo → letra
        slug_a_letra = {}
        for idx, header in enumerate(headers_row):
            if header and idx < 26:
                slug_a_letra[_slug_h(header)] = letras[idx]

        # Mapear campos del JSON a letras reales del Excel
        columnas_inferidas = {}
        for campo in cfg_hoja.get("columnas", {}):
            if campo == "tipo":
                continue
            slug_campo = _slug_h(campo)
            # Buscar coincidencia exacta o parcial
            if slug_campo in slug_a_letra:
                columnas_inferidas[campo] = slug_a_letra[slug_campo]
            else:
                # Buscar por similitud
                for slug_h, letra in slug_a_letra.items():
                    if slug_campo in slug_h or slug_h in slug_campo:
                        if campo not in columnas_inferidas:
                            columnas_inferidas[campo] = letra
                        break

        return columnas_inferidas

    def _leer_filas_consolidadas(self, cfg_hoja: dict) -> list[dict]:
        """
        Lee múltiples hojas fuente y las combina con campo tipo.
        Usa columnas_por_fuente para respetar el mapeo específico de cada hoja.
        """
        import unicodedata, re as _re

        fuentes           = cfg_hoja.get("fuentes", [])
        fila_ini          = cfg_hoja.get("fila_datos", 5)
        ident             = cfg_hoja.get("identificador")
        valores_tipo      = cfg_hoja.get("valores_tipo", [])
        columnas_x_fuente = cfg_hoja.get("columnas_por_fuente", {})

        def _slug(s):
            s = unicodedata.normalize("NFKD", str(s))
            s = s.encode("ascii", "ignore").decode("ascii")
            return _re.sub(r"[^a-z0-9]", "_", s.lower())

        def _es_fila_instruccion(registro: dict, ident_campo: str) -> bool:
            """Detecta filas de ayuda/instrucciones del Excel."""
            if not ident_campo:
                return False
            val = str(registro.get(ident_campo, "") or "")
            # Filas de instrucción tienen SKU con más de 20 chars o palabras clave
            if len(val) > 20:
                return True
            palabras_clave = ("amarillo", "auto =", "verde =", "negro auto", "color")
            if any(p in val.lower() for p in palabras_clave):
                return True
            return False

        # Obtener nombres reales de hojas del Excel
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._db._path, read_only=True, data_only=True)
            hojas_excel = wb.sheetnames
            wb.close()
        except Exception:
            return []

        resultado = []
        for i, fuente_alias in enumerate(fuentes):
            # Buscar hoja real en el Excel por similitud
            slug_fuente = _slug(fuente_alias)
            nombre_real = None
            mejor_score = 0
            for hoja_excel in hojas_excel:
                slug_excel = _slug(hoja_excel)
                s1 = slug_fuente.strip("_")
                s2 = slug_excel.strip("_")
                score = sum(1 for a, b in zip(s1, s2) if a == b)
                if s1 in s2 or s2 in s1:
                    score += 20
                if score > mejor_score:
                    mejor_score = score
                    nombre_real = hoja_excel

            if not nombre_real:
                continue

            tipo_val = valores_tipo[i] if i < len(valores_tipo) else fuente_alias

            # ── USAR COLUMNAS ESPECÍFICAS DE ESTA FUENTE ──────────────────
            columnas = columnas_x_fuente.get(fuente_alias, {})
            if not columnas:
                # columnas_por_fuente no disponible (JSON generado antes del fix)
                # → reconstruir desde los headers reales del Excel
                try:
                    columnas = self._inferir_columnas_desde_excel(
                        nombre_real, fila_ini - 1, cfg_hoja
                    )
                except Exception:
                    columnas = {k: v for k, v in cfg_hoja.get("columnas", {}).items()
                                if v != "[CALCULADO]"}
            else:
                columnas = {k: v for k, v in columnas.items()
                            if v != "[CALCULADO]"}
            columnas.pop("tipo", None)

            try:
                raw = self._db.filas(nombre_real, fila_ini)
            except Exception:
                continue

            for fila in raw:
                fila_ext = list(fila) + [None] * 30
                registro = {"tipo": tipo_val}
                for campo, letra in columnas.items():
                    if not letra or letra == "[CALCULADO]":
                        continue
                    try:
                        idx = _col_idx(letra)
                        registro[campo] = _limpio(fila_ext[idx])
                    except Exception:
                        pass

                # ── FILTRAR FILAS DE INSTRUCCIONES ────────────────────────
                if _es_fila_instruccion(registro, ident):
                    continue

                if ident and not registro.get(ident):
                    continue
                elif not ident and not any(v for k, v in registro.items() if k != "tipo"):
                    continue

                resultado.append(registro)

        return resultado

    def _leer_catalogo(self, alias: str, cfg_hoja: dict, solo_activos: bool) -> list:
        """Lee una hoja de catálogo y agrega metadatos de tipo y precio."""
        filas = self._leer_filas(cfg_hoja)
        filtro = cfg_hoja.get("filtro_activo", {})
        precios_cfg = cfg_hoja.get("precios", {})

        resultado = []
        for f in filas:
            # Filtro activo
            if solo_activos and filtro:
                col = filtro.get("columna")
                val = filtro.get("valor", "")
                campo_val = f.get(col) or ""
                if col and str(campo_val).lower() != str(val).lower():
                    continue

            # Construir precio_N desde los campos mapeados
            p = {**f, "tipo": alias, "catalogo": alias}
            # Buscar precio_1 desde el campo configurado, con fallbacks
            campo_p1  = precios_cfg.get("1", "precio_1")
            campo_p5  = precios_cfg.get("5", "precio_5")
            campo_p10 = precios_cfg.get("10", "precio_10")
            p["precio_1"]  = _num(f.get(campo_p1) or f.get("precio_1") or f.get("precio_unit_____") or f.get("precio_pack____") or 0)
            p["precio_5"]  = _num(f.get(campo_p5) or f.get("precio_5") or f.get("precio_5_____") or 0) or p["precio_1"]
            p["precio_10"] = _num(f.get(campo_p10) or f.get("precio_10") or f.get("precio_10_____") or 0) or p["precio_5"]

            resultado.append(p)
        return resultado

    def _aplicar_tramo(self, producto: dict, cantidad: int) -> dict:
        """Determina precio según tramos de cantidad."""
        if cantidad >= 10:
            precio = producto.get("precio_10") or producto.get("precio_5") or producto.get("precio_1", 0)
            tramo = "10+"
        elif cantidad >= 5:
            precio = producto.get("precio_5") or producto.get("precio_1", 0)
            tramo = "5+"
        else:
            precio = producto.get("precio_1", 0)
            tramo = "unit"

        precio = _num(precio)
        return {
            "sku":             _str(producto.get("sku")),
            "producto":        _str(producto.get("producto")),
            "variante":        _str(producto.get("variante") or producto.get("descripcion")),
            "categoria":       _str(producto.get("categoria") or producto.get("tipo")),
            "cantidad":        cantidad,
            "precio_unitario": int(precio),
            "subtotal":        int(precio * cantidad),
            "tramo":           tramo,
            "tipo":            _str(producto.get("tipo")),
        }


# ── Test ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    s = Sistema("kraftdo")
    print(f"Empresa: {s.cfg['empresa']['nombre']} | Modo: {s.modo}")
    print(f"Hojas: {list(s.hojas_disponibles().keys())}")

    cat = s.catalogo()
    total = sum(len(v) for v in cat.values())
    print(f"\nCatálogo: {total} productos")
    for alias, prods in cat.items():
        print(f"  {alias}: {len(prods)} productos")

    print(f"\nPrecio A01 x1:  {s.precio('A01', 1)}")
    print(f"Precio A01 x5:  {s.precio('A01', 5)}")
    print(f"Precio S01 x10: {s.precio('S01', 10)}")

    cot = s.cotizar([
        {"sku": "A01", "cantidad": 2},
        {"sku": "S01", "cantidad": 5},
    ], cliente="Juan Pérez", telefono="+56912345678")
    print(f"\nCotización: Total ${cot['total']:,} | Anticipo ${cot['anticipo']:,}")
    print(f"KPIs: {s.kpis()}")
