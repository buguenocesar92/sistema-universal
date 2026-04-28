#!/usr/bin/env python3
"""
Importa datos del Excel al MySQL que usa el panel Filament.
Uso: python3 importar_excel_a_mysql.py kraftdo_bd /tmp/kraftdo_laravel_real
"""
import sys, json, os, re
import openpyxl
import mysql.connector
from pathlib import Path

def col_letra_a_num(letra):
    result = 0
    for c in letra.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result

def _construir_mapa_sinonimos(cfg: dict) -> dict:
    """Convierte cfg.sinonimos_modelo {canonico: [variantes]} en un mapa
    plano {variante: canonico, canonico: canonico} para lookup rápido."""
    mapa = {}
    for canonico, variantes in (cfg.get("sinonimos_modelo") or {}).items():
        mapa[canonico] = canonico
        for var in variantes:
            mapa[var] = canonico
    return mapa


def _hojas_que_alimentan_agregado(cfg: dict) -> dict:
    """Devuelve {alias_hoja: campo_grupo_a_canonizar}.

    Para cada hoja tipo "agregado": tanto sus fuentes (el campo_grupo
    en el origen) como la hoja agregada misma (su identificador o
    agrupar_por) deben canonizarse para evitar duplicados.
    """
    mapa = {}
    for alias, hoja in cfg.get("hojas", {}).items():
        if hoja.get("tipo") != "agregado":
            continue
        # La hoja agregada misma: canoniza su agrupar_por / identificador.
        campo_propio = hoja.get("agrupar_por") or hoja.get("identificador")
        if campo_propio:
            mapa[alias] = campo_propio
        for fuente in hoja.get("fuentes", []):
            alias_f  = fuente.get("hoja")
            campo_g  = fuente.get("campo_grupo")
            if alias_f and campo_g:
                mapa[alias_f] = campo_g
    return mapa


def _construir_ai_canonizar(cfg: dict, wb, base: str, empresa: str) -> dict:
    """Pre-pass: para cada hoja con columnas que tienen `ai_canonizar`,
    extrae valores únicos del Excel y obtiene el mapeo IA.
    Devuelve {(alias, campo): {valor_orig: canonico}}."""
    try:
        from ai_cleaner import ai_normalizar_columna
    except Exception:
        return {}
    mapeos = {}
    for alias, hoja_cfg in cfg.get("hojas", {}).items():
        cols = hoja_cfg.get("columnas", {}) or {}
        if not isinstance(cols, dict):
            continue
        nombre_hoja = hoja_cfg.get("nombre")
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        fila_ini = hoja_cfg.get("fila_datos", 5)
        for campo, valor in cols.items():
            # Soporte dos formatos:
            #   "campo": "G"                          (legacy plano)
            #   "campo": {"columna": "G", "ai_canonizar": [...]}
            if not isinstance(valor, dict):
                continue
            cats = valor.get("ai_canonizar")
            if not cats:
                continue
            letra = valor.get("columna")
            if not letra:
                continue
            try:
                col_idx = col_letra_a_num(letra)
            except Exception:
                continue
            valores_unicos = []
            seen = set()
            fin = min(ws.max_row or fila_ini, fila_ini + 500)
            for r in range(fila_ini, fin + 1):
                v = ws.cell(r, col_idx).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s or s in seen:
                    continue
                seen.add(s)
                valores_unicos.append(s)
            if not valores_unicos:
                continue
            cache_path = os.path.join(
                base, "ai_cache", f"{empresa}_{alias}_{campo}.json"
            )
            mapeo = ai_normalizar_columna(
                valores_unicos=valores_unicos,
                categorias=cats,
                cache_path=cache_path,
                campo=f"{alias}.{campo}",
            )
            if mapeo:
                mapeos[(alias, campo)] = mapeo
    return mapeos


def _aplanar_columnas(cols: dict) -> dict:
    """Convierte el formato extendido {campo: {columna: 'G', ai_canonizar: [...]}}
    al formato plano {campo: 'G'} que el resto del importer espera."""
    aplanado = {}
    for k, v in (cols or {}).items():
        if isinstance(v, dict):
            letra = v.get("columna")
            if letra:
                aplanado[k] = letra
        else:
            aplanado[k] = v
    return aplanado


def importar(empresa: str, laravel_dir: str):
    base = os.path.dirname(os.path.abspath(__file__))
    cfg_path = os.path.join(base, "empresas", f"{empresa}.json")
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)

    excel_path = os.path.join(base, cfg["fuente"]["archivo"])
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    sinonimos       = _construir_mapa_sinonimos(cfg)
    canoniza_campos = _hojas_que_alimentan_agregado(cfg)
    if sinonimos:
        print(f"  🔁 sinónimos cargados: {len(sinonimos)} mapeos → "
              f"{set(sinonimos.values())}")

    # ── v25-fase4: Detección de schema drift Excel ↔ JSON ─────────────
    # Compara los headers reales del Excel con las columnas del JSON.
    # No interrumpe; solo advierte para que el usuario actualice el JSON.
    for alias_d, hoja_d in cfg.get("hojas", {}).items():
        if hoja_d.get("tipo") not in ("catalogo", "registros", "agregado"):
            continue
        nombre_hoja_d = hoja_d.get("nombre")
        if nombre_hoja_d not in wb.sheetnames:
            continue
        ws_d = wb[nombre_hoja_d]
        fila_h = max(1, hoja_d.get("fila_datos", 5) - 1)
        cols_json = _aplanar_columnas(hoja_d.get("columnas", {}) or {})
        letras_json = {v.upper() for v in cols_json.values()}
        # Headers Excel
        from openpyxl.utils import get_column_letter
        nuevas = []
        for c in range(1, (ws_d.max_column or 0) + 1):
            h = ws_d.cell(fila_h, c).value
            if h is None or str(h).strip() == "":
                continue
            letra = get_column_letter(c)
            if letra not in letras_json:
                # ¿Tiene datos abajo?
                tiene = False
                fin = min(ws_d.max_row or fila_h + 1, fila_h + 30)
                for r in range(fila_h + 1, fin + 1):
                    if ws_d.cell(r, c).value is not None:
                        tiene = True
                        break
                if tiene:
                    hdr = re.sub(r'\s+', ' ', str(h)).strip()[:40]
                    nuevas.append((letra, hdr))
        if nuevas:
            print(f"  ⚠️  Schema drift en {alias_d}:")
            print(f"     Columnas en Excel NO mapeadas en JSON:")
            for letra, hdr in nuevas[:10]:
                print(f"        col {letra}: {hdr!r}")
            print(f"     → Agrégalas al JSON y regenera para capturarlas.")
        # Columnas en JSON sin existir en Excel (letras fuera del rango)
        max_col_excel = ws_d.max_column or 0
        from openpyxl.utils import column_index_from_string
        faltantes = []
        for campo_j, letra_j in cols_json.items():
            try:
                idx = column_index_from_string(letra_j)
            except Exception:
                continue
            if idx > max_col_excel:
                faltantes.append((campo_j, letra_j))
        if faltantes:
            print(f"  ⚠️  Columnas del JSON {alias_d} fuera del rango del Excel:")
            for c, l in faltantes[:5]:
                print(f"        {c} (col {l})")

    # Pre-pass IA: canonización de columnas con ai_canonizar
    ai_mapeos = _construir_ai_canonizar(cfg, wb, base, empresa)
    if ai_mapeos:
        print(f"  🤖 ai_cleaner: {len(ai_mapeos)} columnas con mapeo activo")

    # Leer .env de Laravel para obtener credenciales DB
    env_path = os.path.join(laravel_dir, ".env")
    env = {}
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()

    conn = mysql.connector.connect(
        host=env.get("DB_HOST", "127.0.0.1"),
        port=int(env.get("DB_PORT", 3306)),
        user=env.get("DB_USERNAME", "kraftdo"),
        password=env.get("DB_PASSWORD", ""),
        database=env.get("DB_DATABASE", "kraftdo")
    )
    cursor = conn.cursor()

    hojas = cfg["hojas"]

    # Pre-pass: hojas tipo matriz_asistencia → despivota a asistencias + pagos_quincena
    for alias, hoja_cfg in hojas.items():
        if hoja_cfg.get("tipo") != "matriz_asistencia":
            continue
        nombre_hoja = hoja_cfg["nombre"]
        ws = next((wb[h] for h in wb.sheetnames if h == nombre_hoja), None)
        if not ws:
            print(f"  ⚠️  Hoja '{nombre_hoja}' no encontrada")
            continue
        mes        = hoja_cfg.get("mes_actual", "")
        f_ini      = hoja_cfg.get("fila_inicio", 5)
        f_fin      = hoja_cfg.get("fila_fin", 18)
        c_codigo   = col_letra_a_num(hoja_cfg.get("col_codigo_obra", "B"))
        c_obra     = col_letra_a_num(hoja_cfg.get("col_obra", "C"))
        c_trab     = col_letra_a_num(hoja_cfg.get("col_trabajador", "D"))
        cols_q1    = [col_letra_a_num(l) for l in hoja_cfg.get("cols_quincena1", [])]
        cols_q2    = [col_letra_a_num(l) for l in hoja_cfg.get("cols_quincena2", [])]
        c_pago_q   = col_letra_a_num(hoja_cfg.get("col_pago_quincena", "T"))
        c_pago_l   = col_letra_a_num(hoja_cfg.get("col_pago_liquidacion", "AH"))

        n_asis = 0
        n_pago = 0
        for fila in range(f_ini, f_fin + 1):
            trabajador = ws.cell(fila, c_trab).value
            if trabajador is None or str(trabajador).strip() == "":
                continue
            trabajador = str(trabajador).strip()
            obra       = ws.cell(fila, c_obra).value
            codigo_obra = ws.cell(fila, c_codigo).value

            # Despivot quincena1 → días 1..15 del mes
            for offset, c in enumerate(cols_q1, start=1):
                v = ws.cell(fila, c).value
                if v is None or str(v).strip() == "":
                    continue
                estado = str(v).strip().upper()[:2]
                fecha = f"{mes}-{offset:02d}"
                try:
                    cursor.execute(
                        "INSERT IGNORE INTO asistencias "
                        "(trabajador, obra, codigo_obra, fecha, mes, estado, created_at, updated_at) "
                        "VALUES (%s,%s,%s,%s,%s,%s,NOW(),NOW())",
                        [trabajador, obra, codigo_obra, fecha, mes, estado]
                    )
                    n_asis += 1
                except Exception:
                    pass

            # Despivot quincena2 → días 16..(15+len)
            for offset, c in enumerate(cols_q2, start=16):
                v = ws.cell(fila, c).value
                if v is None or str(v).strip() == "":
                    continue
                estado = str(v).strip().upper()[:2]
                fecha = f"{mes}-{offset:02d}"
                try:
                    cursor.execute(
                        "INSERT IGNORE INTO asistencias "
                        "(trabajador, obra, codigo_obra, fecha, mes, estado, created_at, updated_at) "
                        "VALUES (%s,%s,%s,%s,%s,%s,NOW(),NOW())",
                        [trabajador, obra, codigo_obra, fecha, mes, estado]
                    )
                    n_asis += 1
                except Exception:
                    pass

            # Pagos quincena + liquidación
            for periodo, col in [("quincena", c_pago_q), ("liquidacion", c_pago_l)]:
                v = ws.cell(fila, col).value
                if v is None:
                    continue
                try:
                    monto = float(v)
                except (TypeError, ValueError):
                    continue
                try:
                    cursor.execute(
                        "INSERT IGNORE INTO pagos_quincena "
                        "(trabajador, mes, periodo, monto, created_at, updated_at) "
                        "VALUES (%s,%s,%s,%s,NOW(),NOW())",
                        [trabajador, mes, periodo, monto]
                    )
                    n_pago += 1
                except Exception:
                    pass

        conn.commit()
        print(f"  ✓ matriz {nombre_hoja}: {n_asis} asistencias, {n_pago} pagos importados")

    for alias, hoja_cfg in hojas.items():
        tipo = hoja_cfg.get("tipo")
        if tipo not in ("catalogo", "registros", "agregado"):
            continue

        nombre_hoja = hoja_cfg["nombre"]
        ws = next((wb[h] for h in wb.sheetnames if h == nombre_hoja), None)
        if not ws:
            print(f"  ⚠️  Hoja '{nombre_hoja}' no encontrada")
            continue

        columnas = _aplanar_columnas(hoja_cfg.get("columnas", {}))
        fila_ini = hoja_cfg.get("fila_datos", 5)
        if tipo == "kpis":
            continue

        # Leer nombre de tabla real desde la migración generada
        import glob, re as _re
        tabla_mysql = None
        patron = os.path.join(laravel_dir, "database", "migrations", f"*create_{alias}*")
        archivos = glob.glob(patron)
        if archivos:
            with open(archivos[0]) as mf:
                m = _re.search(r"Schema::create\(['\"]([\w]+)['\"]", mf.read())
                if m:
                    tabla_mysql = m.group(1)
        if not tabla_mysql:
            # Fallback: plural simple
            tabla_mysql = alias if alias.endswith("s") else alias + "s"
            print(f"  ⚠️  Tabla para {alias} inferida como {tabla_mysql}")

        col_indices = {campo: col_letra_a_num(letra) for campo, letra in columnas.items()}
        campos = list(columnas.keys())
        ident   = hoja_cfg.get("identificador")

        # Palabras que, cuando aparecen como VALOR EXACTO del identificador,
        # marcan filas de header/totales (no datos). Match exacto, no substring,
        # para no descartar conceptos legítimos como "TOTAL BRUTO" si son
        # registros de negocio reales — pero sí filtra filas resumen genéricas.
        SKIP_EXACT = {"TOTALES", "TOTAL", "SUBTOTAL", "INGRESAR DATOS",
                      "[AUTO]", "[INGRESAR]", "AMARILLO"}

        # v25-fase3: importer incremental con upsert diferencial.
        # Métricas por hoja: nuevos / actualizados / sin_cambio / errores.
        import time as _t, hashlib as _hl
        t0 = _t.time()
        n_nuevos = n_upd = n_iguales = n_err = 0
        # Identificador efectivo: si no hay declarado, usar la primera columna.
        ident_efectivo = ident or (campos[0] if campos else None)

        ultimo_ident = None  # carry-forward
        for row in range(fila_ini, ws.max_row + 1):
            valores = {}
            for campo, idx in col_indices.items():
                val = ws.cell(row, idx).value
                valores[campo] = val

            # 1) Filas completamente vacías
            if all(v is None or str(v).strip() == "" for v in valores.values()):
                continue

            # 2) Si hay identificador definido
            if ident:
                ident_val = valores.get(ident)
                ident_vacio = ident_val is None or str(ident_val).strip() == ""
                if ident_vacio:
                    # Fila continuación: hereda identificador previo si existe.
                    # Si nunca hubo identificador antes, sí saltar.
                    if ultimo_ident is None:
                        continue
                    valores[ident] = ultimo_ident
                else:
                    # 3) Identificador es header de totales exacto
                    s_ident = str(ident_val).strip()
                    if s_ident.upper() in SKIP_EXACT:
                        continue
                    # 3b) Notas a pie / fórmulas explicativas: tienen "=" o son muy largas
                    if "=" in s_ident or len(s_ident) > 50:
                        continue
                    ultimo_ident = ident_val

            # 4) Skip si el primer valor empieza con "="
            primer_val = str(valores.get(ident) if ident else list(valores.values())[0] or "").strip()
            if primer_val.startswith("="):
                continue

            # 5) Canonización de modelo: si esta hoja alimenta a un agregado,
            # mapear el campo_grupo de variante → canónico vía sinonimos.
            campo_canonizar = canoniza_campos.get(alias)
            if campo_canonizar and sinonimos:
                v = valores.get(campo_canonizar)
                if v is not None:
                    s = str(v).strip()
                    if s in sinonimos:
                        valores[campo_canonizar] = sinonimos[s]

            # 6) Canonización IA: aplica mapeos {valor_orig → canonico}
            # producidos por ai_cleaner (claude-haiku-3-5 + caché).
            for (alias_m, campo_m), mapeo in ai_mapeos.items():
                if alias_m != alias:
                    continue
                v = valores.get(campo_m)
                if v is None:
                    continue
                s = str(v).strip()
                if s in mapeo and mapeo[s] != s:
                    valores[campo_m] = mapeo[s]

            # Convertir fechas Excel (datetime → string)
            import datetime as _dt
            for k, v in valores.items():
                # Redondear floats que vienen de fórmulas Excel
                if isinstance(v, float) and v == int(v):
                    valores[k] = int(v)
                elif isinstance(v, float):
                    valores[k] = round(v, 2)
                if isinstance(v, (_dt.datetime, _dt.date)):
                    valores[k] = v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, _dt.datetime) else v.strftime("%Y-%m-%d")

            # Hash determinista de los valores actuales (excluye claves
            # autogeneradas/timestamps; se usa solo lo del Excel).
            row_payload = "|".join(
                f"{c}={'' if valores.get(c) is None else str(valores.get(c))}"
                for c in campos
            )
            row_hash = _hl.md5(row_payload.encode("utf-8")).hexdigest()

            ident_val = valores.get(ident_efectivo) if ident_efectivo else None
            cur_dict = conn.cursor(dictionary=True)
            existente = None
            if ident_val is not None and str(ident_val).strip() != "":
                try:
                    cur_dict.execute(
                        f"SELECT id, _row_hash FROM `{tabla_mysql}` WHERE `{ident_efectivo}` = %s LIMIT 1",
                        [ident_val]
                    )
                    existente = cur_dict.fetchone()
                except Exception as e:
                    n_err += 1
                    cur_dict.close()
                    continue
            cur_dict.close()

            # Solo enviamos columnas con valor real para que las NOT NULL
            # con default no rompan cuando el Excel tiene celda vacía.
            cols_present = [
                c for c in campos
                if valores.get(c) is not None and str(valores.get(c)).strip() != ""
            ]
            try:
                if existente is None:
                    cols_with_hash = cols_present + ["_row_hash"]
                    placeholders = ", ".join(["%s"] * len(cols_with_hash))
                    cols_str = ", ".join([f"`{c}`" for c in cols_with_hash])
                    sql = f"INSERT INTO `{tabla_mysql}` ({cols_str}) VALUES ({placeholders})"
                    cursor.execute(sql, [valores.get(c) for c in cols_present] + [row_hash])
                    n_nuevos += 1
                elif existente.get("_row_hash") == row_hash:
                    n_iguales += 1
                else:
                    if not cols_present:
                        n_iguales += 1
                    else:
                        set_cols = ", ".join([f"`{c}`=%s" for c in cols_present]) + ", `_row_hash`=%s"
                        sql = f"UPDATE `{tabla_mysql}` SET {set_cols} WHERE id=%s"
                        cursor.execute(
                            sql,
                            [valores.get(c) for c in cols_present] + [row_hash, existente["id"]]
                        )
                        n_upd += 1
            except Exception as _e:
                n_err += 1
                if n_err <= 1 and os.environ.get("KRAFTDO_DEBUG_IMPORT"):
                    print(f"      ⚠️  ({tabla_mysql}) {type(_e).__name__}: {str(_e)[:150]}")

        conn.commit()
        duracion_ms = int((_t.time() - t0) * 1000)
        try:
            cursor.execute(
                "INSERT INTO import_logs "
                "(empresa, alias_hoja, fecha_inicio, fecha_fin, "
                " nuevos, actualizados, sin_cambio, errores, duracion_ms, "
                " created_at, updated_at) VALUES (%s,%s,NOW(),NOW(),%s,%s,%s,%s,%s,NOW(),NOW())",
                [empresa, alias, n_nuevos, n_upd, n_iguales, n_err, duracion_ms]
            )
            conn.commit()
        except Exception:
            pass  # tabla import_logs ausente: silently skip
        print(
            f"  ✓ {tabla_mysql}: {n_iguales} sin cambio | {n_nuevos} nuevos | "
            f"{n_upd} actualizados | {n_err} errores ({duracion_ms}ms)"
        )

    cursor.close()
    conn.close()
    print("\n✅ Importación completada")

    # v25-fase5: si el proyecto generado tiene verificar_integridad.py,
    # correrlo y mostrar el resumen.
    verif = os.path.join(laravel_dir, "verificar_integridad.py")
    if os.path.isfile(verif):
        import subprocess as _sp
        print("\n🧪 Verificando integridad Excel ↔ BD...")
        r = _sp.run(["python3", verif], cwd=laravel_dir, capture_output=True, text=True)
        sys.stdout.write(r.stdout)
        if r.returncode != 0:
            print("  ⚠️  Hay discrepancias (no bloqueante).")

if __name__ == "__main__":
    empresa = sys.argv[1] if len(sys.argv) > 1 else "kraftdo_bd"
    laravel_dir = sys.argv[2] if len(sys.argv) > 2 else "/tmp/kraftdo_laravel_real"
    print(f"\n🔄 Importando {empresa} → {laravel_dir}\n")
    importar(empresa, laravel_dir)
