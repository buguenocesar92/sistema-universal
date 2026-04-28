# pyright: basic
# type: ignore
"""
KraftDo — generator.py
Lee empresas/{empresa}.json y genera:
  - database/migrations/*.php       (tablas en MySQL)
  - app/Models/*.php                 (modelos Eloquent)
  - app/Filament/Resources/*.php     (CRUD en Filament)
  - routes/api.php                   (API REST lista)
  - install.sh                       (script de instalación completa)

USO:
    python3 generator.py kraftdo           → genera todo
    python3 generator.py kraftdo --solo migraciones
    python3 generator.py kraftdo --preview → muestra sin escribir
"""

import os
import sys
import json
import re
import argparse
from datetime import datetime
from pathlib import Path

# Módulos propios
try:
    from relations import detectar_relaciones, relaciones_por_tabla, gen_foreign_keys_migration, gen_eloquent_relationships, gen_hasmany_relationships, gen_filament_select
    from formula_parser import analizar_excel_formulas, gen_accessors_php
    from consolidator import Consolidator
    from normalizer import analizar_excel_completo as _analizar_patrones
    RELACIONES_OK = True
except ImportError:
    RELACIONES_OK = False
    def _analizar_patrones(path): return {}

def _es_consolidado(cfg_hoja: dict) -> bool:
    return bool(cfg_hoja.get("consolidado"))

def _valores_tipo(cfg_hoja: dict) -> list:
    return cfg_hoja.get("valores_tipo") or [
        _slug(f) for f in cfg_hoja.get("fuentes", [])
    ]

def _slug(s: str) -> str:
    """Convierte a slug ASCII válido para SQL: elimina tildes, emojis y espacios."""
    import re, unicodedata
    # Normalizar unicode: descompone caracteres acentuados (á → a + ́)
    s = unicodedata.normalize("NFKD", str(s))
    # Eliminar caracteres no ASCII
    s = s.encode("ascii", "ignore").decode("ascii")
    # Reemplazar caracteres no alfanuméricos por guión bajo
    s = re.sub(r"[^a-z0-9_]", "_", s.lower().strip())
    # Eliminar guiones bajos duplicados y del inicio/fin
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "hoja"

# ── Tipos de columna: inferencia por nombre ─────────────────────────────────
# Nota: el orden importa. Las reglas más específicas van primero.
REGLAS_TIPO = [
    # Fechas
    (r"^fecha|^f_|_fecha$|_date$|_at$", "timestamp", "nullable"),
    (r"^updated_at$|^created_at$",       "timestamp", "nullable"),
    # Porcentajes y márgenes (deben ir ANTES que la regla monetaria, que es más amplia).
    # Solo decimal(5,4) cuando es realmente fracción 0..1. NO incluye 'descuento'
    # (descuento_faltas es monto), ni '^iva$' (iva en ventas es monto).
    (r"^margen$|^margen_pct$|_pct$|porcentaje", "decimal:5,4", "default:0"),
    # Montos y precios — incluye iva, descuento_faltas, iva_servicio
    (r"precio|monto|costo|total|saldo|anticipo|ganancia|ahorro|^iva|descuento|servicio", "decimal:15,2", "default:0"),
    # Cantidades enteras
    (r"^cantidad$|^stock$|^dias|^horas$|^gramos$|_count$|_qty$", "integer", "default:0"),
    # Estado (select)
    (r"^estado$|^status$|^tipo$|^categoria$|^canal$", "string", "nullable"),
    # Booleanos
    (r"^activo$|^active$|^enabled$|^is_", "boolean", "default:true"),
    # Email
    (r"correo|email", "string", "nullable"),
    # Teléfono / WhatsApp
    (r"telefono|whatsapp|phone|fono", "string:20", "nullable"),
    # URLs
    (r"^web$|^url$|^link$|_url$", "string", "nullable"),
    # Identificadores de referencia
    (r"^sku$|^numero$|^codigo$|^code$|^ref", "string:50", "nullable"),
    # Texto largo
    (r"obs|descripcion|notas|detalle|comment", "text", "nullable"),
    # Default: string
]

def inferir_tipo(nombre: str) -> tuple[str, str]:
    """Retorna (tipo_laravel, modificador) según el nombre del campo."""
    nombre_lower = nombre.lower()
    for patron, tipo, mod in REGLAS_TIPO:
        if re.search(patron, nombre_lower):
            return tipo, mod
    return "string", "nullable"


# ── Reglas de cálculo automático ────────────────────────────────────────────
# Cada regla define un campo destino, sus dependencias y la fórmula PHP.
# El observer aplica las reglas en orden (las posteriores pueden depender
# de campos ya calculados por reglas anteriores).
# Se aplica una regla solo si TODOS los campos (destino + deps) existen
# en la hoja, así nunca se rompe una tabla que no tenga el patrón.
REGLAS_CALCULO = [
    # Producto: costo total = insumo + mano de obra
    {"campo": "costo_total", "deps": ["costo_insumo", "hora_trabajo"],
     "formula": "(float)($model->costo_insumo ?? 0) + (float)($model->hora_trabajo ?? 0)",
     "descripcion": "Suma de costo de insumo + hora de trabajo"},

    # Producto: precio unitario = costo / (1 - margen 40%)
    {"campo": "precio_unit", "deps": ["costo_total"],
     "formula": "(float)($model->costo_total ?? 0) > 0 ? (int) round((float)$model->costo_total / 0.60) : 0",
     "descripcion": "Precio unitario calculado con margen 40%"},

    # Producto: precio mayorista = costo / (1 - margen 37%)
    {"campo": "precio_mayor", "deps": ["costo_total"],
     "formula": "(float)($model->costo_total ?? 0) > 0 ? (int) round((float)$model->costo_total / 0.63) : 0",
     "descripcion": "Precio mayorista calculado con margen 37%"},

    # Liquidación: valor del día = sueldo / días laborales
    {"campo": "valor_dia", "deps": ["sueldo_base", "dias_laborales"],
     "formula": "(float)($model->dias_laborales ?? 0) > 0 ? (int) round((float)$model->sueldo_base / (float)$model->dias_laborales) : 0",
     "descripcion": "Valor diario del trabajador"},

    # Liquidación: descuento por faltas
    {"campo": "descuento_faltas", "deps": ["valor_dia", "faltas"],
     "formula": "(int) round((float)($model->valor_dia ?? 0) * (float)($model->faltas ?? 0))",
     "descripcion": "Descuento por días faltados"},

    # Liquidación: monto a pagar
    {"campo": "a_pagar", "deps": ["sueldo_base", "descuento_faltas"],
     "formula": "(int) round((float)($model->sueldo_base ?? 0) - (float)($model->descuento_faltas ?? 0))",
     "descripcion": "Sueldo base menos descuentos"},

    # Liquidación: saldo pendiente = a pagar - quincena ya pagada
    {"campo": "saldo", "deps": ["a_pagar", "quincena_pagada"],
     "formula": "(int) round((float)($model->a_pagar ?? 0) - (float)($model->quincena_pagada ?? 0))",
     "descripcion": "Saldo pendiente de pago"},

    # Ventas: IVA calculado del neto
    {"campo": "iva", "deps": ["neto"],
     "formula": "(int) round((float)($model->neto ?? 0) * 0.19)",
     "descripcion": "IVA 19% calculado sobre el neto"},

    # Importaciones: IVA del servicio
    {"campo": "iva_servicio", "deps": ["total_neto"],
     "formula": "(int) round((float)($model->total_neto ?? 0) * 0.19)",
     "descripcion": "IVA 19% sobre el total neto"},

    # Promociones / ventas: total con IVA
    {"campo": "total", "deps": ["neto"],
     "formula": "(int) round((float)($model->neto ?? 0) * 1.19)",
     "descripcion": "Total con IVA incluido"},

    # Importaciones: total neto = costos componentes
    {"campo": "total_neto", "deps": ["costo_china", "embarcadero", "agente_aduana"],
     "formula": "(int) round((float)($model->costo_china ?? 0) + (float)($model->embarcadero ?? 0) + (float)($model->agente_aduana ?? 0))",
     "descripcion": "Suma de costos de importación"},

    # Stock: stock disponible = importaciones - ventas - promociones
    {"campo": "stock_disponible", "deps": ["importacion", "ventas"],
     "formula": "(int) ((int)($model->importacion ?? 0) - (int)($model->ventas ?? 0) - (int)($model->promociones ?? 0))",
     "descripcion": "Stock disponible = importado - vendido - promociones"},

    # Obras: total gastado = materiales + mano de obra + otros
    {"campo": "total_gastado", "deps": ["materiales", "mano_obra"],
     "formula": "(int) round((float)($model->materiales ?? 0) + (float)($model->mano_obra ?? 0) + (float)($model->otros ?? 0))",
     "descripcion": "Suma de gastos de la obra"},

    # Obras: resultado = cobrado - total gastado
    {"campo": "resultado", "deps": ["cobrado", "total_gastado"],
     "formula": "(int) round((float)($model->cobrado ?? 0) - (float)($model->total_gastado ?? 0))",
     "descripcion": "Resultado neto de la obra"},

    # Obras: margen porcentual
    {"campo": "margen", "deps": ["cobrado", "resultado"],
     "formula": "(float)($model->cobrado ?? 0) > 0 ? round((float)($model->resultado ?? 0) / (float)$model->cobrado, 4) : 0",
     "descripcion": "Margen como fracción del cobrado"},

    # Ventas: neto con descuento (si no hay descuento explícito, copia neto)
    {"campo": "neto_dsto", "deps": ["neto"],
     "formula": "(int) round((float)($model->neto ?? 0))",
     "descripcion": "Neto con descuento (default = neto)"},
]


SNAPSHOT_SYNONYMS = {
    # campo en el hijo → posible nombre en el padre cuando no coincide
    "producto": "nombre",
    "cliente":  "nombre",
}


def _resolver_accessor(campo: str, alias_hoja: str, hojas_cfg: dict,
                        relaciones: list) -> dict | None:
    """Dado un campo en el hijo, busca la FK cuyo padre lo contiene.

    Prioridad: si existe una FK cuyo padre singular coincide con `campo`
    (ej. campo='producto' → FK a tabla 'productos'), se prefiere esa
    relación incluso si otra también tiene una col con ese nombre vía
    sinónimo. Así `pedidos.producto` resuelve a productos.nombre y no
    a clientes.nombre.
    """
    tabla_origen = nombre_tabla(alias_hoja)
    rels_hoja = [r for r in relaciones if r["tabla_origen"] == tabla_origen]
    # Sort: las FK cuyo padre singular == campo van primero
    rels_hoja.sort(key=lambda r: 0 if r["tabla_destino"].rstrip("s") == campo else 1)

    for rel in rels_hoja:
        alias_padre = next(
            (a for a, h in hojas_cfg.items() if nombre_tabla(a) == rel["tabla_destino"]),
            None
        )
        if not alias_padre:
            continue
        cols_padre = hojas_cfg[alias_padre].get("columnas", {})
        for c in (campo, SNAPSHOT_SYNONYMS.get(campo)):
            if c and c in cols_padre:
                nombre_metodo = rel["campo_origen"].rstrip("s")
                if nombre_metodo == rel["campo_origen"]:
                    nombre_metodo += "Rel"
                return {
                    "rel_method":   nombre_metodo,
                    "modelo_padre": rel["modelo_destino"],
                    "campo_padre":  c,
                    "campo_origen": rel["campo_origen"],
                    "campo_destino": rel["campo_destino"],
                }
    return None


def _calculos_aplicables(cols: dict) -> list[dict]:
    """Devuelve las reglas de cálculo cuyo campo destino + deps están en cols."""
    nombres = set(cols.keys())
    aplicables = []
    for regla in REGLAS_CALCULO:
        if regla["campo"] not in nombres:
            continue
        if not all(d in nombres for d in regla["deps"]):
            continue
        aplicables.append(regla)
    return aplicables


# Conjuntos para clasificar campos en widgets/formularios
CAMPOS_MONEDA = {
    "monto", "total", "precio_unit", "precio_mayor", "neto", "neto_dsto",
    "cobrado", "sueldo_base", "a_pagar", "saldo", "costo_total", "costo_insumo",
    "total_neto", "total_gastado", "ganancia", "anticipo", "ahorro", "iva",
    "iva_servicio", "resultado", "materiales", "mano_obra", "costo_gym",
    "costo_nogales", "gastos_generales", "monto_neto", "costo_china",
    "embarcadero", "agente_aduana", "costo_stand", "descuento_faltas",
    "valor_dia", "quincena_pagada", "precio",
}

# Campos que típicamente NO son enums aunque sean strings cortos
NO_ENUM = {
    "nombre", "descripcion", "obs", "observaciones", "detalle", "comentario",
    "comment", "trabajador", "cliente", "producto", "sku",
    "codigo", "factura", "rut", "empresa", "evento", "lugar", "vehiculo",
    "concepto", "responsable", "contacto", "obra", "item", "variante",
    "ciudad", "whatsapp", "correo", "email", "telefono", "fecha", "url",
    "web", "sitio_web", "link", "ref_pedido", "proveedor", "modelo",
    "aplicaciones", "problema", "f_entrega",
}


def _detectar_enums_excel(excel_path: str, cfg: dict) -> dict:
    """Lee el Excel y detecta columnas string con 2..7 valores únicos como enum.

    Devuelve {alias: {campo: [val1, val2, ...]}}.
    Saltea identificadores y campos en NO_ENUM (nombres, textos abiertos, etc).
    """
    if not os.path.exists(excel_path):
        return {}
    try:
        import openpyxl
        from openpyxl.utils import column_index_from_string
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception:
        return {}

    resultado = {}
    for alias, hoja_cfg in cfg.get("hojas", {}).items():
        if hoja_cfg.get("tipo") not in ("registros", "catalogo"):
            continue
        nombre_hoja = hoja_cfg.get("nombre")
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        fila_inicio = hoja_cfg.get("fila_datos", 2)
        ident       = hoja_cfg.get("identificador")
        enums_hoja  = {}

        for campo, letra in hoja_cfg.get("columnas", {}).items():
            if campo == ident or campo in NO_ENUM or campo in CAMPOS_MONEDA:
                continue
            tipo, _ = inferir_tipo(campo)
            # Sólo strings (los enums numéricos son raros)
            if tipo != "string" and not tipo.startswith("string"):
                continue

            try:
                col_idx = column_index_from_string(letra)
            except Exception:
                continue

            valores = set()
            descartar = False
            try:
                fin = min(ws.max_row or fila_inicio, fila_inicio + 300)
                for row in ws.iter_rows(min_row=fila_inicio, max_row=fin,
                                         min_col=col_idx, max_col=col_idx,
                                         values_only=True):
                    v = row[0]
                    if v is None:
                        continue
                    s = str(v).strip()
                    if not s:
                        continue
                    if len(s) > 60:
                        descartar = True
                        break
                    valores.add(s)
                    if len(valores) > 8:
                        descartar = True
                        break
            except Exception:
                descartar = True
            if descartar:
                continue
            if 2 <= len(valores) <= 7:
                enums_hoja[campo] = sorted(valores)

        if enums_hoja:
            resultado[alias] = enums_hoja
    return resultado


def nombre_tabla(alias: str) -> str:
    """Convierte alias JSON a nombre de tabla Snake_case plural."""
    return alias.lower().replace("-", "_")


def nombre_modelo(alias: str) -> str:
    """Convierte alias JSON a nombre de clase PascalCase singular."""
    palabras = re.split(r'[_\-\s]+', alias)
    ult = palabras[-1].lower()
    if ult.endswith("ores"):   singular = ult[:-2]
    elif ult.endswith("ales"): singular = ult[:-2]
    elif ult.endswith("iones"): singular = ult[:-2]
    elif ult.endswith("entes"): singular = ult[:-1]
    elif ult.endswith("tes"):  singular = ult[:-1]
    elif ult.endswith("enes"): singular = ult[:-2]
    elif ult.endswith("res"):  singular = ult[:-2]
    elif ult.endswith("nes"):  singular = ult[:-1]
    elif ult.endswith("as"):   singular = ult[:-1]
    elif ult.endswith("os"):   singular = ult[:-1]
    elif ult.endswith("es") and len(ult) > 4: singular = ult[:-1]
    elif ult.endswith("s"):    singular = ult[:-1]
    else: singular = ult
    return "".join(w.capitalize() for w in palabras[:-1]) + singular.capitalize()


def nombre_migration(alias: str, idx: int) -> str:
    ts = f"2026_01_01_{idx:06d}"  # fecha fija para evitar migraciones duplicadas
    return f"{ts}_create_{nombre_tabla(alias)}_table"


# ═══════════════════════════════════════════════════════════════════════════════
# GENERADORES
# ═══════════════════════════════════════════════════════════════════════════════

def gen_migracion(alias: str, cfg_hoja: dict, idx: int) -> str:
    # ── Hoja consolidada ──────────────────────────────────────────────────
    if _es_consolidado(cfg_hoja) and RELACIONES_OK:
        try:
            fuentes    = cfg_hoja.get("fuentes", [])
            cols_raw   = {k: v for k, v in cfg_hoja.get("columnas", {}).items()
                          if k not in ("tipo",) and v != "[CALCULADO]"}
            tipos      = _valores_tipo(cfg_hoja)
            tabla_n    = nombre_tabla(alias)
            tiene_meta = bool(cfg_hoja.get("cols_especificas"))
            meta_col   = "            $table->json('metadata')->nullable();" if tiene_meta else ""
            tipos_enum = "', '".join(tipos)

            lineas = []
            for campo, letra in cols_raw.items():
                t, mod = inferir_tipo(campo)
                if t.startswith("decimal"):
                    p, s = t.split(":")[1].split(",")
                    unsigned = "->unsigned()" if campo in CAMPOS_MONEDA or any(x in campo.lower() for x in ("precio", "costo", "total", "cantidad", "stock", "unidades")) else ""
                    lineas.append(
                        "            $table->decimal('" + campo + "', " + p + ", " + s + ")" + unsigned + "->nullable();")
                elif t == "integer":
                    unsigned = "->unsigned()" if any(x in campo.lower() for x in ("cantidad", "stock", "unidades", "dias", "horas")) else ""
                    lineas.append("            $table->integer('" + campo + "')" + unsigned + "->nullable();")
                elif t == "text":
                    lineas.append("            $table->text('" + campo + "')->nullable();")
                elif t == "timestamp":
                    lineas.append("            $table->date('" + campo + "')->nullable();")
                else:
                    lineas.append("            $table->string('" + campo + "', 255)->nullable();")

            cols_str = "\n".join(lineas)

            return (
                "<?php\n\n"
                "use Illuminate\\Database\\Migrations\\Migration;\n"
                "use Illuminate\\Database\\Schema\\Blueprint;\n"
                "use Illuminate\\Support\\Facades\\Schema;\n\n"
                "// Tabla consolidada — fuentes: " + ", ".join(fuentes) + "\n\n"
                "return new class extends Migration\n"
                "{\n"
                "    public function up(): void\n"
                "    {\n"
                "        Schema::create('" + tabla_n + "', function (Blueprint $table) {\n"
                "            $table->id();\n"
                "            $table->enum('tipo', ['" + tipos_enum + "'])->index();\n"
                + cols_str + "\n"
                + ("\n" + meta_col if meta_col else "") + "\n"
                "            $table->string('_row_hash', 64)->nullable()->index();\n"
                "            $table->timestamps();\n"
                "        });\n"
                "    }\n\n"
                "    public function down(): void\n"
                "    {\n"
                "        Schema::dropIfExists('" + tabla_n + "');\n"
                "    }\n"
                "};\n"
            )
        except Exception:
            pass  # fallback a generación estándar
    # ─────────────────────────────────────────────────────────────────────


    tabla  = nombre_tabla(alias)
    clase  = "Create" + "".join(w.capitalize() for w in alias.split("_")) + "Table"
    cols   = cfg_hoja.get("columnas", {})
    estados = cfg_hoja.get("logica", {}).get("estados", [])

    lineas_col = []
    for campo, _ in cols.items():
        if campo in ("numero", "id"):
            continue  # lo maneja $table->id() o el campo numero como PK
        tipo, mod = inferir_tipo(campo)

        # Construir línea de columna
        if tipo.startswith("decimal"):
            partes = tipo.split(":")
            precision = partes[1] if len(partes) > 1 else "10,2"
            p, s = precision.split(",")
            linea = f"            $table->decimal('{campo}', {p}, {s})"
            if campo in CAMPOS_MONEDA or any(x in campo.lower() for x in ("precio", "costo", "total", "cantidad", "stock", "unidades")):
                linea += "->unsigned()"
        elif tipo.startswith("string:"):
            largo = tipo.split(":")[1]
            linea = f"            $table->string('{campo}', {largo})"
        elif tipo == "text":
            linea = f"            $table->text('{campo}')"
        elif tipo == "integer":
            linea = f"            $table->integer('{campo}')"
            if any(x in campo.lower() for x in ("cantidad", "stock", "unidades", "dias", "horas")):
                linea += "->unsigned()"
        elif tipo == "boolean":
            linea = f"            $table->boolean('{campo}')"
        elif tipo == "timestamp":
            linea = f"            $table->timestamp('{campo}')"
        else:
            linea = f"            $table->string('{campo}')"

        # Modificador
        if "nullable" in mod:
            linea += "->nullable()"
        if "default:" in mod:
            val = mod.split("default:")[1]
            if val in ("true", "false"):
                linea += f"->default({val})"
            else:
                linea += f"->default({val})"

        lineas_col.append(linea + ";")

    cols_str = "\n".join(lineas_col)

    # Índices automáticos — detecta fecha, estado, sku, id_, email, etc.
    indices_campos = _auto_indices(cols)
    indices = ""
    for campo in indices_campos:
        if campo not in ("numero", "id"):  # no indexar PKs
            indices += f"\n            $table->index('{campo}');"
    # Hash de fila para upsert diferencial del importer (v25-fase3)
    cols_str += "\n            $table->string('_row_hash', 64)->nullable()->index();"

    return f"""<?php

use Illuminate\\Database\\Migrations\\Migration;
use Illuminate\\Database\\Schema\\Blueprint;
use Illuminate\\Support\\Facades\\Schema;

return new class extends Migration
{{
    public function up(): void
    {{
        Schema::create('{tabla}', function (Blueprint $table) {{
            $table->id();
{cols_str}{indices}
            $table->timestamps();
        }});
    }}

    public function down(): void
    {{
        Schema::dropIfExists('{tabla}');
    }}
}};
"""


def gen_modelo(alias: str, cfg_hoja: dict, empresa_cfg: dict, relaciones=None, formulas=None) -> str:
    modelo = nombre_modelo(alias)
    tabla  = nombre_tabla(alias)
    cols   = cfg_hoja.get("columnas", {})

    # fillable: todos los campos excepto id/timestamps
    fillable = [c for c in cols.keys() if c not in ("id",)]
    fillable_str = ",\n        ".join(f"'{c}'" for c in fillable)

    # casts
    casts = []
    for campo in cols.keys():
        tipo, _ = inferir_tipo(campo)
        if tipo.startswith("decimal"):
            casts.append(f"'{campo}' => 'decimal:2'")
        elif tipo == "integer":
            casts.append(f"'{campo}' => 'integer'")
        elif tipo == "boolean":
            casts.append(f"'{campo}' => 'boolean'")
        elif tipo == "timestamp":
            casts.append(f"'{campo}' => 'datetime'")
    casts_str = ",\n        ".join(casts) if casts else ""
    casts_block = "\n        " + casts_str + "\n    " if casts_str else ""

    # scope de activos si tiene campo estado
    scope = ""
    if "estado" in cols:
        estados_activos = cfg_hoja.get("filtro_activos", {}).get("valores", ["Activo"])
        estados_str = ", ".join(f"'{e}'" for e in estados_activos)
        scope = f"""
    public function scopeActivos($query)
    {{
        return $query->whereIn('estado', [{estados_str}]);
    }}"""

    # Relaciones si están disponibles (belongsTo + hasMany)
    rels_str = ""
    if RELACIONES_OK and hasattr(gen_modelo, '_relaciones'):
        rels = gen_modelo._relaciones
        belongs = gen_eloquent_relationships(rels, tabla)
        hasmany = gen_hasmany_relationships(rels, tabla, {})
        # Deduplicar métodos por nombre de función
        todos_metodos = []
        nombres_vistos = set()
        for bloque in [belongs, hasmany]:
            if not bloque:
                continue
            for metodo in bloque.split("\n\n"):
                metodo = metodo.strip()
                if not metodo:
                    continue
                # Extraer nombre del método
                import re as _re2
                m = _re2.search(r"public function (\w+)\(", metodo)
                nombre = m.group(1) if m else metodo[:30]
                if nombre not in nombres_vistos:
                    nombres_vistos.add(nombre)
                    todos_metodos.append(metodo)
        if todos_metodos:
            rels_str = "\n\n" + "\n\n".join(todos_metodos)

    # Scopes por tipo si la hoja es consolidada
    if _es_consolidado(cfg_hoja):
        tipos = _valores_tipo(cfg_hoja)
        scopes = ""
        for tipo in tipos:
            nombre_scope = "".join(w.capitalize() for w in tipo.split("_"))
            scopes += (
                "\n\n"
                "    public function scope" + nombre_scope + "($query)\n"
                "    {\n"
                "        return $query->where('tipo', '" + tipo + "');\n"
                "    }"
            )
        if scopes:
            rels_str += scopes

    # Accessors de fórmulas si están disponibles
    accessors_str = ""
    if RELACIONES_OK and hasattr(gen_modelo, '_formulas') and alias in gen_modelo._formulas:
        for conv in gen_modelo._formulas[alias]:
            if conv.get("php"):
                accessors_str += "\n\n" + gen_accessors_php(conv["campo"], conv)

    # Accessors "vista lateral" (VLOOKUP triviales) desde campos_accessor
    appends = []
    campos_accessor = cfg_hoja.get("campos_accessor", []) or []
    hojas_cfg = empresa_cfg.get("hojas", {}) if empresa_cfg else {}
    rels_disp = getattr(gen_modelo, '_relaciones', []) or []
    for nombre in campos_accessor:
        info = _resolver_accessor(nombre, alias, hojas_cfg, rels_disp)
        if not info:
            continue
        # Nombre del método get + StudlyCase
        studly = "".join(w.capitalize() for w in re.split(r'[_\-]+', nombre))
        # Usa la forma método ($this->relation()) para no chocar con
        # la columna del mismo nombre. firstOrNew evita N+1 y nullsafes.
        accessors_str += (
            "\n\n"
            f"    public function get{studly}Attribute()\n"
            "    {\n"
            f"        return $this->{info['rel_method']}()->first()?->{info['campo_padre']};\n"
            "    }"
        )
        appends.append(nombre)
    appends_str = ""
    if appends:
        appends_str = (
            "\n\n    protected $appends = ["
            + ", ".join(f"'{a}'" for a in appends)
            + "];"
        )

    # Método estático recalcularModelo para hojas tipo "agregado"
    metodo_agregado_str = ""
    if cfg_hoja.get("tipo") == "agregado" and empresa_cfg:
        metodo_agregado_str = "\n\n" + gen_metodo_recalcular_modelo(
            alias, cfg_hoja, empresa_cfg
        )

    # ¿Tiene observer? Cuando hay reglas de cálculo aplicables, formato_id
    # para auto-numeración, snapshot_at_create, o cascadas registradas.
    formato_id = cfg_hoja.get("formato_id")
    ident      = cfg_hoja.get("identificador")
    tiene_observer = (
        bool(_calculos_aplicables(cols))
        or bool(formato_id and ident and ident in cols)
        or bool(cfg_hoja.get("snapshot_at_create"))
        or bool(cfg_hoja.get("_cascadas"))
        or bool(cfg_hoja.get("auto_aggregate"))
    )
    observer_use   = ""
    observer_attr  = ""
    if tiene_observer:
        observer_use  = (
            "use Illuminate\\Database\\Eloquent\\Attributes\\ObservedBy;\n"
            f"use App\\Observers\\{modelo}Observer;\n"
        )
        observer_attr = f"#[ObservedBy([{modelo}Observer::class])]\n"

    return f"""<?php

namespace App\\Models;

use Illuminate\\Database\\Eloquent\\Model;
use Illuminate\\Database\\Eloquent\\Factories\\HasFactory;
{observer_use}
{observer_attr}class {modelo} extends Model
{{
    use HasFactory;

    protected $table = '{tabla}';

    protected $fillable = [
        {fillable_str},
    ];

    protected $casts = [{casts_block}];{appends_str}{scope}{rels_str}{accessors_str}{metodo_agregado_str}
}}
"""


def _resolver_snapshot(item, alias_hoja: str, hojas_cfg: dict,
                        relaciones: list) -> dict | None:
    """Resuelve una entrada de snapshot_at_create.

    Acepta:
      - string: campo a copiar; resuelve el padre por _resolver_accessor.
      - dict {campo, desde?, campo_padre?, fn?, base?}: explicit override.

    Devuelve dict normalizado con todos los datos para emitir PHP, o None.
    """
    if isinstance(item, str):
        info = _resolver_accessor(item, alias_hoja, hojas_cfg, relaciones)
        if not info:
            return None
        return {
            "campo":         item,
            "desde":         info["campo_origen"],
            "campo_destino": info["campo_destino"],
            "modelo_padre":  info["modelo_padre"],
            "campo_padre":   info["campo_padre"],
            "fn":            None,
        }
    if isinstance(item, dict):
        # Resolver el padre desde el campo `desde` (FK del hijo)
        tabla_origen = nombre_tabla(alias_hoja)
        rel = next(
            (r for r in relaciones
             if r["tabla_origen"] == tabla_origen
             and r["campo_origen"] == item.get("desde")),
            None
        )
        if not rel:
            return None
        return {
            "campo":         item.get("campo"),
            "desde":         rel["campo_origen"],
            "campo_destino": rel["campo_destino"],
            "modelo_padre":  rel["modelo_destino"],
            "campo_padre":   item.get("campo_padre", item.get("campo")),
            "fn":            item.get("fn"),
            "base":          item.get("base"),
        }
    return None


def gen_observer(alias: str, cfg_hoja: dict, empresa_cfg: dict | None = None) -> str | None:
    """Genera un Observer PHP que recalcula campos al guardar.

    Devuelve el contenido PHP, o None si la hoja no tiene ni reglas
    aplicables ni formato_id.
    Usa $model SIN backslash en typehint (regla CLAUDE.md).

    Si la hoja define `formato_id` (ej "KDO-{:03d}"), el observer
    asigna el siguiente identificador disponible al crear (creating).
    """
    cols       = cfg_hoja.get("columnas", {})
    aplicables = _calculos_aplicables(cols)
    formato_id = cfg_hoja.get("formato_id")
    ident      = cfg_hoja.get("identificador")
    auto_id    = bool(formato_id and ident and ident in cols)

    # Snapshot at create: campos del padre que se copian en creating()
    snapshots = []
    snap_cfg = cfg_hoja.get("snapshot_at_create", []) or []
    if snap_cfg:
        rels_disp = getattr(gen_modelo, '_relaciones', []) or []
        hojas_cfg = empresa_cfg.get("hojas", {}) if empresa_cfg else {}
        for item in snap_cfg:
            resolved = _resolver_snapshot(item, alias, hojas_cfg, rels_disp)
            if resolved and resolved["campo"] in cols:
                snapshots.append(resolved)

    # Cascada agregada: cuando esta hoja es fuente de otra hoja tipo "agregado",
    # disparar Stock::recalcularModelo() en saved/deleted.
    cascadas = cfg_hoja.get("_cascadas", []) or []

    # Auto-aggregate: campos que se rellenan desde queries a otras tablas.
    # Se ejecuta ANTES de recalcular() para que los campos derivados las usen.
    auto_agg = cfg_hoja.get("auto_aggregate", []) or []

    if not aplicables and not auto_id and not snapshots and not cascadas and not auto_agg:
        return None

    modelo = nombre_modelo(alias)

    bloques = []
    for regla in aplicables:
        comentario = regla.get("descripcion", "")
        bloques.append(
            f"        // {comentario}\n"
            f"        $model->{regla['campo']} = {regla['formula']};"
        )
    body = "\n\n".join(bloques) if bloques else "        // (sin reglas de cálculo)"

    # Bloque de auto-numeración: en creating, si el identificador está vacío,
    # buscar el último ID con el patrón y emitir el siguiente.
    auto_id_block = ""
    if auto_id:
        # Convertir patrón Python "{:03d}" a regex y formato sprintf-equivalente
        # Soportamos formatos tipo "PREFIJO-{:0Nd}".
        m = re.match(r"^(.*?)\{:0?(\d+)d\}(.*)$", formato_id)
        if m:
            prefijo, ancho, sufijo = m.group(1), int(m.group(2)), m.group(3)
            php_pattern = "/^" + re.escape(prefijo) + r"(\d+)" + re.escape(sufijo) + "$/"
            php_format  = prefijo + "%0" + str(ancho) + "d" + sufijo
            auto_id_block = (
                "\n"
                f"    public function creatingAutoId({modelo} $model): void\n"
                "    {\n"
                f"        if (!empty($model->{ident})) {{\n"
                "            return;\n"
                "        }\n"
                f"        $ultimo = {modelo}::query()\n"
                f"            ->where('{ident}', 'like', '" + prefijo.replace("'", "\\'") + "%')\n"
                f"            ->pluck('{ident}')\n"
                f"            ->map(fn ($v) => preg_match('" + php_pattern + "', (string) $v, $m) ? (int) $m[1] : 0)\n"
                "            ->max() ?? 0;\n"
                f"        $model->{ident} = sprintf('" + php_format + "', $ultimo + 1);\n"
                "    }\n"
            )

    # Bloque de snapshot creating-only
    snapshot_block = ""
    if snapshots:
        # Agrupar por (desde, modelo_padre, campo_destino) para hacer un solo lookup
        grupos = {}
        for s in snapshots:
            k = (s["desde"], s["modelo_padre"], s["campo_destino"])
            grupos.setdefault(k, []).append(s)

        partes = []
        for (desde, modelo_padre, campo_destino), items in grupos.items():
            asignaciones = []
            for s in items:
                campo = s["campo"]
                if s.get("fn") == "fecha_plus_dias":
                    base = s.get("base", "fecha")
                    asignaciones.append(
                        f"            if (empty($model->{campo}) && !empty($model->{base})) {{\n"
                        f"                $model->{campo} = \\Carbon\\Carbon::parse($model->{base})\n"
                        f"                    ->addDays((int) ($padre->{s['campo_padre']} ?? 0));\n"
                        f"            }}"
                    )
                else:
                    asignaciones.append(
                        f"            if (empty($model->{campo})) {{\n"
                        f"                $model->{campo} = $padre->{s['campo_padre']};\n"
                        f"            }}"
                    )
            partes.append(
                f"        if (!empty($model->{desde})) {{\n"
                f"            $padre = \\App\\Models\\{modelo_padre}::query()\n"
                f"                ->where('{campo_destino}', $model->{desde})\n"
                "                ->first();\n"
                "            if ($padre) {\n"
                + "\n".join(asignaciones) + "\n"
                "            }\n"
                "        }"
            )
        snapshot_block = (
            "\n"
            f"    public function creatingSnapshot({modelo} $model): void\n"
            "    {\n"
            + "\n\n".join(partes) + "\n"
            "    }\n"
        )

    # Bloque auto_aggregate: queries a otras tablas para llenar campos.
    aggregate_block = ""
    if auto_agg:
        partes = []
        for agg in auto_agg:
            campo  = agg["campo"]
            modelo_dest = agg["modelo"]
            wm = agg.get("where_match", [])
            we = agg.get("where_extra", [])
            fn = agg.get("fn", "count")

            wheres_php = ""
            for c in wm:
                wheres_php += f"->where('{c}', $model->{c})"
            for w in we:
                if isinstance(w, list) and len(w) >= 2:
                    val = w[1]
                    if isinstance(val, (int, float)):
                        val_php = str(val)
                    else:
                        val_php = "'" + str(val).replace("'", "\\'") + "'"
                    wheres_php += f"->where('{w[0]}', {val_php})"

            if fn == "count":
                expr = f"\\App\\Models\\{modelo_dest}::query(){wheres_php}->count()"
            elif fn.startswith("sum:"):
                campo_v = fn.split(":", 1)[1]
                expr = f"(float) \\App\\Models\\{modelo_dest}::query(){wheres_php}->sum('{campo_v}')"
            elif fn.startswith("avg:"):
                campo_v = fn.split(":", 1)[1]
                expr = f"(float) \\App\\Models\\{modelo_dest}::query(){wheres_php}->avg('{campo_v}')"
            else:
                expr = "0"

            partes.append(f"        $model->{campo} = {expr};")

        aggregate_block = (
            "\n"
            f"    protected function aggregate({modelo} $model): void\n"
            "    {\n"
            + "\n".join(partes) + "\n"
            "    }\n"
        )

    # Bloque cascada: en saved/deleted dispara recálculo del agregado destino.
    cascada_block = ""
    cascada_hooks = []
    if cascadas:
        for c in cascadas:
            modelo_destino = c["modelo_destino"]
            campo_grupo    = c["campo_grupo"]  # campo en este modelo (origen)
            cascada_hooks.append(
                f"        if (!empty($model->{campo_grupo})) {{\n"
                f"            \\App\\Models\\{modelo_destino}::recalcularModelo($model->{campo_grupo});\n"
                "        }"
            )
        cascada_block = (
            "\n"
            f"    public function saved({modelo} $model): void\n"
            "    {\n"
            + "\n".join(cascada_hooks) + "\n"
            "    }\n\n"
            f"    public function deleted({modelo} $model): void\n"
            "    {\n"
            + "\n".join(cascada_hooks) + "\n"
            "    }\n"
        )

    creating_calls = []
    if auto_id_block:
        creating_calls.append("        $this->creatingAutoId($model);")
    if snapshot_block:
        creating_calls.append("        $this->creatingSnapshot($model);")
    if aggregate_block:
        creating_calls.append("        $this->aggregate($model);")
    if aplicables:
        creating_calls.append("        $this->recalcular($model);")
    if not creating_calls:
        creating_calls = ["        // (nada que hacer al crear)"]
    creating_body = "\n".join(creating_calls)

    updating_calls = []
    if aggregate_block:
        updating_calls.append("        $this->aggregate($model);")
    if aplicables:
        updating_calls.append("        $this->recalcular($model);")
    if not updating_calls:
        updating_calls = ["        // (sin recálculo en update)"]
    updating_body = "\n".join(updating_calls)

    return (
        "<?php\n\n"
        "namespace App\\Observers;\n\n"
        f"use App\\Models\\{modelo};\n\n"
        "/**\n"
        f" * Observer auto-generado para {modelo}.\n"
        " * Recalcula campos derivados antes de guardar.\n"
        " */\n"
        f"class {modelo}Observer\n"
        "{\n"
        f"    public function creating({modelo} $model): void\n"
        "    {\n"
        + creating_body + "\n"
        "    }\n\n"
        f"    public function updating({modelo} $model): void\n"
        "    {\n"
        + updating_body + "\n"
        "    }\n\n"
        f"    protected function recalcular({modelo} $model): void\n"
        "    {\n"
        + body + "\n"
        "    }\n"
        + auto_id_block
        + snapshot_block
        + aggregate_block
        + cascada_block
        + "}\n"
    )


def gen_filament_resource(alias: str, cfg_hoja: dict, empresa_cfg: dict,
                           relaciones=None, enums=None) -> str:
    """Genera un Filament Resource completo con ExcelExport incluido.

    enums: {campo: [val1, val2, ...]} para esta hoja (si se detectaron en Excel).
    relaciones: lista global de relaciones detectadas por relations.detectar_relaciones.
    """
    modelo   = nombre_modelo(alias)
    resource = modelo + "Resource"
    cols     = list(cfg_hoja.get("columnas", {}).keys())
    tipo     = cfg_hoja.get("tipo", "registros")
    estados  = cfg_hoja.get("logica", {}).get("estados", [])
    enums    = enums or {}
    tabla_n  = nombre_tabla(alias)
    rels_hoja = [r for r in (relaciones or []) if r["tabla_origen"] == tabla_n]
    rels_por_campo = {r["campo_origen"]: r for r in rels_hoja}

    ns = "\\App\\Filament\\Resources\\"
    ns_models = "\\App\\Models\\"

    # Campos del formulario
    # Campos que nunca van en el formulario (PKs y los calculados por observer)
    cols_dict = cfg_hoja.get("columnas", {})
    calc_fields = {r["campo"] for r in _calculos_aplicables(cols_dict)}
    CAMPOS_AUTO = {"id", "n_pedido", "created_at", "updated_at"} | calc_fields

    form_fields = []
    for campo in cols[:10]:  # max 10 campos visibles
        if campo in CAMPOS_AUTO:
            continue
        t, mod = inferir_tipo(campo)
        label  = campo.replace("_", " ").capitalize()
        req    = "->required()" if "nullable" not in mod else ""

        # Prioridad 1: estado con valores definidos en el JSON
        if campo == "estado" and estados:
            opts = "\n".join(
                "                    '" + _slug(e) + "' => '" + e + "',"
                for e in estados
            )
            form_fields.append(
                "            Forms\\Components\\Select::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->options([\n"
                + opts + "\n"
                "                ])" + req + ","
            )
        # Prioridad 2: relación detectada → Select con opciones del modelo
        elif campo in rels_por_campo:
            rel = rels_por_campo[campo]
            modelo_dest = rel["modelo_destino"]
            campo_dest  = rel["campo_destino"]
            form_fields.append(
                "            Forms\\Components\\Select::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->options(fn() => \\App\\Models\\" + modelo_dest
                + "::query()->pluck('" + campo_dest + "', '" + campo_dest + "')->filter()->toArray())\n"
                "                ->searchable()\n"
                "                ->preload()"
                + req + ","
            )
        # Prioridad 3: enum auto-detectado en Excel
        elif campo in enums:
            opts_lines = "\n".join(
                "                    '" + v.replace("'", "\\'") + "' => '" + v.replace("'", "\\'") + "',"
                for v in enums[campo]
            )
            form_fields.append(
                "            Forms\\Components\\Select::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->options([\n"
                + opts_lines + "\n"
                "                ])" + req + ","
            )
        # Prioridad alta: campo monetario → numeric con prefix $
        elif campo in CAMPOS_MONEDA:
            form_fields.append(
                "            Forms\\Components\\TextInput::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->numeric()\n"
                "                ->minValue(0)\n"
                "                ->prefix('$')" + req + ","
            )
        elif t == "text":
            form_fields.append(
                "            Forms\\Components\\Textarea::make('" + campo + "')\n"
                "                ->label('" + label + "')" + req + ","
            )
        elif t.startswith("decimal") or t == "integer":
            form_fields.append(
                "            Forms\\Components\\TextInput::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->numeric()\n"
                "                ->minValue(0)" + req + ","
            )
        elif t == "timestamp":
            form_fields.append(
                "            Forms\\Components\\DatePicker::make('" + campo + "')\n"
                "                ->label('" + label + "')\n"
                "                ->displayFormat('d/m/Y')\n"
                "                ->native(false)" + req + ","
            )
        else:
            form_fields.append(
                "            Forms\\Components\\TextInput::make('" + campo + "')\n"
                "                ->label('" + label + "')" + req + ","
            )

    # Placeholder readonly para los campos calculados por el observer
    placeholders = []
    for campo in calc_fields:
        if campo not in cols_dict:
            continue
        label = campo.replace("_", " ").capitalize()
        if campo in CAMPOS_MONEDA:
            content_expr = (
                "$record ? '$' . number_format((float) $record->" + campo
                + ", 0, ',', '.') : '— se calcula al guardar —'"
            )
        else:
            content_expr = "$record?->" + campo + " ?? '— se calcula al guardar —'"
        placeholders.append(
            "            Forms\\Components\\Placeholder::make('" + campo + "')\n"
            "                ->label('" + label + " (auto)')\n"
            "                ->content(fn ($record) => " + content_expr + "),"
        )
    placeholders_str = ("\n" + "\n".join(placeholders)) if placeholders else ""

    form_str  = "\n".join(form_fields)

    # Columnas de la tabla
    table_cols = []
    for campo in cols[:8]:
        t, _ = inferir_tipo(campo)
        label_c = campo.replace('_', ' ').capitalize()
        # Monetario → money() con formato CLP
        if campo in CAMPOS_MONEDA:
            table_cols.append(
                "                Tables\\Columns\\TextColumn::make('" + campo + "')\n"
                "                    ->label('" + label_c + "')\n"
                "                    ->money('clp', divideBy: 1, locale: 'es_CL')\n"
                "                    ->sortable(),"
            )
        # Fecha → format d/m/Y
        elif t == "timestamp":
            table_cols.append(
                "                Tables\\Columns\\TextColumn::make('" + campo + "')\n"
                "                    ->label('" + label_c + "')\n"
                "                    ->date('d/m/Y')\n"
                "                    ->sortable(),"
            )
        # Numérico no monetario
        elif t.startswith("decimal") or t == "integer":
            table_cols.append(
                "                Tables\\Columns\\TextColumn::make('" + campo + "')\n"
                "                    ->label('" + label_c + "')\n"
                "                    ->numeric(thousandsSeparator: '.')\n"
                "                    ->sortable()->searchable(),"
            )
        # Badge para enums detectados o estado
        elif campo in enums or (campo == "estado" and estados):
            table_cols.append(
                "                Tables\\Columns\\TextColumn::make('" + campo + "')\n"
                "                    ->label('" + label_c + "')\n"
                "                    ->badge()\n"
                "                    ->sortable()->searchable(),"
            )
        else:
            table_cols.append(
                "                Tables\\Columns\\TextColumn::make('" + campo + "')\n"
                "                    ->label('" + label_c + "')\n"
                "                    ->sortable()->searchable(),"
            )

    table_str = "\n".join(table_cols)

    # Filtros: SelectFilter por estado/enums + DateRangeFilter por fechas
    filters_lines = []
    if "estado" in cols:
        filters_lines.append(
            "                Tables\\Filters\\SelectFilter::make('estado')\n"
            "                    ->options(fn() => " + ns_models + modelo + "::distinct()->pluck('estado', 'estado')->filter()->toArray()),"
        )
    for campo, valores in enums.items():
        if campo == "estado":
            continue
        opts = ", ".join("'" + v.replace("'", "\\'") + "' => '" + v.replace("'", "\\'") + "'" for v in valores)
        filters_lines.append(
            "                Tables\\Filters\\SelectFilter::make('" + campo + "')\n"
            "                    ->label('" + campo.replace('_', ' ').capitalize() + "')\n"
            "                    ->options([" + opts + "]),"
        )
    # DateRangeFilter (un Filter con dos DatePicker)
    for campo in cols:
        t, _ = inferir_tipo(campo)
        if t != "timestamp":
            continue
        filters_lines.append(
            "                Tables\\Filters\\Filter::make('" + campo + "_range')\n"
            "                    ->label('" + campo.replace('_', ' ').capitalize() + " (rango)')\n"
            "                    ->schema([\n"
            "                        Forms\\Components\\DatePicker::make('desde')->native(false),\n"
            "                        Forms\\Components\\DatePicker::make('hasta')->native(false),\n"
            "                    ])\n"
            "                    ->query(fn ($query, array $data) => $query\n"
            "                        ->when($data['desde'] ?? null, fn ($q, $d) => $q->whereDate('" + campo + "', '>=', $d))\n"
            "                        ->when($data['hasta'] ?? null, fn ($q, $d) => $q->whereDate('" + campo + "', '<=', $d))\n"
            "                    ),"
        )
    filters_str = ("\n" + "\n".join(filters_lines)) if filters_lines else ""

    # Global search: primeros 3 campos string no técnicos
    search_attrs = []
    for campo in cols:
        if len(search_attrs) >= 3:
            break
        if campo in ("id", "n_pedido"):
            continue
        if campo in CAMPOS_MONEDA or campo in calc_fields:
            continue
        t, _ = inferir_tipo(campo)
        if t == "string" or t.startswith("string"):
            search_attrs.append(campo)
    global_search = ""
    if search_attrs:
        attrs = ", ".join("'" + c + "'" for c in search_attrs)
        global_search = (
            "    public static function getGloballySearchableAttributes(): array\n"
            "    {\n"
            f"        return [{attrs}];\n"
            "    }\n\n"
        )

    # ExportAction desactivado — requiere pxlrbt/filament-excel
    export_action = ""

    # Relaciones como Select en el formulario
    rel_fields = ""  # desactivado — relaciones en formularios requieren modelos con relationship() definido
    if False and RELACIONES_OK and relaciones:
        from relations import gen_filament_select
        tabla_n = nombre_tabla(alias)
        rels_hoja = [r for r in relaciones if r["tabla_origen"] == tabla_n]
        for rel in rels_hoja[:3]:
            campo = rel["campo_origen"]
            if campo not in cols:
                continue
            rel_fields += (
                "\n            Forms\\Components\\Select::make('" + campo + "')\n"
                "                ->label('" + campo.replace('_',' ').capitalize() + "')\n"
                "                ->relationship('" + campo.rstrip('s') + "', '" + rel['campo_destino'] + "')\n"
                "                ->searchable()->preload()->nullable(),"
            )

    return (
        "<?php\n\n"
        "namespace App\\Filament\\Resources;\n\n"
        "use App\\Filament\\Resources\\" + resource + "\\Pages;\n"
        "use App\\Models\\" + modelo + ";\n"
        "use Filament\\Forms;\n"
        "use Filament\\Schemas\\Schema;\n"
        "use Filament\\Resources\\Resource;\n"
        "use Filament\\Tables;\n"
        "use Filament\\Tables\\Table;\n\n"
        "/**\n"
        " * Para habilitar el historial de actividades de Spatie, puedes agregar:\n"
        " * use Rmswv\\FilamentActivitylog\\Extensions\\Loggable; al Resource\n"
        " * o implementar un Relation Manager para activities.\n"
        " */\n"
        "class " + resource + " extends Resource\n"
        "{\n"
        "    protected static ?string $model = " + modelo + "::class;\n"
        "    protected static \\BackedEnum|string|null $navigationIcon = 'heroicon-o-table-cells';\n"
        "    protected static ?string $navigationLabel = '" + alias.replace('_',' ').title() + "';\n\n"
        "    protected static ?string $pluralModelLabel = '" + alias.replace('_',' ').title() + "';\n\n"
        "    public static function form(Schema $schema): Schema\n"
        "    {\n"
        "        return $schema->components([\n"
        + form_str + rel_fields + placeholders_str + "\n"
        "        ]);\n"
        "    }\n\n"
        "    public static function table(Table $table): Table\n"
        "    {\n"
        "        return $table\n"
        ""  # headerActions removido
        "            ->columns([\n"
        + table_str + "\n"
        "            ])\n"
        "            ->filters([" + filters_str + "\n"
        "            ])\n"
        "            ->actions([\n"
        "                \\Filament\\Actions\\EditAction::make(),\n"
        "                \\Filament\\Actions\\DeleteAction::make(),\n"
        "            ])\n"
        "            ->bulkActions([\n"
        "                \\Filament\\Actions\\BulkActionGroup::make([\n"
        "                    \\Filament\\Actions\\DeleteBulkAction::make(),\n"
        "                ]),\n"
        "            ]);\n"
        "    }\n\n"
        + global_search +
        "    public static function getPages(): array\n"
        "    {\n"
        "        return [\n"
        "            'index'  => Pages\\List" + modelo + "s::route('/'),\n"
        "            'create' => Pages\\Create" + modelo + "::route('/create'),\n"
        "            'edit'   => Pages\\Edit" + modelo + "::route('/{record}/edit'),\n"
        "        ];\n"
        "    }\n"
        "}\n"
    )


def gen_form_request(alias: str, cfg_hoja: dict, enums: dict | None = None) -> str:
    """Genera Laravel FormRequest con reglas de validación desde el JSON.

    Mejoras v19:
      - Identificador → required (no nullable).
      - Campos calculados por observer → no se validan (el observer los rellena).
      - Email con max:255.
      - `in:` para enum cuando hay estados definidos.
    Mejoras v20:
      - Enum auto-detectado del Excel → `in:val1,val2,...`
    """
    modelo_n  = nombre_modelo(alias)
    cols      = cfg_hoja.get("columnas", {})
    estados   = cfg_hoja.get("logica", {}).get("estados", [])
    identif   = cfg_hoja.get("identificador")
    calc      = {r["campo"] for r in _calculos_aplicables(cols)}
    enums     = enums or {}

    reglas = []
    for campo in cols.keys():
        if campo in ("id", "numero"):
            continue
        if campo in calc:
            continue  # los rellena el observer
        tipo, mod = inferir_tipo(campo)
        regla_partes = []

        es_identif = (campo == identif)
        if es_identif or "nullable" not in mod:
            regla_partes.append("required")
        else:
            regla_partes.append("nullable")

        if tipo.startswith("decimal") or tipo == "integer":
            regla_partes.append("numeric")
            regla_partes.append("min:0")
        elif tipo == "timestamp":
            regla_partes.append("date")
        elif tipo == "boolean":
            regla_partes.append("boolean")
        elif campo in ("correo", "email"):
            regla_partes.append("email")
            regla_partes.append("max:255")
        elif campo == "estado" and estados:
            vals = ",".join(estados)
            regla_partes.append(f"in:{vals}")
        elif campo in enums:
            # Comas dentro de valores se escapan (Laravel separa con comas)
            vals = ",".join(v.replace(",", "\\,") for v in enums[campo])
            regla_partes.append(f"in:{vals}")
        else:
            regla_partes.append("string")
            regla_partes.append("max:255")

        reglas.append(f"            '{campo}' => '{"|".join(regla_partes)}',")

    reglas_str = "\n".join(reglas)

    return (
        "<?php\n\n"
        "namespace App\\Http\\Requests;\n\n"
        "use Illuminate\\Foundation\\Http\\FormRequest;\n\n"
        "class " + modelo_n + "Request extends FormRequest\n"
        "{\n"
        "    public function authorize(): bool\n"
        "    {\n"
        "        return true;\n"
        "    }\n\n"
        "    public function rules(): array\n"
        "    {\n"
        "        return [\n"
        + reglas_str + "\n"
        "        ];\n"
        "    }\n"
        "}\n"
    )


def gen_seeder(alias: str, cfg_hoja: dict) -> str:
    """Genera DatabaseSeeder con datos de ejemplo."""
    modelo_n = nombre_modelo(alias)
    cols    = cfg_hoja.get("columnas", {})
    estados = cfg_hoja.get("logica", {}).get("estados", ["Activo"])

    fakes = []
    for campo in list(cols.keys())[:8]:
        if campo in ("id", "numero"):
            continue
        tipo, _ = inferir_tipo(campo)
        n = campo.lower()
        if "nombre" in n or "producto" in n or "descripcion" in n:
            fakes.append(f"            '{campo}' => fake()->words(3, true),")
        elif "cliente" in n:
            fakes.append(f"            '{campo}' => fake()->name(),")
        elif "email" in n or "correo" in n:
            fakes.append(f"            '{campo}' => fake()->email(),")
        elif "telefono" in n or "whatsapp" in n:
            fakes.append(f"            '{campo}' => fake()->phoneNumber(),")
        elif "sku" in n or "codigo" in n:
            fakes.append(f"            '{campo}' => strtoupper(fake()->lexify('???##')),")
        elif "estado" in n:
            opts = "', '".join(estados)
            fakes.append(f"            '{campo}' => fake()->randomElement(['{opts}']),")
        elif tipo.startswith("decimal") or tipo == "integer":
            fakes.append(f"            '{campo}' => fake()->numberBetween(1000, 100000),")
        elif "fecha" in n:
            fakes.append(f"            '{campo}' => fake()->dateTimeBetween('-1 year', 'now'),")
        else:
            fakes.append(f"            '{campo}' => fake()->word(),")

    fakes_str = "\n".join(fakes)

    return (
        "<?php\n\n"
        "namespace Database\\Seeders;\n\n"
        "use App\\Models\\" + modelo_n + ";\n"
        "use Illuminate\\Database\\Seeder;\n\n"
        "class " + modelo_n + "Seeder extends Seeder\n"
        "{\n"
        "    public function run(): void\n"
        "    {\n"
        "        " + modelo_n + "::factory(10)->create();\n"
        "        // O datos de ejemplo fijos:\n"
        "        // " + modelo_n + "::create([\n"
        + fakes_str + "\n"
        "        // ]);\n"
        "    }\n"
        "}\n"
    )


def gen_roles_seeder(cfg: dict) -> str:
    """Genera un Seeder para Roles y Permisos de Spatie."""
    roles = cfg.get("roles", {})
    if not roles:
        return ""

    perms_set = set()
    for p_list in roles.values():
        for p in p_list:
            perms_set.add(p)

    perms_str = "\n".join([f"        Permission::create(['name' => '{p}']);" for p in sorted(list(perms_set))])

    roles_blocks = []
    for role, p_list in roles.items():
        roles_blocks.append(f"        $role = Role::create(['name' => '{role}']);")
        if p_list:
            p_list_str = ", ".join([f"'{p}'" for p in p_list])
            roles_blocks.append(f"        $role->givePermissionTo([{p_list_str}]);")

    roles_str = "\n".join(roles_blocks)

    return f"""<?php

namespace Database\\Seeders;

use Illuminate\\Database\\Seeder;
use Spatie\\Permission\\Models\\Role;
use Spatie\\Permission\\Models\\Permission;

class RolesAndPermissionsSeeder extends Seeder
{{
    public function run(): void
    {{
        // Reset cached roles and permissions
        app()[\\Spatie\\Permission\\PermissionRegistrar::class]->forgetCachedPermissions();

        // Crear permisos
{perms_str}

        // Crear roles y asignar permisos
{roles_str}
    }}
}}
"""


def gen_database_seeder(hojas: dict, has_roles: bool = False) -> str:
    """Genera el DatabaseSeeder principal llamando a todos los seeders."""
    calls = []
    if has_roles:
        calls.append("        $this->call(RolesAndPermissionsSeeder::class);")
    for alias in hojas.keys():
        calls.append(f"        $this->call({nombre_modelo(alias)}Seeder::class);")

    calls_str = "\n".join(calls)
    return f"""<?php

namespace Database\\Seeders;

use Illuminate\\Database\\Seeder;

class DatabaseSeeder extends Seeder
{{
    public function run(): void
    {{
{calls_str}
    }}
}}
"""



def gen_filament_pages(alias: str, cfg_hoja: dict, empresa: str = "") -> dict:
    """Genera las 3 Pages que necesita cada Filament Resource.
    En el ListXxx además del CreateAction y CSV expone:
      - "Exportar Excel" (xlsx via PhpSpreadsheet, respeta orden cols del JSON,
        excluye _row_hash/created_at/updated_at).
      - "Sincronizar Excel" (sube .xlsx, dispara `php artisan kraftdo:sync`)."""
    modelo_n = nombre_modelo(alias)
    pages    = {}
    ns_r = "App\\Filament\\Resources\\"
    ns_f = "Filament\\"
    tabla_n = nombre_tabla(alias)

    # Orden de columnas del JSON (lo que el usuario espera ver/exportar).
    cols_json = list((cfg_hoja.get("columnas") or {}).keys())
    cols_json_php = "[" + ", ".join("'" + c + "'" for c in cols_json) + "]"

    pages["app/Filament/Resources/" + modelo_n + "Resource/Pages/List" + modelo_n + "s.php"] = (
        "<?php\n\n"
        "namespace " + ns_r + modelo_n + "Resource\\Pages;\n\n"
        "use " + ns_r + modelo_n + "Resource;\n"
        "use " + ns_f + "Actions;\n"
        "use " + ns_f + "Forms;\n"
        "use App\\Models\\" + modelo_n + ";\n"
        "use " + ns_f + "Resources\\Pages\\ListRecords;\n"
        "use " + ns_f + "Notifications\\Notification;\n"
        "use Illuminate\\Support\\Facades\\Storage;\n"
        "use Illuminate\\Support\\Facades\\Artisan;\n\n"
        "class List" + modelo_n + "s extends ListRecords\n"
        "{\n"
        "    protected static string $resource = " + modelo_n + "Resource::class;\n\n"
        "    protected function getHeaderActions(): array\n"
        "    {\n"
        "        return [\n"
        "            Actions\\CreateAction::make(),\n"
        # ── CSV (legacy)
        "            Actions\\Action::make('export_csv')\n"
        "                ->label('Exportar CSV')\n"
        "                ->icon('heroicon-o-arrow-down-tray')\n"
        "                ->color('gray')\n"
        "                ->action(function () {\n"
        "                    $registros = " + modelo_n + "::all();\n"
        "                    $filename = '" + tabla_n + "_' . now()->format('Y-m-d_His') . '.csv';\n"
        "                    return response()->streamDownload(function () use ($registros) {\n"
        "                        $out = fopen('php://output', 'w');\n"
        "                        fwrite($out, \"\\xEF\\xBB\\xBF\");\n"
        "                        if ($registros->isNotEmpty()) {\n"
        "                            fputcsv($out, array_keys($registros->first()->toArray()), ';');\n"
        "                            foreach ($registros as $r) {\n"
        "                                fputcsv($out, array_map(fn($v) => is_array($v) ? json_encode($v) : (string) $v, $r->toArray()), ';');\n"
        "                            }\n"
        "                        }\n"
        "                        fclose($out);\n"
        "                    }, $filename, ['Content-Type' => 'text/csv; charset=UTF-8']);\n"
        "                }),\n"
        # ── XLSX (PhpSpreadsheet, vendored por pxlrbt/filament-excel)
        "            Actions\\Action::make('export_xlsx')\n"
        "                ->label('Exportar Excel')\n"
        "                ->icon('heroicon-o-document-arrow-down')\n"
        "                ->color('success')\n"
        "                ->action(function () {\n"
        "                    $cols = " + cols_json_php + ";\n"
        "                    if (empty($cols)) {\n"
        "                        $first = " + modelo_n + "::first();\n"
        "                        $cols = $first ? array_keys($first->toArray()) : [];\n"
        "                    }\n"
        "                    $excluir = ['_row_hash','created_at','updated_at'];\n"
        "                    $cols = array_values(array_diff($cols, $excluir));\n"
        "                    $registros = " + modelo_n + "::all();\n"
        "                    $filename = '" + tabla_n + "_' . now()->format('Y-m-d_His') . '.xlsx';\n"
        "                    $tmp = tempnam(sys_get_temp_dir(), 'xlsx_') . '.xlsx';\n"
        "                    $sheet = new \\PhpOffice\\PhpSpreadsheet\\Spreadsheet();\n"
        "                    $ws = $sheet->getActiveSheet();\n"
        "                    foreach ($cols as $i => $c) {\n"
        "                        $ws->setCellValueByColumnAndRow($i + 1, 1, $c);\n"
        "                    }\n"
        "                    $row = 2;\n"
        "                    foreach ($registros as $r) {\n"
        "                        $arr = $r->toArray();\n"
        "                        foreach ($cols as $i => $c) {\n"
        "                            $v = $arr[$c] ?? null;\n"
        "                            if (is_array($v)) $v = json_encode($v);\n"
        "                            $ws->setCellValueByColumnAndRow($i + 1, $row, $v);\n"
        "                        }\n"
        "                        $row++;\n"
        "                    }\n"
        "                    (new \\PhpOffice\\PhpSpreadsheet\\Writer\\Xlsx($sheet))->save($tmp);\n"
        "                    return response()->download($tmp, $filename)->deleteFileAfterSend();\n"
        "                }),\n"
        # ── Sync desde Excel
        "            Actions\\Action::make('sync_excel')\n"
        "                ->label('Sincronizar Excel')\n"
        "                ->icon('heroicon-o-arrow-path')\n"
        "                ->color('warning')\n"
        "                ->schema([\n"
        "                    Forms\\Components\\FileUpload::make('archivo')\n"
        "                        ->label('Excel (.xlsx)')\n"
        "                        ->acceptedFileTypes(['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','application/vnd.ms-excel'])\n"
        "                        ->required()\n"
        "                        ->disk('local')\n"
        "                        ->directory('sync_temp')\n"
        "                        ->visibility('private'),\n"
        "                ])\n"
        "                ->action(function (array $data) {\n"
        "                    $stored = $data['archivo'] ?? null;\n"
        "                    if (!$stored) {\n"
        "                        Notification::make()->title('Sin archivo')->danger()->send();\n"
        "                        return;\n"
        "                    }\n"
        "                    $absStored = Storage::disk('local')->path($stored);\n"
        "                    $finalDir  = Storage::disk('local')->path('sync_temp');\n"
        "                    if (!is_dir($finalDir)) mkdir($finalDir, 0775, true);\n"
        "                    $finalPath = $finalDir . DIRECTORY_SEPARATOR . '" + empresa + ".xlsx';\n"
        "                    @copy($absStored, $finalPath);\n"
        "                    $exit = Artisan::call('kraftdo:sync', ['empresa' => '" + empresa + "']);\n"
        "                    $out  = trim(Artisan::output());\n"
        "                    if ($exit === 0) {\n"
        "                        Notification::make()->title('Sync OK')\n"
        "                            ->body($out !== '' ? \\Illuminate\\Support\\Str::limit($out, 400) : 'Importación completada.')\n"
        "                            ->success()->send();\n"
        "                    } else {\n"
        "                        Notification::make()->title('Sync con errores')\n"
        "                            ->body(\\Illuminate\\Support\\Str::limit($out, 400))\n"
        "                            ->danger()->send();\n"
        "                    }\n"
        "                }),\n"
        "        ];\n"
        "    }\n"
        "}\n"
    )

    pages["app/Filament/Resources/" + modelo_n + "Resource/Pages/Create" + modelo_n + ".php"] = (
        "<?php\n\n"
        "namespace " + ns_r + modelo_n + "Resource\\Pages;\n\n"
        "use " + ns_r + modelo_n + "Resource;\n"
        "use " + ns_f + "Resources\\Pages\\CreateRecord;\n\n"
        "class Create" + modelo_n + " extends CreateRecord\n"
        "{\n"
        "    protected static string $resource = " + modelo_n + "Resource::class;\n"
        "}\n"
    )

    pages["app/Filament/Resources/" + modelo_n + "Resource/Pages/Edit" + modelo_n + ".php"] = (
        "<?php\n\n"
        "namespace " + ns_r + modelo_n + "Resource\\Pages;\n\n"
        "use " + ns_r + modelo_n + "Resource;\n"
        "use " + ns_f + "Actions;\n"
        "use " + ns_f + "Resources\\Pages\\EditRecord;\n\n"
        "class Edit" + modelo_n + " extends EditRecord\n"
        "{\n"
        "    protected static string $resource = " + modelo_n + "Resource::class;\n\n"
        "    protected function getHeaderActions(): array\n"
        "    {\n"
        "        return [Actions\\DeleteAction::make()];\n"
        "    }\n"
        "}\n"
    )

    return pages


def gen_api_routes(hojas: dict) -> str:
    rutas = []
    for alias, cfg_hoja in hojas.items():
        if cfg_hoja.get("tipo") not in ("catalogo", "registros"):
            continue
        modelo  = nombre_modelo(alias)
        tabla   = nombre_tabla(alias)
        rutas.append(
            f"Route::apiResource('{tabla}', \\App\\Http\\Controllers\\Api\\{modelo}Controller::class);"
        )
    rutas_str = "\n".join(rutas)
    return f"""<?php

use Illuminate\\Support\\Facades\\Route;

/*
|--------------------------------------------------------------------------
| API Routes — Generadas automáticamente desde el JSON de configuración
|--------------------------------------------------------------------------
*/

{rutas_str}
"""


def gen_api_controller(alias: str, cfg_hoja: dict) -> str:
    modelo = nombre_modelo(alias)
    ns = "\\App\\"
    return (
        "<?php\n\n"
        "namespace App\\Http\\Controllers\\Api;\n\n"
        "use App\\Http\\Controllers\\Controller;\n"
        "use " + ns + "Models\\" + modelo + ";\n"
        "use " + ns + "Http\\Requests\\" + modelo + "Request;\n\n"
        "class " + modelo + "Controller extends Controller\n"
        "{\n"
        "    public function index()\n"
        "    {\n"
        "        return " + modelo + "::all();\n"
        "    }\n\n"
        "    public function store(" + modelo + "Request $request)\n"
        "    {\n"
        "        return " + modelo + "::create($request->validated());\n"
        "    }\n\n"
        "    public function show(" + modelo + " $record)\n"
        "    {\n"
        "        return $record;\n"
        "    }\n\n"
        "    public function update(" + modelo + "Request $request, " + modelo + " $record)\n"
        "    {\n"
        "        $record->update($request->validated());\n"
        "        return $record;\n"
        "    }\n\n"
        "    public function destroy(" + modelo + " $record)\n"
        "    {\n"
        "        $record->delete();\n"
        "        return response()->noContent();\n"
        "    }\n"
        "}\n"
    )



STOCK_FIELDS = {"stock", "stock_disponible", "disponible", "cantidad"}
STOCK_UMBRAL = 5


def gen_filament_widget(cfg: dict) -> str:
    """Genera un widget de KPIs para el dashboard de Filament.

    v21: incluye:
      - Conteo y sumas financieras por hoja (existente).
      - Stat 'Stock crítico' en color danger cuando hay registros con
        stock|cantidad|disponible < STOCK_UMBRAL.
      - KPIs custom desde cfg.logica_negocios.kpis_custom (si está definido).
    """
    hojas     = cfg.get("hojas", {})
    registros = {a: h for a, h in hojas.items() if h.get("tipo") in ("registros", "catalogo")}
    ns        = "\\App\\Models\\"
    logica    = cfg.get("logica_negocios", {})
    moneda    = logica.get("moneda", "CLP")
    simbolo   = "$"

    stats_lines = []

    # 1) KPIs custom desde el JSON: logica_negocios.kpis_custom = [{...}]
    for kpi in logica.get("kpis_custom", []):
        label   = kpi.get("label", "KPI")
        modelo_n = kpi.get("modelo")
        if not modelo_n:
            continue
        agreg   = kpi.get("agregacion", "count")
        color   = kpi.get("color", "primary")
        formato = kpi.get("formato", "numero")
        wheres  = kpi.get("where", [])
        descr   = kpi.get("descripcion", label)

        # Construir cadena de wheres en PHP
        where_php = ""
        for w in wheres:
            if not isinstance(w, list) or len(w) < 3:
                continue
            campo, op, val = w[0], w[1], w[2]
            if val == "today":
                val_php = "now()->startOfDay()"
            elif isinstance(val, (int, float)):
                val_php = str(val)
            else:
                val_php = "'" + str(val).replace("'", "\\'") + "'"
            where_php += "->where('" + campo + "', '" + op + "', " + val_php + ")"

        # Construir agregación
        if agreg == "count":
            valor = ns + modelo_n + "::query()" + where_php + "->count()"
        elif agreg.startswith("sum:"):
            campo = agreg.split(":", 1)[1]
            valor = "(float) " + ns + modelo_n + "::query()" + where_php + "->sum('" + campo + "')"
        elif agreg.startswith("avg:"):
            campo = agreg.split(":", 1)[1]
            valor = "(float) " + ns + modelo_n + "::query()" + where_php + "->avg('" + campo + "')"
        else:
            valor = ns + modelo_n + "::count()"

        if formato == "moneda":
            stat_value = "'" + simbolo + " ' . number_format(" + valor + ", 0, ',', '.')"
            icono = "heroicon-m-banknotes"
        else:
            stat_value = "(string) (" + valor + ")"
            icono = "heroicon-m-chart-bar"

        stats_lines.append(
            "            Stat::make('" + label + "', fn() => " + stat_value + ")\n"
            "                ->description('" + descr.replace("'", "\\'") + "')\n"
            "                ->descriptionIcon('" + icono + "')\n"
            "                ->color('" + color + "'),"
        )

    # 2) Stock crítico (auto-detectado)
    for alias, hoja in registros.items():
        modelo_n = nombre_modelo(alias)
        label    = alias.replace("_", " ").title()
        cols     = list(hoja.get("columnas", {}).keys())
        for campo in cols:
            if campo not in STOCK_FIELDS:
                continue
            stats_lines.append(
                "            Stat::make('" + label + " — Stock crítico', fn() => "
                + ns + modelo_n + "::where('" + campo + "', '<', " + str(STOCK_UMBRAL) + ")->count())\n"
                "                ->description('Registros con " + campo + " < " + str(STOCK_UMBRAL) + "')\n"
                "                ->descriptionIcon('heroicon-m-exclamation-triangle')\n"
                "                ->color(fn() => " + ns + modelo_n + "::where('" + campo + "', '<', " + str(STOCK_UMBRAL) + ")->count() > 0 ? 'danger' : 'gray'),"
            )
            break  # solo un stock-stat por hoja

    # 3) Conteos y sumas financieras por hoja
    for alias, hoja in list(registros.items())[:6]:
        modelo_n = nombre_modelo(alias)
        label    = alias.replace("_", " ").title()
        cols     = list(hoja.get("columnas", {}).keys())

        stats_lines.append(
            "            Stat::make('" + label + " — Total', fn() => "
            + ns + modelo_n + "::count())\n"
            "                ->description('Registros en " + label + "')\n"
            "                ->descriptionIcon('heroicon-m-rectangle-stack')\n"
            "                ->color('primary'),"
        )

        sumas = 0
        for campo in cols:
            if sumas >= 2:
                break
            if campo not in CAMPOS_MONEDA:
                continue
            label_campo = campo.replace("_", " ").capitalize()
            stats_lines.append(
                "            Stat::make('" + label + " — " + label_campo + "', "
                "fn() => '" + simbolo + " ' . number_format((float) "
                + ns + modelo_n + "::sum('" + campo + "'), 0, ',', '.'))\n"
                "                ->description('Suma " + label_campo + " (" + moneda + ")')\n"
                "                ->descriptionIcon('heroicon-m-banknotes')\n"
                "                ->color('success'),"
            )
            sumas += 1

    stats_str = "\n".join(stats_lines) if stats_lines else (
        "            Stat::make('Sin datos', '0')->color('gray'),"
    )
    ns_w = "Filament\\Widgets"

    return (
        "<?php\n\n"
        "namespace App\\" + ns_w + ";\n\n"
        "use " + ns_w + "\\StatsOverviewWidget as BaseWidget;\n"
        "use " + ns_w + "\\StatsOverviewWidget\\Stat;\n\n"
        "class KraftDoStatsWidget extends BaseWidget\n"
        "{\n"
        "    protected ?string $heading = 'Indicadores generales';\n\n"
        "    protected function getStats(): array\n"
        "    {\n"
        "        return [\n"
        + stats_str + "\n"
        "        ];\n"
        "    }\n"
        "}\n"
    )


def gen_widgets(cfg: dict) -> dict:
    """Genera múltiples Widgets de Filament basados en el bloque 'dashboard' del JSON.

    Estructura JSON esperada:
    \"dashboard\": [
      {
        \"nombre\": \"ResumenVentas\",
        \"titulo\": \"Ventas del Mes\",
        \"stats\": [
          { \"label\": \"Total CLP\", \"modelo\": \"Venta\", \"fn\": \"sum:total\", \"color\": \"success\" },
          { \"label\": \"Pedidos\", \"modelo\": \"Venta\", \"fn\": \"count\" }
        ]
      }
    ]
    """
    dashboard_cfg = cfg.get("dashboard", [])
    if not dashboard_cfg:
        return {}

    archivos = {}
    ns_models = "\\App\\Models\\"
    ns_widgets = "App\\Filament\\Widgets"

    for d in dashboard_cfg:
        nombre_clase = d.get("nombre", "CustomDashboardWidget")
        titulo = d.get("titulo", "Dashboard")
        stats_cfg = d.get("stats", [])

        stats_lines = []
        for s in stats_cfg:
            label = s.get("label", "Stat")
            modelo = s.get("modelo")
            fn = s.get("fn", "count")
            color = s.get("color", "primary")
            icon = s.get("icon", "heroicon-m-chart-bar")

            if not modelo:
                continue

            if fn == "count":
                expr = f"{ns_models}{modelo}::count()"
            elif fn.startswith("sum:"):
                campo = fn.split(":")[1]
                expr = f"(float) {ns_models}{modelo}::sum('{campo}')"
            elif fn.startswith("avg:"):
                campo = fn.split(":")[1]
                expr = f"(float) {ns_models}{modelo}::avg('{campo}')"
            else:
                expr = "0"

            # Formateo si parece dinero
            if "total" in fn or "precio" in fn or "costo" in fn or "monto" in fn:
                value_php = f"'$ ' . number_format({expr}, 0, ',', '.')"
            else:
                value_php = f"number_format({expr}, 0, ',', '.')"

            stats_lines.append(
                f"            Stat::make('{label}', fn() => {value_php})\n"
                f"                ->descriptionIcon('{icon}')\n"
                f"                ->color('{color}'),"
            )

        stats_str = "\n".join(stats_lines)
        contenido = f"""<?php

namespace {ns_widgets};

use Filament\\Widgets\\StatsOverviewWidget as BaseWidget;
use Filament\\Widgets\\StatsOverviewWidget\\Stat;

class {nombre_clase} extends BaseWidget
{{
    protected ?string $heading = '{titulo}';

    protected function getStats(): array
    {{
        return [
{stats_str}
        ];
    }}
}}
"""
        archivos[f"app/Filament/Widgets/{nombre_clase}.php"] = contenido

    return archivos


def gen_import_log_migration() -> str:
    """v25-fase3: tabla de auditoría del importer (incremental)."""
    return (
        "<?php\n\n"
        "use Illuminate\\Database\\Migrations\\Migration;\n"
        "use Illuminate\\Database\\Schema\\Blueprint;\n"
        "use Illuminate\\Support\\Facades\\Schema;\n\n"
        "return new class extends Migration\n"
        "{\n"
        "    public function up(): void\n"
        "    {\n"
        "        Schema::create('import_logs', function (Blueprint $table) {\n"
        "            $table->id();\n"
        "            $table->string('empresa', 100)->index();\n"
        "            $table->string('alias_hoja', 100)->index();\n"
        "            $table->dateTime('fecha_inicio')->index();\n"
        "            $table->dateTime('fecha_fin')->nullable();\n"
        "            $table->integer('nuevos')->default(0);\n"
        "            $table->integer('actualizados')->default(0);\n"
        "            $table->integer('sin_cambio')->default(0);\n"
        "            $table->integer('errores')->default(0);\n"
        "            $table->integer('duracion_ms')->default(0);\n"
        "            $table->text('mensaje')->nullable();\n"
        "            $table->timestamps();\n"
        "        });\n"
        "    }\n\n"
        "    public function down(): void\n"
        "    {\n"
        "        Schema::dropIfExists('import_logs');\n"
        "    }\n"
        "};\n"
    )


def gen_import_log_modelo() -> str:
    """Modelo Eloquent para import_logs."""
    return (
        "<?php\n\n"
        "namespace App\\Models;\n\n"
        "use Illuminate\\Database\\Eloquent\\Model;\n\n"
        "class ImportLog extends Model\n"
        "{\n"
        "    protected $table = 'import_logs';\n"
        "    protected $fillable = ['empresa','alias_hoja','fecha_inicio',\n"
        "        'fecha_fin','nuevos','actualizados','sin_cambio',\n"
        "        'errores','duracion_ms','mensaje'];\n"
        "    protected $casts = [\n"
        "        'fecha_inicio' => 'datetime',\n"
        "        'fecha_fin'    => 'datetime',\n"
        "        'nuevos'       => 'integer',\n"
        "        'actualizados' => 'integer',\n"
        "        'sin_cambio'   => 'integer',\n"
        "        'errores'      => 'integer',\n"
        "        'duracion_ms'  => 'integer',\n"
        "    ];\n"
        "}\n"
    )


def gen_import_log_resource() -> dict:
    """Filament Resource read-only para auditoría de imports."""
    archivos = {}
    archivos["app/Filament/Resources/ImportLogResource.php"] = (
        "<?php\n\n"
        "namespace App\\Filament\\Resources;\n\n"
        "use App\\Filament\\Resources\\ImportLogResource\\Pages;\n"
        "use App\\Models\\ImportLog;\n"
        "use Filament\\Schemas\\Schema;\n"
        "use Filament\\Resources\\Resource;\n"
        "use Filament\\Tables;\n"
        "use Filament\\Tables\\Table;\n\n"
        "class ImportLogResource extends Resource\n"
        "{\n"
        "    protected static ?string $model = ImportLog::class;\n"
        "    protected static \\BackedEnum|string|null $navigationIcon = 'heroicon-o-arrow-down-tray';\n"
        "    protected static ?string $navigationLabel = 'Import Logs';\n"
        "    protected static ?string $pluralModelLabel = 'Import Logs';\n"
        "    protected static ?int $navigationSort = 99;\n\n"
        "    public static function form(Schema $schema): Schema\n"
        "    {\n"
        "        return $schema->components([]);\n"
        "    }\n\n"
        "    public static function canCreate(): bool { return false; }\n\n"
        "    public static function table(Table $table): Table\n"
        "    {\n"
        "        return $table\n"
        "            ->defaultSort('fecha_inicio', 'desc')\n"
        "            ->columns([\n"
        "                Tables\\Columns\\TextColumn::make('fecha_inicio')\n"
        "                    ->label('Inicio')->dateTime('d/m/Y H:i:s')->sortable(),\n"
        "                Tables\\Columns\\TextColumn::make('empresa')->sortable()->searchable(),\n"
        "                Tables\\Columns\\TextColumn::make('alias_hoja')\n"
        "                    ->label('Hoja')->sortable()->searchable(),\n"
        "                Tables\\Columns\\TextColumn::make('nuevos')\n"
        "                    ->numeric()->color('success')->sortable(),\n"
        "                Tables\\Columns\\TextColumn::make('actualizados')\n"
        "                    ->numeric()->color('warning')->sortable(),\n"
        "                Tables\\Columns\\TextColumn::make('sin_cambio')\n"
        "                    ->numeric()->color('gray')->sortable(),\n"
        "                Tables\\Columns\\TextColumn::make('errores')\n"
        "                    ->numeric()->color('danger')->sortable(),\n"
        "                Tables\\Columns\\TextColumn::make('duracion_ms')\n"
        "                    ->label('ms')->numeric()->sortable(),\n"
        "            ])\n"
        "            ->actions([])  // read-only\n"
        "            ->bulkActions([]);\n"
        "    }\n\n"
        "    public static function getPages(): array\n"
        "    {\n"
        "        return ['index' => Pages\\ListImportLogs::route('/')];\n"
        "    }\n"
        "}\n"
    )
    archivos["app/Filament/Resources/ImportLogResource/Pages/ListImportLogs.php"] = (
        "<?php\n\n"
        "namespace App\\Filament\\Resources\\ImportLogResource\\Pages;\n\n"
        "use App\\Filament\\Resources\\ImportLogResource;\n"
        "use Filament\\Resources\\Pages\\ListRecords;\n\n"
        "class ListImportLogs extends ListRecords\n"
        "{\n"
        "    protected static string $resource = ImportLogResource::class;\n"
        "}\n"
    )
    return archivos


def gen_archivos_matriz_asistencia(idx_base: int, matriz_cfg: dict | None = None) -> dict:
    """Genera migrations + modelos + Filament Resources + Pages para
    tablas auxiliares de matriz_asistencia (asistencias, pagos_quincena).

    Devuelve {ruta: contenido} listo para inyectar en el dict archivos.
    Las tablas son fijas — no dependen del JSON de la matriz; solo de
    su existencia. La vista pivot usa matriz_cfg.mes_actual como mes
    por defecto.
    """
    archivos = {}
    mes_default = (matriz_cfg or {}).get("mes_actual", "2026-03")

    # 1) Migración asistencias
    mig_a_name = f"2026_01_01_{idx_base:06d}_create_asistencias_table"
    archivos[f"database/migrations/{mig_a_name}.php"] = (
        "<?php\n\n"
        "use Illuminate\\Database\\Migrations\\Migration;\n"
        "use Illuminate\\Database\\Schema\\Blueprint;\n"
        "use Illuminate\\Support\\Facades\\Schema;\n\n"
        "return new class extends Migration\n"
        "{\n"
        "    public function up(): void\n"
        "    {\n"
        "        Schema::create('asistencias', function (Blueprint $table) {\n"
        "            $table->id();\n"
        "            $table->string('trabajador', 100)->index();\n"
        "            $table->string('obra', 50)->nullable();\n"
        "            $table->string('codigo_obra', 20)->nullable();\n"
        "            $table->date('fecha')->index();\n"
        "            $table->string('mes', 7)->index(); // 'YYYY-MM'\n"
        "            $table->string('estado', 2)->index(); // A | F | L | ''\n"
        "            $table->timestamps();\n"
        "            $table->unique(['trabajador','fecha']);\n"
        "        });\n"
        "    }\n\n"
        "    public function down(): void\n"
        "    {\n"
        "        Schema::dropIfExists('asistencias');\n"
        "    }\n"
        "};\n"
    )

    # 2) Migración pagos_quincena
    mig_p_name = f"2026_01_01_{idx_base+1:06d}_create_pagos_quincena_table"
    archivos[f"database/migrations/{mig_p_name}.php"] = (
        "<?php\n\n"
        "use Illuminate\\Database\\Migrations\\Migration;\n"
        "use Illuminate\\Database\\Schema\\Blueprint;\n"
        "use Illuminate\\Support\\Facades\\Schema;\n\n"
        "return new class extends Migration\n"
        "{\n"
        "    public function up(): void\n"
        "    {\n"
        "        Schema::create('pagos_quincena', function (Blueprint $table) {\n"
        "            $table->id();\n"
        "            $table->string('trabajador', 100)->index();\n"
        "            $table->string('mes', 7)->index();\n"
        "            $table->string('periodo', 20); // 'quincena' | 'liquidacion'\n"
        "            $table->decimal('monto', 15, 2)->default(0);\n"
        "            $table->timestamps();\n"
        "            $table->unique(['trabajador','mes','periodo']);\n"
        "        });\n"
        "    }\n\n"
        "    public function down(): void\n"
        "    {\n"
        "        Schema::dropIfExists('pagos_quincena');\n"
        "    }\n"
        "};\n"
    )

    # 3) Modelo Asistencia
    archivos["app/Models/Asistencia.php"] = (
        "<?php\n\n"
        "namespace App\\Models;\n\n"
        "use Illuminate\\Database\\Eloquent\\Model;\n\n"
        "class Asistencia extends Model\n"
        "{\n"
        "    protected $table = 'asistencias';\n"
        "    protected $fillable = ['trabajador','obra','codigo_obra','fecha','mes','estado'];\n"
        "    protected $casts = ['fecha' => 'date'];\n"
        "}\n"
    )

    # 4) Modelo PagoQuincena
    archivos["app/Models/PagoQuincena.php"] = (
        "<?php\n\n"
        "namespace App\\Models;\n\n"
        "use Illuminate\\Database\\Eloquent\\Model;\n\n"
        "class PagoQuincena extends Model\n"
        "{\n"
        "    protected $table = 'pagos_quincena';\n"
        "    protected $fillable = ['trabajador','mes','periodo','monto'];\n"
        "    protected $casts = ['monto' => 'decimal:2'];\n"
        "}\n"
    )

    # 5) Filament Resources básicos
    for modelo, slug, plural in [("Asistencia","asistencias","Asistencias"),
                                  ("PagoQuincena","pagos-quincena","Pagos Quincena")]:
        ns_r = "App\\Filament\\Resources\\"
        archivos[f"app/Filament/Resources/{modelo}Resource.php"] = (
            "<?php\n\n"
            "namespace App\\Filament\\Resources;\n\n"
            f"use App\\Filament\\Resources\\{modelo}Resource\\Pages;\n"
            f"use App\\Models\\{modelo};\n"
            "use Filament\\Forms;\n"
            "use Filament\\Schemas\\Schema;\n"
            "use Filament\\Resources\\Resource;\n"
            "use Filament\\Tables;\n"
            "use Filament\\Tables\\Table;\n\n"
            f"class {modelo}Resource extends Resource\n"
            "{\n"
            f"    protected static ?string $model = {modelo}::class;\n"
            "    protected static \\BackedEnum|string|null $navigationIcon = 'heroicon-o-calendar-days';\n"
            f"    protected static ?string $navigationLabel = '{plural}';\n"
            f"    protected static ?string $pluralModelLabel = '{plural}';\n\n"
            "    public static function form(Schema $schema): Schema\n"
            "    {\n"
            "        return $schema->components([\n"
            "            Forms\\Components\\TextInput::make('trabajador')->required(),\n"
            + ("            Forms\\Components\\DatePicker::make('fecha')->required()->native(false),\n"
               "            Forms\\Components\\TextInput::make('mes')->placeholder('YYYY-MM'),\n"
               "            Forms\\Components\\TextInput::make('estado')->maxLength(2)->placeholder('A | F | L'),\n"
               "            Forms\\Components\\TextInput::make('obra'),\n"
               if modelo == "Asistencia" else
               "            Forms\\Components\\TextInput::make('mes')->placeholder('YYYY-MM'),\n"
               "            Forms\\Components\\Select::make('periodo')->options(['quincena'=>'Quincena','liquidacion'=>'Liquidación']),\n"
               "            Forms\\Components\\TextInput::make('monto')->numeric()->prefix('$'),\n")
            + "        ]);\n"
            "    }\n\n"
            "    public static function table(Table $table): Table\n"
            "    {\n"
            "        return $table\n"
            "            ->columns([\n"
            "                Tables\\Columns\\TextColumn::make('trabajador')->searchable()->sortable(),\n"
            + ("                Tables\\Columns\\TextColumn::make('fecha')->date('d/m/Y')->sortable(),\n"
               "                Tables\\Columns\\TextColumn::make('mes')->sortable(),\n"
               "                Tables\\Columns\\TextColumn::make('estado')->badge(),\n"
               "                Tables\\Columns\\TextColumn::make('obra')->sortable(),\n"
               if modelo == "Asistencia" else
               "                Tables\\Columns\\TextColumn::make('mes')->sortable(),\n"
               "                Tables\\Columns\\TextColumn::make('periodo')->badge(),\n"
               "                Tables\\Columns\\TextColumn::make('monto')->money('clp', divideBy: 1, locale: 'es_CL')->sortable(),\n")
            + "            ])\n"
            "            ->filters([])\n"
            "            ->actions([\n"
            "                \\Filament\\Actions\\EditAction::make(),\n"
            "                \\Filament\\Actions\\DeleteAction::make(),\n"
            "            ])\n"
            "            ->bulkActions([\n"
            "                \\Filament\\Actions\\BulkActionGroup::make([\n"
            "                    \\Filament\\Actions\\DeleteBulkAction::make(),\n"
            "                ]),\n"
            "            ]);\n"
            "    }\n\n"
            "    public static function getPages(): array\n"
            "    {\n"
            "        return [\n"
            f"            'index'  => Pages\\List{modelo}s::route('/'),\n"
            f"            'create' => Pages\\Create{modelo}::route('/create'),\n"
            f"            'edit'   => Pages\\Edit{modelo}::route('/{{record}}/edit'),\n"
            "        ];\n"
            "    }\n"
            "}\n"
        )

        # Pages básicas
        archivos[f"app/Filament/Resources/{modelo}Resource/Pages/List{modelo}s.php"] = (
            "<?php\n\n"
            f"namespace App\\Filament\\Resources\\{modelo}Resource\\Pages;\n\n"
            f"use App\\Filament\\Resources\\{modelo}Resource;\n"
            "use Filament\\Actions;\n"
            "use Filament\\Resources\\Pages\\ListRecords;\n\n"
            f"class List{modelo}s extends ListRecords\n"
            "{\n"
            f"    protected static string $resource = {modelo}Resource::class;\n\n"
            "    protected function getHeaderActions(): array\n"
            "    {\n"
            "        return [Actions\\CreateAction::make()];\n"
            "    }\n"
            "}\n"
        )
        archivos[f"app/Filament/Resources/{modelo}Resource/Pages/Create{modelo}.php"] = (
            "<?php\n\n"
            f"namespace App\\Filament\\Resources\\{modelo}Resource\\Pages;\n\n"
            f"use App\\Filament\\Resources\\{modelo}Resource;\n"
            "use Filament\\Resources\\Pages\\CreateRecord;\n\n"
            f"class Create{modelo} extends CreateRecord\n"
            "{\n"
            f"    protected static string $resource = {modelo}Resource::class;\n"
            "}\n"
        )
        archivos[f"app/Filament/Resources/{modelo}Resource/Pages/Edit{modelo}.php"] = (
            "<?php\n\n"
            f"namespace App\\Filament\\Resources\\{modelo}Resource\\Pages;\n\n"
            f"use App\\Filament\\Resources\\{modelo}Resource;\n"
            "use Filament\\Actions;\n"
            "use Filament\\Resources\\Pages\\EditRecord;\n\n"
            f"class Edit{modelo} extends EditRecord\n"
            "{\n"
            f"    protected static string $resource = {modelo}Resource::class;\n\n"
            "    protected function getHeaderActions(): array\n"
            "    {\n"
            "        return [Actions\\DeleteAction::make()];\n"
            "    }\n"
            "}\n"
        )

    # 6) Página Filament custom: matriz pivotada (trabajador × día)
    archivos["app/Filament/Pages/MatrizAsistencia.php"] = (
        "<?php\n\n"
        "namespace App\\Filament\\Pages;\n\n"
        "use Filament\\Pages\\Page;\n"
        "use App\\Models\\Asistencia;\n\n"
        "class MatrizAsistencia extends Page\n"
        "{\n"
        "    protected string $view = 'filament.pages.matriz-asistencia';\n"
        "    protected static \\BackedEnum|string|null $navigationIcon = 'heroicon-o-table-cells';\n"
        "    protected static ?string $navigationLabel = 'Matriz de Asistencia';\n"
        "    protected static ?string $title = 'Matriz de Asistencia';\n"
        "    protected static ?int $navigationSort = 5;\n\n"
        f"    public string $mes = '{mes_default}';\n\n"
        "    public function getTrabajadoresProperty(): array\n"
        "    {\n"
        "        return Asistencia::where('mes', $this->mes)\n"
        "            ->distinct()\n"
        "            ->orderBy('trabajador')\n"
        "            ->pluck('trabajador')\n"
        "            ->filter()\n"
        "            ->values()\n"
        "            ->all();\n"
        "    }\n\n"
        "    public function getMatrizProperty(): array\n"
        "    {\n"
        "        $rows = Asistencia::where('mes', $this->mes)->get();\n"
        "        $matriz = [];\n"
        "        foreach ($rows as $r) {\n"
        "            $dia = (int) (\\Carbon\\Carbon::parse($r->fecha)->format('d'));\n"
        "            $matriz[$r->trabajador][$dia] = $r->estado;\n"
        "        }\n"
        "        return $matriz;\n"
        "    }\n\n"
        "    public function getDiasMesProperty(): array\n"
        "    {\n"
        "        $fecha = \\Carbon\\Carbon::createFromFormat('Y-m', $this->mes);\n"
        "        return range(1, $fecha->daysInMonth);\n"
        "    }\n\n"
        "    /**\n"
        "     * Cicla el estado de una celda: '' → A → F → L → ''.\n"
        "     */\n"
        "    public function toggleAsistencia(string $trabajador, int $dia): void\n"
        "    {\n"
        "        $fecha = sprintf('%s-%02d', $this->mes, $dia);\n"
        "        $registro = Asistencia::firstOrNew([\n"
        "            'trabajador' => $trabajador,\n"
        "            'fecha'      => $fecha,\n"
        "        ]);\n"
        "        $estados = ['', 'A', 'F', 'L'];\n"
        "        $idx = array_search($registro->estado ?? '', $estados, true);\n"
        "        $next = $estados[(($idx === false ? 0 : $idx) + 1) % count($estados)];\n"
        "        if ($next === '') {\n"
        "            if ($registro->exists) {\n"
        "                $registro->delete();\n"
        "            }\n"
        "        } else {\n"
        "            $registro->mes    = $this->mes;\n"
        "            $registro->estado = $next;\n"
        "            $registro->save();\n"
        "        }\n"
        "    }\n"
        "}\n"
    )

    # 7) Vista Blade
    archivos["resources/views/filament/pages/matriz-asistencia.blade.php"] = """<x-filament-panels::page>
    <div class=\"space-y-4\">
        <div class=\"flex items-center gap-4\">
            <label class=\"font-medium\" for=\"mes\">Mes:</label>
            <input id=\"mes\" type=\"month\" wire:model.live=\"mes\"
                   class=\"rounded-lg border border-gray-300 px-3 py-1.5 text-sm
                          bg-white dark:bg-gray-900 dark:border-gray-700\"/>
            <span class=\"text-xs text-gray-500\">Click en una celda → cicla · → A → F → L → ·</span>
        </div>

        @php
            $trabajadores = $this->trabajadores;
            $matriz       = $this->matriz;
            $dias         = $this->diasMes;
        @endphp

        <div class=\"overflow-x-auto rounded-xl border border-gray-200 dark:border-gray-700
                    bg-white dark:bg-gray-900 shadow\">
            <table class=\"min-w-full text-sm\">
                <thead class=\"bg-gray-50 dark:bg-gray-800\">
                    <tr>
                        <th class=\"sticky left-0 z-10 bg-gray-50 dark:bg-gray-800
                                   px-3 py-2 text-left font-semibold border-r
                                   border-gray-200 dark:border-gray-700\">
                            Trabajador
                        </th>
                        @foreach($dias as $d)
                            <th class=\"px-1 py-2 w-9 text-center text-xs text-gray-600 dark:text-gray-400\">
                                {{ $d }}
                            </th>
                        @endforeach
                    </tr>
                </thead>
                <tbody>
                    @forelse($trabajadores as $trab)
                        <tr class=\"border-t border-gray-100 dark:border-gray-800\">
                            <td class=\"sticky left-0 z-10 bg-white dark:bg-gray-900
                                       px-3 py-2 font-medium whitespace-nowrap
                                       border-r border-gray-200 dark:border-gray-700\">
                                {{ $trab }}
                            </td>
                            @foreach($dias as $d)
                                @php
                                    $estado = $matriz[$trab][$d] ?? '';
                                    [$bg, $txt] = match($estado) {
                                        'A' => ['bg-green-200 hover:bg-green-300 dark:bg-green-700 dark:hover:bg-green-600', 'text-green-900 dark:text-green-100'],
                                        'F' => ['bg-red-200 hover:bg-red-300 dark:bg-red-700 dark:hover:bg-red-600',         'text-red-900 dark:text-red-100'],
                                        'L' => ['bg-yellow-200 hover:bg-yellow-300 dark:bg-yellow-700 dark:hover:bg-yellow-600','text-yellow-900 dark:text-yellow-100'],
                                        default => ['bg-gray-50 hover:bg-gray-200 dark:bg-gray-800 dark:hover:bg-gray-700',  'text-gray-400'],
                                    };
                                @endphp
                                <td class=\"p-0.5 text-center\">
                                    <button type=\"button\"
                                        wire:click=\"toggleAsistencia(@js($trab), {{ $d }})\"
                                        wire:loading.attr=\"disabled\"
                                        class=\"w-8 h-8 rounded {{ $bg }} {{ $txt }} text-xs font-bold transition\">
                                        {{ $estado ?: '·' }}
                                    </button>
                                </td>
                            @endforeach
                        </tr>
                    @empty
                        <tr>
                            <td colspan=\"{{ count($dias) + 1 }}\" class=\"px-3 py-8 text-center text-gray-500\">
                                Sin asistencias registradas para {{ $mes }}.
                            </td>
                        </tr>
                    @endforelse
                </tbody>
            </table>
        </div>

        <div class=\"flex flex-wrap gap-4 text-xs text-gray-700 dark:text-gray-300\">
            <span class=\"flex items-center gap-1.5\"><span class=\"inline-block w-4 h-4 rounded bg-green-200 dark:bg-green-700\"></span> A — Asistió</span>
            <span class=\"flex items-center gap-1.5\"><span class=\"inline-block w-4 h-4 rounded bg-red-200 dark:bg-red-700\"></span> F — Falta</span>
            <span class=\"flex items-center gap-1.5\"><span class=\"inline-block w-4 h-4 rounded bg-yellow-200 dark:bg-yellow-700\"></span> L — Licencia</span>
            <span class=\"flex items-center gap-1.5\"><span class=\"inline-block w-4 h-4 rounded bg-gray-50 dark:bg-gray-800 border border-gray-300 dark:border-gray-700\"></span> · — Sin marcar</span>
        </div>
    </div>
</x-filament-panels::page>
"""

    return archivos


def gen_metodo_recalcular_modelo(alias_agregado: str, cfg_hoja: dict,
                                  empresa_cfg: dict) -> str:
    """Genera el método estático Stock::recalcularModelo($valor) que se
    inyecta como trait/método en el modelo agregado.

    Aplica los SUM por fuente y hace UPSERT en la tabla agregada.
    """
    modelo = nombre_modelo(alias_agregado)
    grupo  = cfg_hoja.get("agrupar_por", cfg_hoja.get("identificador", "modelo"))
    fuentes = cfg_hoja.get("fuentes", [])
    hojas   = empresa_cfg.get("hojas", {})

    sumas = []
    for fuente in fuentes:
        alias_f   = fuente.get("hoja")
        if alias_f not in hojas:
            continue
        modelo_f  = nombre_modelo(alias_f)
        cgrupo_f  = fuente.get("campo_grupo", grupo)
        cval_f    = fuente.get("campo_valor")
        destino   = fuente.get("destino")
        if not (cval_f and destino):
            continue
        sumas.append(
            f"            '{destino}' => (float) \\App\\Models\\{modelo_f}::query()\n"
            f"                ->where('{cgrupo_f}', $valor)\n"
            f"                ->sum('{cval_f}'),"
        )
    sumas_str = "\n".join(sumas) if sumas else "            // (sin fuentes)"

    return (
        "    public static function recalcularModelo($valor): self\n"
        "    {\n"
        "        $valores = [\n"
        + sumas_str + "\n"
        "        ];\n"
        f"        $registro = static::query()->firstOrNew(['{grupo}' => $valor]);\n"
        "        foreach ($valores as $k => $v) {\n"
        "            $registro->$k = $v;\n"
        "        }\n"
        "        $registro->save();\n"
        "        return $registro;\n"
        "    }\n"
    )


def gen_alert_commands(cfg: dict) -> dict:
    """Genera Comandos de Laravel para alertas proactivas."""
    alertas = cfg.get("alertas", [])
    if not alertas:
        return {}

    archivos = {}
    ns_models = "\\App\\Models\\"

    for a in alertas:
        nombre = a.get("nombre", "AlertaCustom")
        modelo = a.get("modelo")
        condicion = a.get("condicion", "id > 0")
        mensaje = a.get("mensaje", f"Alerta en {modelo}")
        signature = f"alerta:{nombre.lower()}"

        if not modelo:
            continue

        contenido = f"""<?php

namespace App\\Console\\Commands;

use Illuminate\\Console\\Command;
use Illuminate\\Support\\Facades\\Log;
use {ns_models}{modelo};

class {nombre}Command extends Command
{{
    protected $signature = '{signature}';
    protected $description = 'Alerta proactiva: {nombre}';

    public function handle(): int
    {{
        $registros = {modelo}::whereRaw('{condicion}')->get();

        if ($registros->count() > 0) {{
            foreach ($registros as $r) {{
                $msg = "{mensaje}";
                foreach ($r->toArray() as $key => $val) {{
                    if (!is_array($val)) {{
                        $msg = str_replace('{{' . $key . '}}', (string) $val, $msg);
                    }}
                }}
                Log::warning("[ALERTA {nombre}] " . $msg);
                $this->warn($msg);
            }}
            $this->info("Se detectaron " . $registros->count() . " registros.");
        }}

        return self::SUCCESS;
    }}
}}
"""
        archivos[f"app/Console/Commands/{nombre}Command.php"] = contenido

    return archivos


def gen_console_routes(cfg: dict) -> str:
    """Genera routes/console.php con la programación de alertas."""
    alertas = cfg.get("alertas", [])
    schedules = []

    for a in alertas:
        nombre = a.get("nombre", "AlertaCustom")
        prog = a.get("programacion", "daily()")
        if not prog.endswith(')'):
            prog += '()'
        signature = f"alerta:{nombre.lower()}"
        schedules.append(f"Schedule::command('{signature}')->{prog};")

    schedules_str = "\n".join(schedules)

    return f"""<?php

use Illuminate\\Support\\Facades\\Schedule;

/*
|--------------------------------------------------------------------------
| Console Routes
|--------------------------------------------------------------------------
*/

{schedules_str}

Schedule::command('kraftdo:recalcular')->daily();
"""


def gen_recalcular_command(modelos_observers: list[str]) -> str:
    """Genera app/Console/Commands/RecalcularTodo.php que llama save() en cada
    registro de los modelos con observer, forzando recálculo de campos derivados."""
    if not modelos_observers:
        modelos_observers = []
    lineas = ",\n".join(
        "            \\App\\Models\\" + m + "::class"
        for m in modelos_observers
    )
    return (
        "<?php\n\n"
        "namespace App\\Console\\Commands;\n\n"
        "use Illuminate\\Console\\Command;\n\n"
        "class RecalcularTodo extends Command\n"
        "{\n"
        "    protected $signature = 'kraftdo:recalcular';\n"
        "    protected $description = 'Recalcula todos los campos derivados llamando save() en cada registro';\n\n"
        "    public function handle(): int\n"
        "    {\n"
        "        $modelos = [\n"
        + lineas + "\n"
        "        ];\n\n"
        "        foreach ($modelos as $clase) {\n"
        "            if (!class_exists($clase)) {\n"
        "                continue;\n"
        "            }\n"
        "            $count = 0;\n"
        "            $clase::query()->lazy()->each(function ($r) use (&$count) {\n"
        "                // Forzar dirty para que el observer dispare en cada registro,\n"
        "                // aunque el record no haya cambiado desde el último save().\n"
        "                $r->updated_at = now();\n"
        "                $r->save();\n"
        "                $count++;\n"
        "            });\n"
        "            $this->info(\"  {$clase}: {$count} registros recalculados\");\n"
        "        }\n\n"
        "        $this->info('✅ Recalculo completo.');\n"
        "        return self::SUCCESS;\n"
        "    }\n"
        "}\n"
    )


def gen_sync_command(empresa: str) -> str:
    """v25-fase4: php artisan kraftdo:sync {empresa} — invoca el importer
    Python sobre el .xlsx que el panel acaba de subir a sync_temp/."""
    base_repo = os.path.dirname(os.path.abspath(__file__))
    importer  = os.path.join(base_repo, "importar_excel_a_mysql.py")
    return (
        "<?php\n\n"
        "namespace App\\Console\\Commands;\n\n"
        "use Illuminate\\Console\\Command;\n"
        "use Illuminate\\Support\\Facades\\Storage;\n\n"
        "class SyncDesdeExcel extends Command\n"
        "{\n"
        "    protected $signature = 'kraftdo:sync {empresa}';\n"
        "    protected $description = 'Sincroniza datos desde un Excel subido al panel via importar_excel_a_mysql.py';\n\n"
        "    public function handle(): int\n"
        "    {\n"
        "        $empresa = (string) $this->argument('empresa');\n"
        f"        $repoBase    = '{base_repo}';\n"
        f"        $importerPy  = '{importer}';\n"
        "        $cfgPath     = $repoBase . DIRECTORY_SEPARATOR . 'empresas' . DIRECTORY_SEPARATOR . $empresa . '.json';\n"
        "        if (!is_file($cfgPath)) {\n"
        "            $this->error('No existe ' . $cfgPath);\n"
        "            return self::FAILURE;\n"
        "        }\n"
        "        $cfg = json_decode(file_get_contents($cfgPath), true);\n"
        "        $excelDest = $repoBase . DIRECTORY_SEPARATOR . ($cfg['fuente']['archivo'] ?? '');\n"
        "        $excelTmp  = Storage::disk('local')->path('sync_temp/' . $empresa . '.xlsx');\n"
        "        if (!is_file($excelTmp)) {\n"
        "            $this->error('No se encontro ' . $excelTmp . ' (subir primero desde el panel)');\n"
        "            return self::FAILURE;\n"
        "        }\n"
        "        if (!@copy($excelTmp, $excelDest)) {\n"
        "            $this->error('No se pudo copiar el Excel a ' . $excelDest);\n"
        "            return self::FAILURE;\n"
        "        }\n"
        "        $cmd = 'python3 ' . escapeshellarg($importerPy) . ' '\n"
        "             . escapeshellarg($empresa) . ' ' . escapeshellarg(base_path()) . ' 2>&1';\n"
        "        $out = shell_exec($cmd);\n"
        "        $this->line($out ?? '(sin salida)');\n"
        "        // Resumen del último import_log\n"
        "        $logs = \\App\\Models\\ImportLog::where('empresa', $empresa)\n"
        "            ->orderByDesc('id')->take(10)->get();\n"
        "        if ($logs->count()) {\n"
        "            $this->line('--- Últimos imports ---');\n"
        "            foreach ($logs as $l) {\n"
        "                $this->line(sprintf('%s: %d nuevos, %d upd, %d igual, %d err (%dms)',\n"
        "                    $l->alias_hoja, $l->nuevos, $l->actualizados,\n"
        "                    $l->sin_cambio, $l->errores, $l->duracion_ms));\n"
        "            }\n"
        "        }\n"
        "        return self::SUCCESS;\n"
        "    }\n"
        "}\n"
    )


def gen_install_script(empresa: str, hojas: dict) -> str:
    modelos = [nombre_modelo(a) for a, h in hojas.items()
               if h.get("tipo") in ("catalogo", "registros")]
    resources_cmd = "\n".join(
        f"php artisan make:filament-resource {m} --generate --force"
        for m in modelos
    )
    seeders_cmd = "\n".join(
        f"php artisan db:seed --class={nombre_modelo(a)}Seeder"
        for a in hojas.keys()
    )
    return (
        "#!/bin/bash\n"
        "# KraftDo — install.sh\n"
        "# Instala el sistema generado para: " + empresa + "\n"
        "# Ejecutar desde la raíz del proyecto Laravel\n\n"
        "set -e\n"
        'echo "🚀 Instalando sistema ' + empresa + '..."\n\n'
        "# 1. Dependencias\n"
        "composer require filament/filament\\n"
        "composer require bezhansalleh/filament-shield\\n"
        "composer require pxlrbt/filament-excel\\n"
        "composer require leandrocfe/filament-apex-charts\\n"
        "composer require spatie/laravel-medialibrary\\n"
        "composer require filament/spatie-laravel-media-library-plugin\\n"
        "composer require spatie/laravel-permission\\n"
        "composer require spatie/laravel-activitylog\\n"
        'echo "✅ Dependencias + plugins Filament instalados"\n\n'
        "# 2. Migraciones\n"
        "php artisan vendor:publish --provider=\"Spatie\\Permission\\PermissionServiceProvider\"\n"
        "php artisan vendor:publish --provider=\"Spatie\\Activitylog\\ActivitylogServiceProvider\" --tag=\"activitylog-migrations\"\n"
        "php artisan migrate --force\n"
        'echo "✅ Tablas creadas"\n\n'
        "# 3. Seeders (datos de ejemplo)\n"
        + seeders_cmd + "\n"
        'echo "✅ Datos de ejemplo cargados"\n\n'
        "# 4. Recursos Filament\n"
        + resources_cmd + "\n"
        'echo "✅ Recursos Filament generados"\n\n'
        "# 5. Compilar assets\n"
        "npm install && npm run build\n"
        'echo "✅ Assets compilados"\n\n'
        "# 6. Recalcular campos derivados (observers se ejecutan al save)\n"
        "php artisan kraftdo:recalcular || true\n"
        'echo "✅ Campos derivados recalculados"\n\n'
        "# 7. Crear usuario admin\n"
        "php artisan make:filament-user\n"
        'echo "✅ Listo! Abre /admin en tu navegador"\n'
    )


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def _auto_indices(columnas: dict) -> list[str]:
    """Detecta campos que deberían tener índice MySQL automáticamente."""
    indices = []
    patrones_indice = [
        "fecha", "date", "created", "updated",
        "estado", "status", "tipo", "type",
        "id_", "_id", "sku", "email",
        "cliente", "proveedor", "empresa",
    ]
    for campo in columnas.keys():
        campo_lower = campo.lower()
        if any(p in campo_lower for p in patrones_indice):
            indices.append(campo)
    return indices



def _crear_base_laravel(output_dir: str, empresa: str, cfg: dict) -> bool:
    """
    Crea un proyecto Laravel completo con Filament instalado.
    Retorna True si se creó correctamente.
    """
    import subprocess, shutil

    nombre = cfg["empresa"]["nombre"]
    db_name = empresa  # usar el nombre de empresa directamente como nombre de BD
    db_pass = "4c4e99bc4d1c2e6a"  # password MySQL del usuario kraftdo

    print(f"\n🚀 Creando base Laravel en {output_dir}...")

    # Verificar que composer esté disponible
    if not shutil.which("composer"):
        print("❌ composer no encontrado — instalarlo primero")
        print("   curl -sS https://getcomposer.org/installer | php -- --install-dir=/usr/local/bin --filename=composer")
        return False

    # Si ya existe y tiene artisan, no recrear
    if os.path.exists(os.path.join(output_dir, "artisan")):
        print(f"  ✓ Base Laravel ya existe en {output_dir}")
        return True

    parent = os.path.dirname(os.path.abspath(output_dir))
    dirname = os.path.basename(output_dir)

    # 1. composer create-project
    print("  📦 Ejecutando composer create-project laravel/laravel ...")
    r = subprocess.run(
        ["composer", "create-project", "laravel/laravel", dirname, "--no-interaction", "--quiet"],
        cwd=parent, capture_output=True, text=True
    )
    if r.returncode != 0:
        print(f"  ❌ Error en composer create-project:\n{r.stderr[:500]}")
        return False
    print("  ✓ Laravel instalado")

    # 2. composer require filament
    print("  🎛️  Instalando Filament 3 y extensiones (Shield, Excel, Spatie) ...")
    r = subprocess.run(
        ["composer", "require", "filament/filament:^4.0", 
         "bezhansalleh/filament-shield", "pxlrbt/filament-excel", 
         "leandrocfe/filament-apex-charts", "spatie/laravel-medialibrary", 
         "filament/spatie-laravel-media-library-plugin",
         "spatie/laravel-permission", "spatie/laravel-activitylog", 
         "--no-interaction", "--quiet"],
        cwd=output_dir, capture_output=True, text=True
    )
    if r.returncode != 0:
        print(f"  ❌ Error instalando Filament:\n{r.stderr[:500]}")
        return False
    print("  ✓ Filament instalado")

    # 3. php artisan filament:install --panels
    print("  ⚙️  Configurando panel Filament ...")
    r = subprocess.run(
        ["php", "artisan", "filament:install", "--panels", "--no-interaction"],
        cwd=output_dir, capture_output=True, text=True
    )
    if r.returncode != 0:
        print(f"  ❌ Error en filament:install:\n{r.stderr[:300]}")
        return False
    print("  ✓ Panel Filament configurado")

    # 4. Escribir .env con datos de la empresa
    env_path = os.path.join(output_dir, ".env")
    if os.path.exists(env_path):
        with open(env_path, encoding="utf-8") as f:
            env = f.read()
        env = re.sub(r'APP_NAME=.*', 'APP_NAME="' + nombre + '"', env)
        # Laravel 13 tiene DB_ comentadas — descomentar y configurar
        env = re.sub(r"#\s*DB_CONNECTION=.*", "DB_CONNECTION=mysql", env)
        env = re.sub(r"DB_CONNECTION=.*", "DB_CONNECTION=mysql", env)
        env = re.sub(r"#\s*DB_HOST=.*", "DB_HOST=127.0.0.1", env)
        env = re.sub(r"DB_HOST=.*", "DB_HOST=127.0.0.1", env)
        env = re.sub(r"#\s*DB_PORT=.*", "DB_PORT=3306", env)
        env = re.sub(r"DB_PORT=.*", "DB_PORT=3307", env)
        env = re.sub(r"#\s*DB_DATABASE=.*", f"DB_DATABASE={db_name}", env)
        env = re.sub(r"DB_DATABASE=.*", f"DB_DATABASE={db_name}", env)
        env = re.sub(r"#\s*DB_USERNAME=.*", "DB_USERNAME=kraftdo", env)
        env = re.sub(r"DB_USERNAME=.*", "DB_USERNAME=kraftdo", env)
        env = re.sub(r"DB_PASSWORD=\S*", f"DB_PASSWORD={db_pass}", env)
        env = re.sub(r"DB_PASSWORD=\S*", f"DB_PASSWORD={db_pass}", env)
        env = re.sub(r"DB_HOST=.*", "DB_HOST=127.0.0.1", env)
        env = re.sub(r"DB_PORT=.*", "DB_PORT=3307", env)
        env = re.sub(r"DB_DATABASE=.*", f"DB_DATABASE={db_name}", env)
        env = re.sub(r"DB_USERNAME=.*", "DB_USERNAME=kraftdo", env)
        env = re.sub(r"DB_PASSWORD=.*", "DB_PASSWORD=", env)
        with open(env_path, "w", encoding="utf-8") as f:
            f.write(env)
        # Asegurar DB_PASSWORD en el .env
        with open(env_path, "a", encoding="utf-8") as f:
            f.write(f"\nDB_PASSWORD={db_pass}\n")
        print(f"  ✓ .env configurado para {nombre} (mysql)")

    # 5. php artisan key:generate
    subprocess.run(
        ["php", "artisan", "key:generate", "--no-interaction"],
        cwd=output_dir, capture_output=True
    )
    print("  ✓ APP_KEY generado")

    print(f"\n✅ Base Laravel lista en {output_dir}\n")
    return True

def generar(empresa: str = None, output_dir: str = "./laravel_output", preview: bool = False, solo: str = None, config_path: str = None):
    # Cargar config
    if config_path:
        cfg_path = config_path
        # Si no existe literal, buscar en empresas/
        if not os.path.exists(cfg_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            cfg_path = os.path.join(base_dir, "empresas", config_path)
    else:
        if not empresa:
            print("❌ Falta nombre de empresa o --config")
            sys.exit(1)
        base = os.path.dirname(os.path.abspath(__file__))
        cfg_path = os.path.join(base, "empresas", f"{empresa}.json")

    if not os.path.exists(cfg_path):
        print(f"❌ No encontré: {cfg_path}")
        sys.exit(1)

    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)

    # El nombre de la empresa para la BD y carpetas
    empresa_id = empresa if empresa else os.path.basename(cfg_path).replace(".json", "")

    nombre_empresa = cfg["empresa"]["nombre"]
    hojas = cfg["hojas"]

    print(f"\n{'='*60}")
    print(f"  KraftDo Generator — {nombre_empresa}")
    print(f"  Output: {output_dir}")
    print(f"{'='*60}\n")

    # Opción B: crear base Laravel completa si no existe
    if not preview:
        if not _crear_base_laravel(output_dir, empresa_id, cfg):
            print("⚠️  Continuando sin crear base Laravel — solo se generan los archivos")

    archivos = {}  # path → contenido

    hojas_generables = {
        a: h for a, h in hojas.items()
        if h.get("tipo") in ("catalogo", "registros", "agregado")
    }

    # Detectar cascadas: por cada hoja "agregado" con fuentes, anotar en
    # cada hoja fuente un _cascadas[] que el observer leerá. Mutación en
    # memoria sólo, no toca el JSON en disco.
    for alias_agg, hoja_agg in hojas.items():
        if hoja_agg.get("tipo") != "agregado":
            continue
        modelo_agg = nombre_modelo(alias_agg)
        for fuente in hoja_agg.get("fuentes", []):
            alias_fuente = fuente.get("hoja")
            if alias_fuente not in hojas_generables:
                continue
            hoja_f = hojas_generables[alias_fuente]
            hoja_f.setdefault("_cascadas", []).append({
                "modelo_destino": modelo_agg,
                "campo_grupo":    fuente.get("campo_grupo", "modelo"),
            })

    # 1. Migraciones
    if not solo or solo == "migraciones":
        for idx, (alias, cfg_hoja) in enumerate(hojas_generables.items(), start=1):
            nombre_arch = f"database/migrations/{nombre_migration(alias, idx)}.php"
            archivos[nombre_arch] = gen_migracion(alias, cfg_hoja, idx)
            print(f"  📄 {nombre_arch}")

    # 2. Modelos
    # Analizar patrones del Excel si está disponible
    excel_path_check = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     cfg.get("fuente", {}).get("archivo", ""))
    patrones = {}
    if RELACIONES_OK and os.path.exists(excel_path_check):
        try:
            patrones = _analizar_patrones(excel_path_check)
            hojas_raras = {h: d for h, d in patrones.items()
                           if d.get("patron") not in ("vertical", "con_totales", "vacia")}
            if hojas_raras:
                print(f"  ⚠️  Patrones no estándar detectados:")
                for hoja, diag in hojas_raras.items():
                    print(f"     {hoja}: {diag['patron']} — {diag['transformacion'][:60]}")
                print()
        except Exception:
            pass

    # Calcular relaciones y fórmulas una sola vez
    rels = detectar_relaciones(cfg) if RELACIONES_OK else []
    rels_x_tabla = relaciones_por_tabla(rels) if RELACIONES_OK else {}
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), cfg.get("fuente", {}).get("archivo", ""))
    formulas_all = analizar_excel_formulas(excel_path, cfg) if RELACIONES_OK and os.path.exists(excel_path) else {}
    # Pasar datos a gen_modelo via atributos
    gen_modelo._relaciones = rels
    gen_modelo._formulas   = formulas_all

    # Detectar enums leyendo el Excel (string columns con pocos valores únicos)
    enums_por_hoja = _detectar_enums_excel(excel_path, cfg)
    if enums_por_hoja:
        n = sum(len(e) for e in enums_por_hoja.values())
        print(f"  🔠 {n} columnas detectadas como enum auto:")
        for a, e in enums_por_hoja.items():
            for campo, vals in e.items():
                muestra = ", ".join(vals[:4]) + (" …" if len(vals) > 4 else "")
                print(f"     {a}.{campo} = [{muestra}]")
        print()

    if not solo or solo == "modelos":
        for alias, cfg_hoja in hojas_generables.items():
            modelo_n = nombre_modelo(alias)
            nombre_arch = f"app/Models/{modelo_n}.php"
            archivos[nombre_arch] = gen_modelo(alias, cfg_hoja, cfg, rels, formulas_all)
            print(f"  📦 {nombre_arch}")

    # 3. Filament Resources
    if not solo or solo == "filament":
        for alias, cfg_hoja in hojas_generables.items():
            modelo = nombre_modelo(alias)
            nombre_arch = f"app/Filament/Resources/{modelo}Resource.php"
            archivos[nombre_arch] = gen_filament_resource(
                alias, cfg_hoja, cfg, rels, enums_por_hoja.get(alias, {})
            )
            print(f"  🎛️  {nombre_arch}")

    # 3a-bis. Observers (campos calculados auto al guardar)
    modelos_con_observer = []
    if not solo or solo in ("modelos", "observers"):
        for alias, cfg_hoja in hojas_generables.items():
            obs = gen_observer(alias, cfg_hoja, cfg)
            if obs is None:
                continue
            modelo_n = nombre_modelo(alias)
            nombre_arch = f"app/Observers/{modelo_n}Observer.php"
            archivos[nombre_arch] = obs
            modelos_con_observer.append(modelo_n)
            print(f"  👁️  {nombre_arch}")

    # 3a-ter. Comando artisan kraftdo:recalcular
    if (not solo or solo in ("modelos", "observers")) and modelos_con_observer:
        nombre_arch = "app/Console/Commands/RecalcularTodo.php"
        archivos[nombre_arch] = gen_recalcular_command(modelos_con_observer)
        print(f"  🔁 {nombre_arch}")

    # 3b. FormRequests
    if not solo or solo in ("modelos", "requests"):
        for alias, cfg_hoja in hojas_generables.items():
            modelo_n = nombre_modelo(alias)
            nombre_arch = f"app/Http/Requests/{modelo_n}Request.php"
            archivos[nombre_arch] = gen_form_request(
                alias, cfg_hoja, enums_por_hoja.get(alias, {})
            )
            print(f"  ✅ {nombre_arch}")

    # 3c. Seeders
    if not solo or solo in ("modelos", "seeders"):
        # Roles y Permisos (v24b)
        roles_content = gen_roles_seeder(cfg)
        if roles_content:
            archivos["database/seeders/RolesAndPermissionsSeeder.php"] = roles_content
            print("  🌱 database/seeders/RolesAndPermissionsSeeder.php")

        for alias, cfg_hoja in hojas_generables.items():
            modelo_n = nombre_modelo(alias)
            nombre_arch = f"database/seeders/{modelo_n}Seeder.php"
            archivos[nombre_arch] = gen_seeder(alias, cfg_hoja)
            print(f"  🌱 {nombre_arch}")

        # DatabaseSeeder principal
        archivos["database/seeders/DatabaseSeeder.php"] = gen_database_seeder(
            hojas_generables, has_roles=bool(roles_content)
        )
        print("  🌱 database/seeders/DatabaseSeeder.php")

    # 3d. Filament Pages
    if not solo or solo in ("filament", "pages"):
        for alias, cfg_hoja in hojas_generables.items():
            pages = gen_filament_pages(alias, cfg_hoja, empresa)
            for path, contenido in pages.items():
                archivos[path] = contenido
                print(f"  📄 {path}")

    # 4. API Routes y Controllers
    if not solo or solo == "api":
        archivos["routes/api.php"] = gen_api_routes(hojas_generables)
        print(f"  🔌 routes/api.php")
        for alias, cfg_hoja in hojas_generables.items():
            modelo = nombre_modelo(alias)
            nombre_arch = f"app/Http/Controllers/Api/{modelo}Controller.php"
            archivos[nombre_arch] = gen_api_controller(alias, cfg_hoja)
            print(f"  🔌 {nombre_arch}")

    # 6. Dashboard Widget de Filament
    if not solo or solo in ("filament", "widgets"):
        nombre_arch = "app/Filament/Widgets/KraftDoStatsWidget.php"
        archivos[nombre_arch] = gen_filament_widget(cfg)
        print(f"  📊 {nombre_arch}")

        # Widgets dinámicos (v24b)
        widgets_dinamicos = gen_widgets(cfg)
        for path, contenido in widgets_dinamicos.items():
            archivos[path] = contenido
            print(f"  📊 {path}")

    # 7. Alertas proactivas y Programación (v24b)
    alert_commands = gen_alert_commands(cfg)
    for path, contenido in alert_commands.items():
        archivos[path] = contenido
        print(f"  🔔 {path}")

    archivos["routes/console.php"] = gen_console_routes(cfg)
    print(f"  ⏰ routes/console.php")

    # 4a-bis. Tabla import_logs (auditoría incremental, v25-fase3)
    archivos["database/migrations/0000_00_00_000000_create_import_logs_table.php"] = gen_import_log_migration()
    archivos["app/Models/ImportLog.php"] = gen_import_log_modelo()
    for path, contenido in gen_import_log_resource().items():
        archivos[path] = contenido
    print("  📋 import_logs (migration + modelo + resource)")

    # 4a-ter. Comando artisan kraftdo:sync (v25-fase4)
    archivos["app/Console/Commands/SyncDesdeExcel.php"] = gen_sync_command(empresa)
    print("  🔄 app/Console/Commands/SyncDesdeExcel.php (kraftdo:sync)")

    # 4b. Tablas auxiliares para hojas tipo "matriz_asistencia"
    matriz_cfg = next(
        (h for h in hojas.values() if h.get("tipo") == "matriz_asistencia"),
        None
    )
    if matriz_cfg:
        idx_base = len(hojas_generables) + 100  # después de las migrations regulares
        for path, contenido in gen_archivos_matriz_asistencia(idx_base, matriz_cfg).items():
            archivos[path] = contenido
            print(f"  📅 {path}")

    # 5. Script de instalación
    archivos["install.sh"] = gen_install_script(empresa_id, hojas_generables)
    print(f"  ⚙️  install.sh")

    print(f"\n  Total: {len(archivos)} archivos\n")

    if preview:
        print("── PREVIEW (no se escriben archivos) ──────────────────────")
        for path, contenido in list(archivos.items())[:3]:
            print(f"\n{'─'*60}\n{path}\n{'─'*60}")
            print(contenido[:800] + ("..." if len(contenido) > 800 else ""))
        return archivos

    # Escribir archivos
    for path, contenido in archivos.items():
        full_path = os.path.join(output_dir, path)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(contenido)

    print(f"✅ Generado en: {output_dir}")

    # Auto-setup: migrar y crear usuario admin si no es preview
    if not preview:
        import subprocess
        db_pass = "4c4e99bc4d1c2e6a"
        db_name = empresa  # usar el nombre de empresa directamente como nombre de BD
        db_root_pass = "51381c69b62f87a6"
        print("\n🔧 Configurando base de datos...")
        # Crear BD si no existe
        try:
            import mysql.connector
            conn = mysql.connector.connect(
                host="127.0.0.1", port=3307,
                user="root", password=db_root_pass
            )
            conn.cursor().execute(f"CREATE DATABASE IF NOT EXISTS `{db_name}`")
            cur = conn.cursor()
            cur.execute(f"GRANT ALL ON `{db_name}`.* TO 'kraftdo'@'%'")
            cur.execute("FLUSH PRIVILEGES")
            conn.commit()
            conn.close()
            print(f"  ✓ BD {db_name} creada")
        except Exception as e:
            print(f"  ⚠️  No se pudo crear la BD: {e}")

        # Publicar migraciones de Spatie (v24b)
        subprocess.run(
            ["php", "artisan", "vendor:publish", "--provider=Spatie\\Permission\\PermissionServiceProvider"],
            cwd=output_dir, capture_output=True
        )
        subprocess.run(
            ["php", "artisan", "vendor:publish", "--provider=Spatie\\Activitylog\\ActivitylogServiceProvider", "--tag=activitylog-migrations"],
            cwd=output_dir, capture_output=True
        )

        # Migrar
        r = subprocess.run(
            ["php", "artisan", "migrate:fresh", "--force"],
            cwd=output_dir, capture_output=True, text=True
        )
        if r.returncode == 0:
            print("  ✓ Migraciones ejecutadas")
        else:
            print(f"  ⚠️  Error en migrate: {r.stderr[:200]}")

        # Crear usuario admin
        tinker_cmd = (
            "App\\Models\\User::create(["
            "'name'=>'Admin',"
            f"'email'=>'{cfg['empresa'].get('email','admin@kraftdo.cl')}',"
            "'password'=>bcrypt('kraftdo2026')"
            "]);echo 'ok';"
        )
        subprocess.run(
            ["php", "artisan", "tinker", "--execute", tinker_cmd],
            cwd=output_dir, capture_output=True
        )
        print("  ✓ Usuario admin creado (password: kraftdo2026)")

        # Importar Excel
        importer = os.path.join(os.path.dirname(os.path.abspath(__file__)), "importar_excel_a_mysql.py")
        if os.path.exists(importer):
            print("\n📊 Importando datos del Excel...")
            r = subprocess.run(
                ["python3", importer, empresa, output_dir],
                capture_output=True, text=True
            )
            print(r.stdout)
            if r.returncode != 0:
                print(f"  ⚠️  {r.stderr[:200]}")

        # Recalcular campos derivados (observers se ejecutan en save())
        print("\n🔁 Recalculando campos derivados...")
        r = subprocess.run(
            ["php", "artisan", "kraftdo:recalcular"],
            cwd=output_dir, capture_output=True, text=True
        )
        if r.returncode == 0:
            for line in r.stdout.strip().splitlines()[-6:]:
                print(line)
        else:
            print(f"  ⚠️  {r.stderr[:200]}")

        # Puerto único por empresa basado en hash del nombre
        import hashlib
        puerto = 8080 + int(hashlib.md5(empresa.encode()).hexdigest(), 16) % 900
        print(f"\n🚀 Panel listo en: php artisan serve --host=0.0.0.0 --port={puerto}")
        print(f"   URL: http://localhost:{puerto}/admin")
        print(f"   Email: {cfg['empresa'].get('email','admin@kraftdo.cl')}")
        print(f"   Password: kraftdo2026\n")


    return archivos


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="KraftDo Generator")
    parser.add_argument("empresa", nargs="?", help="Nombre de la empresa (ej: kraftdo)")
    parser.add_argument("--config", help="Ruta al archivo JSON de configuración")
    parser.add_argument("--output", default="./laravel_output", help="Directorio de salida")
    parser.add_argument("--preview", action="store_true", help="Mostrar sin escribir")
    parser.add_argument("--solo",
                        choices=["migraciones", "modelos", "filament", "api",
                                 "pages", "seeders", "requests", "observers", "widgets"],
                        help="Generar solo una capa")
    args = parser.parse_args()

    generar(args.empresa, args.output, args.preview, args.solo, args.config)
