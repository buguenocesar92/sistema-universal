
def _split_if_args(formula: str) -> list:
    """Divide los 3 argumentos de IF() contando paréntesis — maneja anidados."""
    # Encontrar el primer paréntesis abierto
    start = formula.index("(") + 1
    depth = 1
    args = []
    current = ""
    i = start
    while i < len(formula) and depth > 0:
        c = formula[i]
        if c == "(":
            depth += 1
            current += c
        elif c == ")":
            depth -= 1
            if depth == 0:
                args.append(current.strip())
            else:
                current += c
        elif c == "," and depth == 1:
            args.append(current.strip())
            current = ""
        else:
            current += c
        i += 1
    return args if len(args) == 3 else []

"""
KraftDo — formula_parser.py
Parsea fórmulas Excel y las convierte a:
  - Accessors de Eloquent (PHP)
  - Campos computados de Filament
  - Documentación de lógica de negocio

Cubre los patrones más comunes en Excel de PYMEs:
  =A*B              → multiplicación simple
  =A+B              → suma
  =A-B              → resta
  =A/B              → división
  =A*(1+B)          → precio con margen
  =A*(1-B)          → descuento
  =IF(cond, a, b)   → condicional
  =ROUND(x, n)      → redondeo
  =SUM(A:B)         → suma de rango (→ nota, no se convierte)
  =VLOOKUP(...)     → (→ nota, usar relación Eloquent)
"""

import re
from typing import Optional


# ── Mapeo de columna letra → nombre de campo ─────────────────────────────────
def col_a_campo(letra: str, columnas: dict) -> Optional[str]:
    """Convierte letra de columna a nombre de campo según el config."""
    letra = letra.upper()
    for campo, col_cfg in columnas.items():
        if isinstance(col_cfg, str) and col_cfg.upper() == letra:
            return campo
    return None


def cols_a_campos(formula: str, columnas: dict) -> str:
    """Reemplaza referencias de columna (A, B, C...) por nombres de campo."""
    def reemplazar(m):
        letra = m.group(1).upper()
        campo = col_a_campo(letra, columnas)
        return f"$this->{campo}" if campo else f"$this->col_{letra.lower()}"
    
    # Reemplazar referencias tipo A2, B5, C (con o sin número de fila)
    resultado = re.sub(r'\b([A-Z]+)\d*\b(?!\s*\()', reemplazar, formula)
    return resultado


# ── Convertidor de fórmulas ───────────────────────────────────────────────────
def formula_a_php(formula: str, columnas: dict) -> dict:
    """
    Convierte una fórmula Excel a código PHP.
    
    Returns:
        {
          "php":       "return $this->costo * (1 + $this->margen);",
          "tipo":      "simple" | "condicional" | "complejo" | "no_convertible",
          "descripcion": "Precio calculado aplicando margen sobre costo",
          "formula_original": "=H7*(1+K7)",
        }
    """
    f = formula.strip()
    if not f.startswith("="):
        return {"tipo": "no_formula", "php": None}
    
    f = f[1:].strip()  # quitar el =
    formula_original = f

    # Detectar referencias a otras hojas: 'Hoja'!Celda o Hoja!Celda
    if "!" in f:
        return {
            "tipo": "no_convertible",
            "php": None,
            "descripcion": "Referencia a otra hoja — implementar manualmente",
            "formula_original": formula_original,
        }

    # Fórmulas no convertibles (requieren lógica compleja)
    no_convertibles = ["VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "OFFSET",
                       "INDIRECT", "SUMIF", "COUNTIF", "AVERAGEIF"]
    for nc in no_convertibles:
        if nc in f.upper():
            return {
                "tipo":      "no_convertible",
                "php":       None,
                "descripcion": f"Fórmula compleja ({nc}) → implementar manualmente como relación Eloquent",
                "formula_original": formula_original,
            }

    # SUM de rango → no convertible directamente
    if re.search(r'SUM\s*\(', f, re.IGNORECASE):
        return {
            "tipo":      "agregado",
            "php":       None,
            "descripcion": "Suma de rango → usar scope/query en el modelo",
            "formula_original": formula_original,
        }

    # IF(condicion, valor_si, valor_no) — solo IFs simples, no anidados
    # Detectar IF anidado antes de intentar parsear
    if re.match(r'IF\s*\(', f, re.IGNORECASE):
        # Contar IFs anidados
        if f.upper().count("IF(") > 1:
            return {
                "tipo": "no_convertible",
                "php": None,
                "descripcion": "IF anidado → implementar manualmente",
                "formula_original": formula_original,
            }

    # Parser de IF con conteo de paréntesis — maneja IFs anidados
    if re.match(r'IF\s*\(', f, re.IGNORECASE):
        partes = _split_if_args(f)
        if partes and len(partes) == 3:
            cond_raw, v_si_raw, v_no_raw = partes
            cond = _expr_a_php(cond_raw, columnas)
            v_si = _expr_a_php(v_si_raw, columnas)
            v_no = _expr_a_php(v_no_raw, columnas)
            # Si alguna parte contiene IF, aplicar recursión
            if "IF(" in v_si.upper():
                v_si_conv = formula_a_php("=" + v_si_raw, columnas)
                v_si = v_si_conv["php"].replace("return ", "").rstrip(";") if v_si_conv["php"] else v_si
            if "IF(" in v_no.upper():
                v_no_conv = formula_a_php("=" + v_no_raw, columnas)
                v_no = v_no_conv["php"].replace("return ", "").rstrip(";") if v_no_conv["php"] else v_no
            php = f"return ({cond}) ? ({v_si}) : ({v_no});"
            return {
                "tipo":      "condicional",
                "php":       php,
                "descripcion": f"Valor condicional: si {cond_raw}",
                "formula_original": formula_original,
            }

    # ROUND(expr, decimales)
    round_match = re.match(r'ROUND\s*\(\s*(.+?)\s*,\s*(\d+)\s*\)$', f, re.IGNORECASE)
    if round_match:
        expr_raw, decimales = round_match.groups()
        expr = _expr_a_php(expr_raw, columnas)
        php = f"return round({expr}, {decimales});"
        return {
            "tipo":      "simple",
            "php":       php,
            "descripcion": f"Redondeo a {decimales} decimales de: {expr_raw}",
            "formula_original": formula_original,
        }

    # Fórmula aritmética simple
    php_expr = _expr_a_php(f, columnas)
    if php_expr:
        return {
            "tipo":      "simple",
            "php":       f"return {php_expr};",
            "descripcion": _describir(f, columnas),
            "formula_original": formula_original,
        }

    return {
        "tipo":      "no_convertible",
        "php":       None,
        "descripcion": "Fórmula no reconocida → implementar manualmente",
        "formula_original": formula_original,
    }


def _expr_a_php(expr: str, columnas: dict) -> str:
    """Convierte expresión aritmética Excel a PHP."""
    e = expr.strip()
    # Reemplazar referencias de celda
    e = cols_a_campos(e, columnas)
    # Limpiar + unario al inicio
    e = re.sub(r"^\+", "", e.strip())
    # Funciones Excel → PHP
    e = re.sub(r'TODAY\(\)', 'date("Y-m-d")', e, flags=re.IGNORECASE)
    e = re.sub(r'NOW\(\)', 'now()', e, flags=re.IGNORECASE)
    e = re.sub(r'ROUND\((.+?),\s*(\d+)\)', lambda m: f"round({m.group(1)}, {m.group(2)})", e, flags=re.IGNORECASE)
    e = re.sub(r'AND\((.+?)\)', lambda m: "(" + " && ".join(a.strip() for a in m.group(1).split(",")) + ")", e, flags=re.IGNORECASE)
    e = re.sub(r'OR\((.+?)\)', lambda m: "(" + " || ".join(a.strip() for a in m.group(1).split(",")) + ")", e, flags=re.IGNORECASE)
    e = re.sub(r'TEXT\((.+?),\s*["\'].*?["\']\)', lambda m: m.group(1), e, flags=re.IGNORECASE)
    e = re.sub(r'IFERROR\((.+?),\s*(.+?)\)', lambda m: f"(function(){{ try {{ return {m.group(1)}; }} catch(\$e) {{ return {m.group(2)}; }} }})()", e, flags=re.IGNORECASE)
    e = re.sub(r'LEN\((.+?)\)', lambda m: f"strlen({m.group(1)})", e, flags=re.IGNORECASE)
    e = re.sub(r'UPPER\((.+?)\)', lambda m: f"strtoupper({m.group(1)})", e, flags=re.IGNORECASE)
    e = re.sub(r'LOWER\((.+?)\)', lambda m: f"strtolower({m.group(1)})", e, flags=re.IGNORECASE)
    # Comparadores Excel → PHP
    e = e.replace("<>", "!=")
    # Operadores potencia
    e = e.replace("^", "**")
    # Limpiar espacios extra
    e = re.sub(r'\s+', ' ', e).strip()
    return e


def _describir(formula: str, columnas: dict) -> str:
    """Genera descripción legible de la fórmula."""
    patrones = [
        (r'(\w+)\s*\*\s*\(\s*1\s*\+\s*(\w+)\s*\)', "Aplicar {1} como margen sobre {0}"),
        (r'(\w+)\s*\*\s*\(\s*1\s*-\s*(\w+)\s*\)', "Aplicar descuento {1} sobre {0}"),
        (r'(\w+)\s*\*\s*0\.19',                    "IVA (19%) sobre {0}"),
        (r'(\w+)\s*\*\s*(\w+)',                    "Multiplicar {0} por {1}"),
        (r'(\w+)\s*\+\s*(\w+)',                    "Suma de {0} y {1}"),
        (r'(\w+)\s*-\s*(\w+)',                     "Diferencia entre {0} y {1}"),
        (r'(\w+)\s*/\s*(\w+)',                     "Dividir {0} entre {1}"),
    ]
    f_campos = cols_a_campos(formula, columnas)
    for patron, template in patrones:
        m = re.search(patron, f_campos)
        if m:
            return template.format(*[g.replace("$this->", "") for g in m.groups()])
    return f"Cálculo: {formula}"


# ── Analizador de hoja completa ───────────────────────────────────────────────
def analizar_formulas_hoja(ws, hoja_cfg: dict) -> list[dict]:
    """
    Lee las fórmulas de una hoja de Excel y genera accessors PHP.
    
    ws: worksheet de openpyxl (cargado SIN data_only=True)
    hoja_cfg: config de la hoja del JSON
    """
    columnas = hoja_cfg.get("columnas", {})
    fila_ini = hoja_cfg.get("fila_datos", 5)
    resultados = []

    # Buscar celdas con fórmula en las primeras 3 filas de datos
    campos_vistos = set()
    for row_idx in range(fila_ini, min(fila_ini + 3, (ws.max_row or fila_ini + 3) + 1)):
        for col_idx in range(1, (ws.max_column or 20) + 1):
            celda = ws.cell(row_idx, col_idx)
            valor = celda.value
            if not isinstance(valor, str) or not valor.startswith("="):
                continue
            
            # Determinar a qué campo corresponde esta columna
            
            def col_letra_local(idx):
                result = ""
                while idx > 0:
                    idx, rem = divmod(idx - 1, 26)
                    result = chr(65 + rem) + result
                return result
            
            letra = col_letra_local(col_idx)
            campo = col_a_campo(letra, columnas)
            if not campo or campo in campos_vistos:
                continue
            
            conversion = formula_a_php(valor, columnas)
            if conversion["tipo"] in ("simple", "condicional") and conversion["php"]:
                campos_vistos.add(campo)
                resultados.append({
                    "campo":   campo,
                    "celda":   f"{letra}{row_idx}",
                    **conversion,
                })

    return resultados


def gen_accessors_php(campo: str, conversion: dict) -> str:
    """Genera accessor PHP para Eloquent."""
    nombre_accessor = "".join(w.capitalize() for w in campo.split("_")) + "Computed"
    return (
        f"    /**\n"
        f"     * {conversion.get('descripcion', campo)}\n"
        f"     * Fórmula Excel: ={conversion.get('formula_original', '')}\n"
        f"     */\n"
        f"    public function get{nombre_accessor}Attribute()\n"
        f"    {{\n"
        f"        {conversion['php']}\n"
        f"    }}"
    )


# ── Analizar Excel completo ───────────────────────────────────────────────────
def analizar_excel_formulas(excel_path: str, cfg: dict) -> dict:
    """
    Lee el Excel con fórmulas (no data_only) y extrae todas las conversiones.
    Retorna dict: {alias_hoja: [conversiones]}
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        return {}

    resultado = {}
    for alias, hoja_cfg in cfg.get("hojas", {}).items():
        nombre_hoja = hoja_cfg.get("nombre", "")
        ws = next((wb[h] for h in wb.sheetnames if nombre_hoja.lower() in h.lower()), None)
        if not ws:
            continue
        
        conversiones = analizar_formulas_hoja(ws, hoja_cfg)
        if conversiones:
            resultado[alias] = conversiones

    return resultado


if __name__ == "__main__":
    import json, sys
    sys.path.insert(0, '.')
    cfg = json.load(open("empresas/kraftdo.json"))
    formulas = analizar_excel_formulas("KraftDo_BD_Maestra_v5.xlsx", cfg)
    
    if not formulas:
        print("No se encontraron fórmulas convertibles")
    else:
        for alias, convs in formulas.items():
            print(f"\n── {alias} ──")
            for c in convs:
                print(f"  {c['campo']}: ={c['formula_original']}")
                print(f"  PHP: {c['php']}")
                print(f"  Tipo: {c['tipo']} | {c['descripcion']}")
