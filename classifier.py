"""
KraftDo — classifier.py
UI web para clasificar hojas de un Excel y generar el JSON de configuración.

USO:
    python3 classifier.py          → abre en http://localhost:8001
    python3 classifier.py --port 8080

FLUJO:
    1. Sube un archivo Excel
    2. La app detecta todas las hojas y muestra una preview de cada una
    3. Clasificas cada hoja: Modelo | Vista | Config | Ignorar
    4. Defines el nombre de la empresa y configuración básica
    5. Descarga el JSON listo para usar con core.py y generator.py
"""

import os
import sys
import json
import re
import argparse
import tempfile
from pathlib import Path
from datetime import datetime

from typing import Any, Optional
import openpyxl

# FastAPI/Pydantic son opcionales: classifier funciona como librería
# (analizar_excel + generar_json) incluso sin la UI web instalada.
try:
    from fastapi import FastAPI, UploadFile, File, HTTPException, Request
    from fastapi.responses import HTMLResponse, JSONResponse
    from fastapi.middleware.cors import CORSMiddleware
    from pydantic import BaseModel
    _HAS_WEB = True
except ImportError:
    _HAS_WEB = False
    class BaseModel:  # type: ignore
        pass

# Modelos para consolidación y Fase 0
class DescripcionNegocio(BaseModel):
    empresa: str
    descripcion: str
    hojas_meta: dict = {}

class GrupoConsolidacion(BaseModel):
    nombre_entidad: str
    aliases: list[str]
    empresa_cfg: dict = {}

if _HAS_WEB:
    app = FastAPI(title="KraftDo Classifier", docs_url=None)
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
else:
    # Stub: permite cargar el archivo en modo librería (sin uvicorn/fastapi)
    # para usar analizar_excel / generar_json sin instalar la web stack.
    class _AppStub:
        def _noop(self, *_a, **_kw):
            def deco(fn): return fn
            return deco
        get = post = put = delete = _noop  # type: ignore
        def add_middleware(self, *_a, **_kw): pass
    app = _AppStub()
    HTMLResponse = JSONResponse = lambda *a, **kw: None  # type: ignore
    File = lambda *a, **kw: None  # type: ignore
    UploadFile = type("UploadFile", (), {})  # type: ignore
    HTTPException = Exception  # type: ignore
    Request = type("Request", (), {})  # type: ignore

class GenerarRequest(BaseModel):
    empresa: dict
    clasificaciones: dict
    hojas_meta: dict
    grupos_consolidacion: list = []

UPLOAD_DIR = tempfile.mkdtemp()
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Helpers ──────────────────────────────────────────────────────────────────
def col_letra(idx: int) -> str:
    """Índice 0-based → letra de columna."""
    result = ""
    idx += 1
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result

def detectar_tipo_col(nombre: str) -> str:
    """Inferir tipo de columna para el JSON."""
    n = nombre.lower()
    if any(p in n for p in ("precio", "monto", "costo", "total", "saldo", "anticipo", "ganancia", "ahorro")):
        return "decimal"
    if any(p in n for p in ("cantidad", "stock", "dias", "horas", "gramos")):
        return "integer"
    if any(p in n for p in ("fecha", "f_", "_at")):
        return "date"
    if any(p in n for p in ("estado", "tipo", "categoria", "canal")):
        return "select"
    if any(p in n for p in ("obs", "descripcion", "notas", "detalle")):
        return "text"
    if any(p in n for p in ("correo", "email")):
        return "email"
    if any(p in n for p in ("telefono", "whatsapp", "phone")):
        return "phone"
    return "string"

# ────────────────────────────────────────────────────────────────────────
# Detección de capas avanzadas (v25-fase1)
# Cada función infiere una capa específica del JSON sin depender de las otras.
# ────────────────────────────────────────────────────────────────────────

# Vista lateral: campos que describen al padre, sin valor histórico para el hijo.
ACCESSOR_HINTS = {
    "whatsapp", "telefono", "fono", "email", "correo", "ciudad",
    "direccion", "region", "comuna", "rut",
}
# Snapshot: campos cuyo valor en el momento de creación importa históricamente.
SNAPSHOT_HINTS = {
    "precio", "precio_unit", "precio_unitario", "precio_mayor",
    "categoria", "nombre", "tiempo_produccion", "plazo",
    "costo_unit", "costo",
}
# Hojas que típicamente sirven como "agregado" cuando contienen sumas
# de otras y un campo "disponible/stock/saldo".
AGREGADO_HINTS_DISPONIBLE = {"disponible", "stock", "stock_disponible", "saldo"}
# Reglas de cálculo conocidas — espejo de generator.REGLAS_CALCULO.
# Solo necesitamos los nombres + descripciones para proponer en el JSON.
REGLAS_CALCULO_CONOCIDAS = {
    "costo_total":      ["costo_insumo", "hora_trabajo"],
    "precio_unit":      ["costo_total"],
    "precio_mayor":     ["costo_total"],
    "valor_dia":        ["sueldo_base", "dias_laborales"],
    "descuento_faltas": ["valor_dia", "faltas"],
    "a_pagar":          ["sueldo_base", "descuento_faltas"],
    "saldo":            ["a_pagar", "quincena_pagada"],
    "iva":              ["neto"],
    "iva_servicio":     ["total_neto"],
    "total":            ["neto"],
    "total_neto":       ["costo_china", "embarcadero", "agente_aduana"],
    "stock_disponible": ["importacion", "ventas"],
    "total_gastado":    ["materiales", "mano_obra"],
    "resultado":        ["cobrado", "total_gastado"],
    "margen":           ["cobrado", "resultado"],
    "neto_dsto":        ["neto"],
}


def _levenshtein(a: str, b: str) -> int:
    """Distancia Levenshtein iterativa (sin deps externas)."""
    if a == b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i] + [0] * len(b)
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            cur[j] = min(cur[j-1] + 1, prev[j] + 1, prev[j-1] + cost)
        prev = cur
    return prev[-1]


def _valores_unicos_columna(ws, col_letter: str, fila_datos: int,
                             max_filas: int = 200) -> list:
    """Devuelve hasta max_filas valores únicos no-vacíos de una columna."""
    try:
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
    except Exception:
        return []
    vistos = []
    seen = set()
    fin = min(ws.max_row or fila_datos, fila_datos + max_filas)
    for r in range(fila_datos, fin + 1):
        v = ws.cell(r, col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        vistos.append(s)
    return vistos


def detectar_formato_id(columnas: dict, ws, fila_datos: int,
                         alias_hoja: str = "") -> dict | None:
    """1a) Detecta columna con valores tipo PREFIJO-NNN. Solo propone
    formato_id si encuentra el patrón en datos REALES o si el campo es
    claramente un identificador autonumérico (n_, n_pedido, folio, num_)
    Y la hoja parece transaccional (no catálogo).
    Devuelve {campo, formato_id} o None."""
    HINTS_AUTO = ("n_pedido", "n_factura", "folio", "n_orden", "n_orden_de_compra")
    patron = re.compile(r'^([A-Z]{2,6})-(\d{2,4})$')
    for campo, letra in columnas.items():
        # Normalizar dobles underscores que vienen de slugify de "Nº PEDIDO"
        campo_norm = re.sub(r'_+', '_', campo).strip("_")
        valores = _valores_unicos_columna(ws, letra, fila_datos, 60)
        prefijos = []
        for v in valores:
            m = patron.match(str(v).strip())
            if m:
                prefijos.append((m.group(1), len(m.group(2))))
        if prefijos:
            from collections import Counter
            p_more = Counter(p for p, _ in prefijos).most_common(1)[0][0]
            ancho  = Counter(w for _, w in prefijos).most_common(1)[0][0]
            return {"campo": campo, "formato_id": f"{p_more}-{{:0{ancho}d}}"}
        # Sin valores pero el nombre es claramente un id auto
        if not valores and any(h == campo_norm or h in campo_norm for h in HINTS_AUTO):
            base = re.sub(r'[^a-zA-Z]', '', alias_hoja)[:3].upper() or "ID"
            return {"campo": campo, "formato_id": f"{base}-{{:03d}}"}
    return None


def detectar_fk_y_padres(alias_hoja: str, columnas: dict,
                          todas_hojas: dict) -> list:
    """1b) Lista de FKs detectadas: cada item con info del padre.
    Hace match flexible: campo='cliente' encuentra alias 'clientes'
    o 'control_clientes', etc."""
    rels = []
    for campo in columnas:
        if campo == "id":
            continue
        for alias_b, meta_b in todas_hojas.items():
            if alias_b == alias_hoja or meta_b.get("vacia"):
                continue
            ident_b = meta_b.get("identificador")
            cols_b  = meta_b.get("columnas", {}) or {}
            sing_b  = alias_b.rstrip("s")
            es_fk = (
                (ident_b and campo == ident_b)
                or campo == alias_b
                or campo == sing_b
                or alias_b.endswith("_" + campo)            # control_clientes ← cliente
                or alias_b.endswith("_" + campo + "s")      # control_pedidos ← pedido
                or sing_b.endswith("_" + campo)
            )
            if es_fk:
                rels.append({
                    "campo_origen":  campo,
                    "alias_padre":   alias_b,
                    "ident_padre":   ident_b,
                    "cols_padre":    set(cols_b.keys()),
                })
                break
    return rels


def proponer_accessor_y_snapshot(columnas: dict, fks: list) -> tuple[dict, list, set]:
    """1b) A partir de las FKs, decide qué columnas mover a campos_accessor
    o snapshot_at_create. Devuelve (columnas_finales, snapshot_lista,
    accessors_set)."""
    accessors = []
    snapshots = []
    a_quitar  = set()
    nombres_cols = list(columnas.keys())

    for fk in fks:
        cols_padre = fk["cols_padre"]
        for campo in nombres_cols:
            if campo == fk["campo_origen"]:
                continue
            # ¿Existe columna con ese nombre en el padre?
            if campo not in cols_padre:
                continue
            if campo in ACCESSOR_HINTS and campo not in accessors:
                accessors.append(campo)
                a_quitar.add(campo)
            elif campo in SNAPSHOT_HINTS and campo not in snapshots:
                snapshots.append(campo)
                a_quitar.add(campo)

    cols_finales = {k: v for k, v in columnas.items() if k not in a_quitar}
    return cols_finales, snapshots, set(accessors)


def detectar_agregado(alias: str, columnas: dict, todas_hojas: dict) -> dict | None:
    """1c) Si la hoja tiene cols cuyo nombre coincide con (substring de)
    otros alias y un campo 'disponible/stock/saldo/total' → tipo agregado
    con fuentes."""
    nombres_cols = set(columnas.keys())
    if not (nombres_cols & AGREGADO_HINTS_DISPONIBLE) and "total" not in nombres_cols:
        return None

    def _stem(s: str) -> str:
        """Reduce un slug a su 'núcleo': sin underscores múltiples ni vocales
        sueltas finales ('importaci_n' → 'importacion')."""
        s = re.sub(r'_+', '_', s).strip("_")
        # Casos comunes: importaci_n → importacion, etc.
        s = s.replace("ci_n", "cion").replace("si_n", "sion")
        return s

    def _buscar_alias_para(campo: str) -> str | None:
        """Match flexible: 'ventas' encuentra 'control_de_ventas',
        'importaci_n' encuentra 'control_importacion', etc."""
        c = _stem(campo)
        if c in todas_hojas: return c
        if c + "s" in todas_hojas: return c + "s"
        sing = c.rstrip("s")
        for alias_b in todas_hojas:
            if alias_b == alias:
                continue
            ab_stem = _stem(alias_b)
            if c in ab_stem or sing in ab_stem or ab_stem.endswith("_" + c):
                return alias_b
        return None

    def _es_disponible(c: str) -> bool:
        c = _stem(c)
        return c in AGREGADO_HINTS_DISPONIBLE or c == "total" or "disponible" in c

    fuentes = []
    for campo in nombres_cols:
        if _es_disponible(campo) or campo == "total":
            continue
        alias_real = _buscar_alias_para(campo)
        if not alias_real:
            continue
        candidato = todas_hojas.get(alias_real, {})
        if candidato.get("vacia"):
            continue
        cols_fuente = candidato.get("columnas", {}) or {}
        cols_stems  = {_stem(k): k for k in cols_fuente}
        campo_grupo = next((cols_stems[s] for s in ("modelo", "codigo", "sku") if s in cols_stems), None)
        campo_valor = next(
            (cols_stems[s] for s in ("cantidad", "cant", "unidades", "unidad", "qty", "monto", "total")
             if s in cols_stems),
            None
        )
        if not (campo_grupo and campo_valor):
            continue
        fuentes.append({
            "hoja":         alias_real,
            "campo_grupo":  campo_grupo,
            "campo_valor":  campo_valor,
            "destino":      campo,
        })
    if not fuentes:
        return None

    grupo = next((c for c in ("modelo", "codigo", "sku") if c in nombres_cols), "modelo")
    return {"agrupar_por": grupo, "fuentes": fuentes}


def detectar_sinonimos_modelo(hojas_meta: dict, wb) -> dict:
    """1d) Compara valores únicos de columnas modelo/codigo/sku entre
    hojas y agrupa similares. Devuelve {canonico: [variantes]}."""
    valores_por_hoja = {}
    for alias, meta in hojas_meta.items():
        if meta.get("vacia"):
            continue
        cols = meta.get("columnas", {}) or {}
        nombre_hoja = meta.get("nombre")
        ws = wb[nombre_hoja] if nombre_hoja in wb.sheetnames else None
        if not ws:
            continue
        for campo in ("modelo", "codigo", "sku"):
            if campo not in cols:
                continue
            valores = _valores_unicos_columna(
                ws, cols[campo], meta.get("fila_datos", 5), 200
            )
            valores_por_hoja[(alias, campo)] = valores
            break

    todos = []
    for vals in valores_por_hoja.values():
        for v in vals:
            if v not in todos:
                todos.append(v)
    if len(todos) < 2:
        return {}

    def _son_sinonimos(a: str, b: str) -> bool:
        """Criterio estricto: misma cadena difiriendo solo en case, o uno
        contiene al otro como substring distinto, o levenshtein <= 1 sobre
        cadenas de >= 4 chars (evita false positives entre A01/A02)."""
        al, bl = a.lower(), b.lower()
        if al == bl and a != b:                 # solo case
            return True
        if al != bl and (al in bl or bl in al): # substring
            return True
        if min(len(al), len(bl)) >= 4 and _levenshtein(al, bl) == 1:
            return True
        return False

    grupos = []
    for v in todos:
        encontrado = False
        for g in grupos:
            if any(_son_sinonimos(v, w) for w in g["variantes"]):
                g["variantes"].add(v)
                if len(v) < len(g["canonico"]):
                    g["canonico"] = v
                encontrado = True
                break
        if not encontrado:
            grupos.append({"canonico": v, "variantes": {v}})

    sinonimos = {}
    for g in grupos:
        if len(g["variantes"]) >= 2:
            sinonimos[g["canonico"]] = sorted(g["variantes"])
    return sinonimos


def detectar_matriz_asistencia(meta: dict, ws) -> dict | None:
    """1e) Hoja con >= 15 cols cuyo header es número 1..31 o letra A/F/L
    → tipo matriz_asistencia. col_trabajador se infiere como la primera
    columna a la izquierda de las cols-día que tenga texto en r4 o r5
    (segunda fila de header en formato 'multi_header')."""
    from openpyxl.utils import get_column_letter, column_index_from_string

    fila_h = meta.get("fila_headers", 1)
    cols_dia = []
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        v = ws.cell(fila_h, c).value
        if v is None:
            continue
        s = str(v).strip()
        try:
            n = int(float(s))
            if 1 <= n <= 31:
                cols_dia.append(get_column_letter(c))
                continue
        except (ValueError, TypeError):
            pass
    if len(cols_dia) < 15:
        return None

    # col_trabajador = la primera col a la izquierda de cols_dia con texto
    # en alguna de las primeras 5 filas (catch headers multi-fila o nombres).
    primer_dia_idx = column_index_from_string(cols_dia[0])
    col_trabajador = None
    cols_persona_hints = ("personal", "trabajador", "persona", "empleado", "nombre")
    for c in range(primer_dia_idx - 1, 0, -1):
        for r in range(1, min(6, ws.max_row + 1) if ws.max_row else 6):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if any(h in s for h in cols_persona_hints):
                col_trabajador = get_column_letter(c)
                break
        if col_trabajador:
            break
    if not col_trabajador:
        # Fallback: simplemente la columna inmediatamente a la izquierda.
        col_trabajador = get_column_letter(max(primer_dia_idx - 1, 1))

    n = len(cols_dia)
    mitad = (n + 1) // 2
    return {
        "fila_inicio":          meta.get("fila_datos", 5),
        "fila_fin":             ws.max_row or meta.get("fila_datos", 5),
        "col_codigo_obra":      None,
        "col_obra":             None,
        "col_trabajador":       col_trabajador,
        "cols_quincena1":       cols_dia[:mitad],
        "cols_quincena2":       cols_dia[mitad:],
        "col_pago_quincena":    None,
        "col_pago_liquidacion": None,
        "mes_actual":           "",
    }


def detectar_reglas_desde_formulas(ws, columnas: dict,
                                    fila_datos: int) -> list:
    """1f) Recorre formulas Excel y propone reglas calc conocidas."""
    try:
        from formula_parser import formula_a_php
    except Exception:
        formula_a_php = None

    # Mapa letra → campo
    letra_a_campo = {v: k for k, v in columnas.items()}
    encontradas = set()
    propuestas = []

    fin = min(ws.max_row or fila_datos, fila_datos + 50)
    from openpyxl.utils import get_column_letter
    for r in range(fila_datos, fin + 1):
        for letra_col, campo in letra_a_campo.items():
            try:
                from openpyxl.utils import column_index_from_string
                c = column_index_from_string(letra_col)
            except Exception:
                continue
            v = ws.cell(r, c).value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            if campo in encontradas:
                continue
            # ¿Es una regla conocida cuyas deps están todas en columnas?
            if campo in REGLAS_CALCULO_CONOCIDAS:
                deps = REGLAS_CALCULO_CONOCIDAS[campo]
                if all(d in columnas for d in deps):
                    propuestas.append({
                        "campo":         campo,
                        "deps":          deps,
                        "formula_excel": v[:120],
                        "implementado":  True,
                    })
                    encontradas.add(campo)
            else:
                propuestas.append({
                    "campo":         campo,
                    "deps":          [],
                    "formula_excel": v[:120],
                    "implementado":  False,
                })
                encontradas.add(campo)
    return propuestas


def analizar_excel(path: str) -> dict:
    """Analiza el Excel y retorna metadata de cada hoja."""
    wb = openpyxl.load_workbook(path, data_only=True)
    hojas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0

        # Detectar fila de headers (primera con 3+ celdas llenas)
        fila_headers = None
        for r in range(1, min(8, max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            if sum(1 for v in row_vals if v is not None) >= 3:
                fila_headers = r
                break

        if not fila_headers:
            hojas[sheet_name] = {
                "nombre":       sheet_name,
                "vacia":        True,
                "filas":        max_row,
                "columnas":     {},
                "preview":      [],
                "fila_headers": 1,
                "fila_datos":   2,
            }
            continue

        # Leer headers
        headers = {}
        for c in range(1, max_col + 1):
            val = ws.cell(fila_headers, c).value
            if val:
                nombre_limpio = re.sub(r'\[.*?\]|\n.*', '', str(val)).strip()
                headers[col_letra(c - 1)] = nombre_limpio

        # Preview: primeras 3 filas de datos
        fila_datos = fila_headers + 1
        preview = []
        from openpyxl.utils import column_index_from_string
        for r in range(fila_datos, min(fila_datos + 3, max_row + 1)):
            fila = {}
            for letra, nombre in headers.items():
                try:
                    col_idx = column_index_from_string(letra)
                except Exception:
                    continue
                v = ws.cell(r, col_idx).value
                fila[nombre] = str(v)[:30] if v is not None else ""
            if any(fila.values()):
                preview.append(fila)

        # Clasificación automática heurística
        n_filas_datos = max_row - fila_datos
        n_formulas = sum(
            1 for r in range(fila_headers, min(fila_headers + 5, max_row + 1))
            for c in range(1, max_col + 1)
            if isinstance(ws.cell(r, c).value, str)
            and str(ws.cell(r, c).value).startswith("=")
        )
        # Plantillas por nombre de hoja
        nombre_lower = sheet_name.lower().replace(" ", "").replace("_", "")
        plantilla = None
        PLANTILLAS = {
            "catalogo":  ["producto", "catálogo", "catalog", "precio", "artículo", "item", "inventario", "stock", "nfc", "sublimacion", "impresion"],
            "registros": ["cliente", "pedido", "orden", "venta", "caja", "pago", "factura", "proveedor", "trabajador", "empleado", "produccion"],
            "kpis":      ["resumen", "dashboard", "kpi", "inicio", "inicio", "summary", "report"],
            "vista":     ["cotizador", "calculador", "template", "formato", "plantilla"],
        }
        for tipo_p, palabras in PLANTILLAS.items():
            if any(p in nombre_lower for p in palabras):
                plantilla = tipo_p
                break

        clasificacion_auto = plantilla or (
            "kpis"      if n_filas_datos < 5 and max_col < 8 else
            "vista"     if n_formulas > 3 else
            "registros" if n_filas_datos > 5 else
            "catalogo"
        )

        # Slug fuerte: normaliza unicode (á→a, ñ→n), quita anotaciones tipo
        # [AUTO]/(UND)/$, colapsa underscores múltiples y arregla casos
        # comunes (categor_a → categoria) para que las heurísticas de
        # detección funcionen sin tropezar con ruido tipográfico.
        import unicodedata
        def _slug_fuerte(s: str) -> str:
            s = unicodedata.normalize("NFKD", str(s))
            s = s.encode("ascii", "ignore").decode("ascii")
            s = re.sub(r'\(.*?\)|\[.*?\]', '', s)  # parentéticos
            s = s.lower().replace(" ", "_").replace("/", "_").replace("$", "")
            s = re.sub(r'[^a-z0-9_]', '_', s)
            s = re.sub(r'_+', '_', s).strip("_")
            return s or "col"
        cols_slug = {_slug_fuerte(k): letra for letra, k in headers.items()}

        # Identificador candidato (heurística)
        ident = None
        for cand in ("n_pedido", "numero", "sku", "codigo", "id", "nombre", "trabajador", "evento", "modelo"):
            if cand in cols_slug:
                ident = cand
                break

        meta_hoja = {
            "nombre":           sheet_name,
            "vacia":            False,
            "filas":            n_filas_datos,
            "columnas_raw":     headers,
            "plantilla":        plantilla or "heuristica",
            "columnas":         cols_slug,
            "identificador":    ident,
            "preview":          preview,
            "fila_headers":     fila_headers,
            "fila_datos":       fila_datos,
            "clasificacion":    clasificacion_auto,
            "tipos_col":        {
                re.sub(r'[^a-z0-9_]', '_', k.lower().replace(" ", "_")): detectar_tipo_col(k)
                for k in headers.values()
            }
        }

        # Capas avanzadas que dependen solo de la hoja en sí
        meta_hoja["formato_id_propuesto"]    = detectar_formato_id(cols_slug, ws, fila_datos, sheet_name)
        meta_hoja["formulas_detectadas"]     = detectar_reglas_desde_formulas(ws, cols_slug, fila_datos)
        matriz = detectar_matriz_asistencia(meta_hoja, ws)
        if matriz:
            meta_hoja["matriz_asistencia"] = matriz
            meta_hoja["clasificacion"] = "matriz_asistencia"

        hojas[sheet_name] = meta_hoja

    # Capas que requieren mirar todas las hojas a la vez:
    # detectar_fk_y_padres / accessor / snapshot / agregado.
    # Indexamos por slug fuerte (sin emojis, unicode normalizado, single _).
    import unicodedata as _ud
    def _alias_slug(s: str) -> str:
        s = _ud.normalize("NFKD", str(s))
        s = s.encode("ascii", "ignore").decode("ascii")
        s = re.sub(r'\(.*?\)|\[.*?\]', '', s)
        s = s.lower().replace(" ", "_")
        s = re.sub(r'[^a-z0-9_]', '_', s)
        s = re.sub(r'_+', '_', s).strip("_")
        return s or "hoja"
    indexado = {}
    for sn, m in hojas.items():
        indexado[_alias_slug(sn)] = m

    for sn, meta in hojas.items():
        if meta.get("vacia"):
            continue
        cols   = meta.get("columnas", {}) or {}
        alias  = _alias_slug(sn)
        fks    = detectar_fk_y_padres(alias, cols, indexado)
        meta["fks_detectadas"] = fks

        cols_finales, snapshots, accessors = proponer_accessor_y_snapshot(cols, fks)
        if accessors or snapshots:
            meta["columnas_post_accessor"] = cols_finales
            meta["accessors_propuestos"]   = sorted(accessors)
            meta["snapshots_propuestos"]   = snapshots

        agregado = detectar_agregado(alias, cols, indexado)
        if agregado:
            meta["agregado_propuesto"] = agregado
            meta["clasificacion"] = "agregado"

    # Sinónimos a nivel global del Excel
    hojas["__sinonimos_modelo__"] = detectar_sinonimos_modelo(hojas, wb)
    return hojas


def generar_json(empresa_data: dict, hojas_clasificadas: dict, hojas_meta: dict) -> dict:
    """Genera el JSON de configuración final con capas avanzadas (v25-fase1)."""
    hojas_json = {}
    sinonimos_globales = hojas_meta.get("__sinonimos_modelo__", {}) or {}

    for alias, clasificacion in hojas_clasificadas.items():
        if clasificacion in ("ignorar", "vista"):
            continue

        meta = hojas_meta.get(alias, {})
        if meta.get("vacia"):
            continue

        # Si el classifier detectó accessors/snapshots, las cols_finales
        # ya excluyen esos campos. Si no, usamos las columnas crudas.
        columnas = meta.get("columnas_post_accessor") or meta.get("columnas", {})

        # Si la hoja es matriz_asistencia, su contrato de JSON es distinto.
        if clasificacion == "matriz_asistencia" or meta.get("matriz_asistencia"):
            mat = meta.get("matriz_asistencia") or {}
            cfg_hoja = {
                "nombre":               meta.get("nombre", alias),
                "tipo":                 "matriz_asistencia",
                "fila_inicio":          mat.get("fila_inicio", meta.get("fila_datos", 5)),
                "fila_fin":             mat.get("fila_fin",    meta.get("fila_datos", 5) + 14),
                "col_codigo_obra":      mat.get("col_codigo_obra"),
                "col_obra":             mat.get("col_obra"),
                "col_trabajador":       mat.get("col_trabajador"),
                "cols_quincena1":       mat.get("cols_quincena1", []),
                "cols_quincena2":       mat.get("cols_quincena2", []),
                "col_pago_quincena":    mat.get("col_pago_quincena"),
                "col_pago_liquidacion": mat.get("col_pago_liquidacion"),
                "mes_actual":           mat.get("mes_actual", ""),
            }
            alias_limpio = re.sub(r'[^a-z0-9_]', '_', alias.lower().replace(" ", "_"))
            hojas_json[alias_limpio] = cfg_hoja
            continue

        cfg_hoja = {
            "nombre":      meta.get("nombre", alias),
            "tipo":        clasificacion,
            "descripcion": f"Hoja {alias}",
            "fila_datos":  meta.get("fila_datos", 5),
            "columnas":    columnas,
        }

        # Identificador automático (heurística amplia).
        # Prioridad: si hay formato_id_propuesto, ese campo es el ident.
        cols = list(columnas.keys())
        fid  = meta.get("formato_id_propuesto")

        ident = None
        if fid and fid["campo"] in cols:
            ident = fid["campo"]
        elif meta.get("identificador") in cols:
            ident = meta["identificador"]
        else:
            for candidato in ("n_pedido", "n__pedido", "numero", "nombre",
                              "sku", "codigo", "trabajador", "evento",
                              "modelo", "item", "proveedor"):
                if candidato in cols:
                    ident = candidato
                    break
        if ident:
            cfg_hoja["identificador"] = ident

        # 1a) formato_id si el ident coincide con el campo detectado
        if fid and fid["campo"] == ident:
            cfg_hoja["formato_id"] = fid["formato_id"]

        # 1b) campos_accessor / snapshot_at_create
        accessors = meta.get("accessors_propuestos") or []
        snapshots = meta.get("snapshots_propuestos") or []
        if accessors:
            cfg_hoja["campos_accessor"] = list(accessors)
        if snapshots:
            cfg_hoja["snapshot_at_create"] = list(snapshots)

        # 1c) agregado con fuentes
        agg = meta.get("agregado_propuesto")
        if agg and clasificacion == "agregado":
            cfg_hoja["agrupar_por"] = agg["agrupar_por"]
            cfg_hoja["fuentes"]     = agg["fuentes"]

        # 1f) reglas_detectadas — fórmulas Excel mapeadas a REGLAS_CALCULO
        formulas = meta.get("formulas_detectadas") or []
        if formulas:
            cfg_hoja["reglas_detectadas"] = formulas

        # Filtro activo para catálogos con campo estado
        if clasificacion == "catalogo" and "estado" in cols:
            cfg_hoja["filtro_activo"] = {"columna": "estado", "valor": "Activo"}
            # Detectar campos de precio reales desde las columnas
            cols_precio_1 = next((c for c in cols if "precio" in c and ("unit" in c or c.endswith("_1") or "1" in c)), None)
            cols_precio_5 = next((c for c in cols if "precio" in c and "5" in c), None)
            cols_precio_10 = next((c for c in cols if "precio" in c and "10" in c), None)
            precios_cfg = {}
            if cols_precio_1: precios_cfg["1"] = cols_precio_1
            if cols_precio_5: precios_cfg["5"] = cols_precio_5
            if cols_precio_10: precios_cfg["10"] = cols_precio_10
            if precios_cfg:
                cfg_hoja["precios"] = precios_cfg

        # Filtro activos para registros con campo estado
        if clasificacion == "registros" and "estado" in cols:
            cfg_hoja["filtro_activos"] = {
                "columna": "estado",
                "valores": ["Activo", "Confirmado", "En producción"]
            }

        # KPIs para hojas de tipo kpis
        if clasificacion == "kpis":
            cfg_hoja.pop("columnas", None)
            cfg_hoja.pop("fila_datos", None)
            cfg_hoja["kpis"] = {}  # el usuario completará las celdas

        alias_limpio = re.sub(r'[^a-z0-9_]', '_', alias.lower().replace(" ", "_"))
        hojas_json[alias_limpio] = cfg_hoja

    cfg_final = {
        "empresa": {
            "nombre":       empresa_data.get("nombre") or empresa_data.get("archivo", "").replace(".xlsx","").replace("_"," ").title() or "Mi Empresa",
            "rut":          empresa_data.get("rut", ""),
            "email":        empresa_data.get("email", ""),
            "telefono":     empresa_data.get("telefono", ""),
            "web":          empresa_data.get("web", ""),
            "color_primary": empresa_data.get("color_primary", "1E3A5F"),
            "color_accent":  empresa_data.get("color_accent", "F5A623"),
        },
        "fuente": {
            "tipo":      "local",
            "archivo":   empresa_data.get("archivo", "BD_Maestra.xlsx"),
            "sheets_id": "",
            "creds_json": "creds.json",
        },
        "logica_negocios": {
            "iva":          0.19,
            "anticipo_pct": 0.5,
            "moneda":       "CLP",
        },
        "hojas": hojas_json,
    }
    # 1d) sinónimos a nivel raíz si se detectaron
    if sinonimos_globales:
        cfg_final["sinonimos_modelo"] = sinonimos_globales
    return cfg_final


# ── Rutas Flask ──────────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>KraftDo — Clasificador de Excel</title>
<style>
:root{--primary:#1E3A5F;--accent:#F5A623;--bg:#f8f7f4;--card:#fff;--border:#e5e4dd;--text:#1c1c1a;--muted:#6b6b67}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
header{background:var(--primary);color:#fff;padding:1rem 2rem;display:flex;align-items:center;gap:1rem}
header h1{font-size:1.2rem;font-weight:500}
header span{font-size:.85rem;opacity:.7}
.container{max-width:900px;margin:0 auto;padding:2rem 1rem}
.card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:1.5rem;margin-bottom:1rem}
.card h2{font-size:.9rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin-bottom:1rem}
.upload-zone{border:2px dashed var(--border);border-radius:8px;padding:3rem;text-align:center;cursor:pointer;transition:.2s}
.upload-zone:hover{border-color:var(--accent);background:#fffbf0}
.upload-zone input{display:none}
.btn{display:inline-flex;align-items:center;gap:.5rem;padding:.6rem 1.2rem;border-radius:8px;border:none;cursor:pointer;font-size:.9rem;font-weight:500;transition:.15s}
.btn-primary{background:var(--primary);color:#fff}
.btn-primary:hover{background:#16304f}
.btn-accent{background:var(--accent);color:#fff}
.btn-accent:hover{background:#e09520}
.btn-outline{background:transparent;border:1px solid var(--border);color:var(--text)}
.btn-outline:hover{background:var(--bg)}
.grid-empresa{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
@media(max-width:560px){.grid-empresa{grid-template-columns:1fr}}
.field{display:flex;flex-direction:column;gap:.4rem}
.field label{font-size:.8rem;color:var(--muted);font-weight:500}
.field input{padding:.5rem .75rem;border:1px solid var(--border);border-radius:6px;font-size:.9rem;background:var(--bg)}
.field input:focus{outline:none;border-color:var(--primary)}
.hoja-card{border:1px solid var(--border);border-radius:8px;margin-bottom:.75rem;overflow:hidden}
.hoja-header{display:flex;align-items:center;gap:1rem;padding:.75rem 1rem;background:var(--bg);cursor:pointer}
.hoja-nombre{font-weight:500;flex:1}
.hoja-meta{font-size:.8rem;color:var(--muted)}
.tipos{display:flex;gap:.5rem;flex-wrap:wrap}
.tipo-btn{padding:.3rem .8rem;border-radius:999px;font-size:.8rem;border:1px solid var(--border);cursor:pointer;transition:.15s;background:#fff}
.tipo-btn:hover{border-color:var(--primary)}
.tipo-btn.activo{background:var(--primary);color:#fff;border-color:var(--primary)}
.tipo-btn[data-tipo="catalogo"].activo{background:#059669}
.tipo-btn[data-tipo="registros"].activo{background:#0369a1}
.tipo-btn[data-tipo="kpis"].activo{background:#7c3aed}
.tipo-btn[data-tipo="vista"].activo{background:#b45309}
.tipo-btn[data-tipo="ignorar"].activo{background:#9ca3af}
.preview-table{width:100%;font-size:.75rem;border-collapse:collapse;overflow:auto}
.preview-table th{background:var(--bg);padding:.4rem .5rem;text-align:left;border-bottom:1px solid var(--border);color:var(--muted);white-space:nowrap}
.preview-table td{padding:.3rem .5rem;border-bottom:1px solid #f0efe8;color:var(--text)}
.hoja-body{padding:1rem;display:none}
.hoja-body.open{display:block}
.preview-wrap{overflow-x:auto;margin-top:.75rem;border-radius:6px;border:1px solid var(--border)}
.step{display:none}.step.active{display:block}
.steps-nav{display:flex;gap:.5rem;margin-bottom:1.5rem}
.step-dot{width:28px;height:28px;border-radius:50%;background:var(--border);display:flex;align-items:center;justify-content:center;font-size:.8rem;font-weight:600;cursor:default}
.step-dot.done{background:var(--primary);color:#fff}
.step-dot.active{background:var(--accent);color:#fff}
.actions{display:flex;gap:.75rem;margin-top:1.5rem}
.badge{display:inline-block;padding:.2rem .6rem;border-radius:999px;font-size:.75rem;font-weight:500}
.badge-green{background:#d1fae5;color:#065f46}
.badge-blue{background:#dbeafe;color:#1e40af}
.badge-purple{background:#ede9fe;color:#5b21b6}
.badge-gray{background:#f3f4f6;color:#6b7280}
#resultado{display:none}
#resultado pre{background:#1c1c1a;color:#a3e635;padding:1rem;border-radius:8px;font-size:.75rem;overflow:auto;max-height:400px}
.progreso{height:4px;background:var(--border);border-radius:2px;margin-bottom:1.5rem}
.progreso-bar{height:100%;background:var(--accent);border-radius:2px;transition:width .3s}
</style>
</head>
<body>
<header>
  <div>
    <h1>KraftDo — Clasificador de Excel</h1>
    <span>Excel → JSON de configuración → Sistema Laravel+Filament</span>
  </div>
</header>
<div class="container">

<!-- Progreso -->
<div class="progreso"><div class="progreso-bar" id="barra" style="width:33%"></div></div>

<div class="steps-nav">
  <div class="step-dot active" id="dot1" title="Upload">1</div>
  <div style="flex:1;height:1px;background:var(--border);align-self:center"></div>
  <div class="step-dot" id="dot2" title="Clasificar">2</div>
  <div style="flex:1;height:1px;background:var(--border);align-self:center"></div>
  <div class="step-dot" id="dot_norm" title="Normalizar">N</div>
  <div style="flex:1;height:1px;background:var(--border);align-self:center"></div>
  <div class="step-dot" id="dot_consolida" title="Consolidar">C</div>
  <div style="flex:1;height:1px;background:var(--border);align-self:center"></div>
  <div class="step-dot" id="dot3" title="Resultado">3</div>
</div>

<!-- PASO 1: Upload -->
<div class="step active" id="step1">
  <div class="card">
    <h2>Paso 1 — Sube el Excel del negocio</h2>
    <div class="upload-zone" id="dropzone" onclick="document.getElementById('fileInput').click()">
      <input type="file" id="fileInput" accept=".xlsx,.xlsm,.xls">
      <div style="font-size:2rem;margin-bottom:.5rem">📊</div>
      <div style="font-weight:500;margin-bottom:.25rem">Arrastra tu Excel aquí o haz click</div>
      <div style="font-size:.85rem;color:var(--muted)">.xlsx, .xlsm — cualquier estructura</div>
    </div>
    <div id="upload-status" style="margin-top:1rem;display:none"></div>
  </div>
</div>

<!-- PASO 2: Empresa + Clasificación -->
<div class="step" id="step2">
  <div class="card">
    <h2>Paso 2 — Datos de la empresa</h2>
    <div class="grid-empresa">
      <div class="field"><label>Nombre empresa</label><input id="e_nombre" placeholder="KraftDo SpA"></div>
      <div class="field"><label>Archivo Excel</label><input id="e_archivo" placeholder="BD_Maestra.xlsx" readonly></div>
      <div class="field"><label>Email</label><input id="e_email" placeholder="hola@empresa.cl"></div>
      <div class="field"><label>Teléfono</label><input id="e_telefono" placeholder="+569 1234 5678"></div>
      <div class="field"><label>Web</label><input id="e_web" placeholder="www.empresa.cl"></div>
      <div class="field"><label>RUT</label><input id="e_rut" placeholder="12.345.678-9"></div>
    </div>
  </div>

  <div class="card">
    <h2>Clasificación de hojas</h2>
    <p style="font-size:.85rem;color:var(--muted);margin-bottom:1rem">
      Clasifica cada hoja. La app pre-clasifica automáticamente — solo corrige las que estén mal.
    </p>
    <div style="display:flex;gap:.5rem;font-size:.8rem;margin-bottom:1rem;flex-wrap:wrap">
      <span><span class="badge badge-green">Catálogo</span> Productos con precios</span>
      <span><span class="badge badge-blue">Registros</span> Clientes, pedidos, caja</span>
      <span><span class="badge badge-purple">KPIs</span> Dashboard / resumen</span>
      <span style="color:var(--muted)">Vista → no se importa | Ignorar → se omite</span>
    </div>
    <div id="hojas-container"></div>
  </div>

  <div class="card" style="background:#f0f9ff;border-color:#bae6fd">
    <h2 style="color:#0369a1">Fase 0 — Describe tu negocio (opcional)</h2>
    <p style="font-size:.85rem;color:var(--muted);margin-bottom:.75rem">
      Si describes cómo funciona tu negocio, Claude analizará el Excel y sugerirá
      qué hojas consolidar, relaciones entre tablas y problemas de normalización.
    </p>
    <div class="field">
      <label>Descripción del negocio</label>
      <textarea id="descripcion-negocio" rows="3"
        placeholder="Ej: Tenemos 3 hojas de productos (NFC, sublimación e impresión 3D) porque los separamos por tipo pero en realidad son el mismo catálogo. Los pedidos llegan por WhatsApp..."
        oninput="this.value.length > 30 && clearTimeout(window._iaTimer) && (window._iaTimer = setTimeout(() => obtenerSugerenciasIA(this.value), 1500))"></textarea>
    </div>
  </div>

  <div class="actions">
    <button class="btn btn-outline" onclick="irStep(1)">← Volver</button>
    <button class="btn btn-primary" onclick="irStep('norm')">Normalización →</button>
  </div>
</div>

<!-- PASO 3: Normalización de patrones raros -->
<div class="step" id="step_norm">
  <div class="card">
    <h2>Patrones detectados en el Excel</h2>
    <p style="font-size:.85rem;color:var(--muted);margin-bottom:1rem">
      El sistema analizó la estructura de cada hoja. Revisa y confirma los patrones raros antes de continuar.
    </p>
    <div id="patrones-container"></div>
  </div>
  <div class="actions">
    <button class="btn btn-outline" onclick="irStep(2)">← Volver</button>
    <button class="btn btn-primary" onclick="irStep('consolida')">Consolidar entidades →</button>
  </div>
</div>

<!-- PASO 4: Consolidación de entidades -->
<div class="step" id="step_consolida">
  <div class="card">
    <h2>Consolidación de entidades</h2>
    <p style="font-size:.85rem;color:var(--muted);margin-bottom:1rem">
      Si varias hojas representan la misma entidad (ej: 3 hojas de productos), agrúpalas aquí.
      Generará una sola tabla MySQL con campo <code>tipo</code> discriminador.
    </p>

    <div id="grupos-container">
      <div class="badge badge-gray" style="font-size:.8rem">Sin grupos definidos — las hojas se generarán como tablas separadas</div>
    </div>

    <div style="margin-top:1rem;display:flex;gap:.5rem;flex-wrap:wrap;align-items:center">
      <select id="sel-alias-1" style="padding:.4rem .6rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem"></select>
      <span style="color:var(--muted)">+</span>
      <select id="sel-alias-2" style="padding:.4rem .6rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem"></select>
      <input id="nombre-entidad" placeholder="nombre_tabla" style="padding:.4rem .6rem;border:1px solid var(--border);border-radius:6px;font-size:.85rem;width:130px">
      <button class="btn btn-outline" onclick="agregarGrupo()" style="padding:.4rem .8rem">Agrupar</button>
    </div>

    <div id="sugerencias-ia" style="margin-top:1rem;display:none">
      <div class="card" style="background:#fffbf0;border-color:var(--accent);margin-bottom:0">
        <h2 style="color:#b45309;margin-bottom:.5rem">💡 Sugerencias de Claude</h2>
        <div id="sugerencias-contenido"></div>
      </div>
    </div>
  </div>

  <div class="actions">
    <button class="btn btn-outline" onclick="irStep('norm')">← Volver</button>
    <button class="btn btn-primary" onclick="irStep(3)">Generar JSON →</button>
  </div>
</div>

<!-- PASO 3: Resultado -->
<div class="step" id="step3">
  <div class="card">
    <h2>Paso 3 — JSON generado</h2>
    <div id="resumen-final" style="margin-bottom:1rem"></div>
    <div id="resultado">
      <pre id="json-output"></pre>
      <div class="actions" style="margin-top:1rem">
        <button class="btn btn-accent" onclick="descargarJSON()">⬇ Descargar JSON</button>
        <button class="btn btn-outline" onclick="copiarJSON()">📋 Copiar</button>
        <button class="btn btn-outline" onclick="irStep(2)">← Ajustar</button>
      </div>
    </div>
  </div>

  <div class="card" style="background:#fffbf0;border-color:var(--accent)">
    <h2 style="color:#b45309">Próximos pasos</h2>
    <ol style="font-size:.85rem;line-height:2;padding-left:1.2rem;margin-top:.5rem">
      <li>Guarda el JSON descargado como <code>empresas/{nombre}.json</code></li>
      <li>Crea un proyecto Laravel vacío</li>
      <li>Ejecuta: <code>python3 generator.py {nombre} --output ./mi-sistema</code></li>
      <li>Copia los archivos generados al proyecto Laravel</li>
      <li>Ejecuta: <code>chmod +x install.sh && ./install.sh</code></li>
      <li>Importa los datos: <code>python3 importer.py {nombre}</code></li>
    </ol>
  </div>
</div>

</div>
<script>
let hojasData = {};
let jsonGenerado = null;
let archivoNombre = "";

// Upload
const fileInput = document.getElementById("fileInput");
const dropzone  = document.getElementById("dropzone");

dropzone.addEventListener("dragover", e => { e.preventDefault(); dropzone.style.borderColor="#F5A623" });
dropzone.addEventListener("dragleave", () => { dropzone.style.borderColor="" });
dropzone.addEventListener("drop", e => {
  e.preventDefault();
  dropzone.style.borderColor = "";
  const f = e.dataTransfer.files[0];
  if (f) procesarArchivo(f);
});
fileInput.addEventListener("change", () => {
  if (fileInput.files[0]) procesarArchivo(fileInput.files[0]);
});

function procesarArchivo(file) {
  archivoNombre = file.name;
  document.getElementById("e_archivo").value = file.name;
  const status = document.getElementById("upload-status");
  status.style.display = "block";
  status.innerHTML = '<span style="color:var(--muted)">⏳ Analizando hojas...</span>';

  const fd = new FormData();
  fd.append("file", file);

  fetch("/analizar", { method: "POST", body: fd })
    .then(r => r.json())
    .then(data => {
      hojasData = data;
      status.innerHTML = `<span style="color:#059669">✅ ${Object.keys(data).length} hojas detectadas</span>`;
      renderHojas(data);
      irStep(2);
    })
    .catch(e => {
      status.innerHTML = `<span style="color:#dc2626">❌ Error: ${e.message}</span>`;
    });
}

function renderHojas(hojas) {
  const cont = document.getElementById("hojas-container");
  cont.innerHTML = "";

  for (const [nombre, meta] of Object.entries(hojas)) {
    const alias = nombre.toLowerCase().replace(/[^a-z0-9]/g,"_");
    const clasif = meta.clasificacion || "registros";

    const card = document.createElement("div");
    card.className = "hoja-card";
    card.dataset.alias = alias;
    card.dataset.nombre = nombre;

    const tipos = [
      {id:"catalogo", label:"Catálogo"},
      {id:"registros", label:"Registros"},
      {id:"kpis", label:"KPIs"},
      {id:"vista", label:"Vista"},
      {id:"ignorar", label:"Ignorar"},
    ];

    const tiposBtns = tipos.map(t =>
      `<button class="tipo-btn ${t.id===clasif?'activo':''}" data-tipo="${t.id}"
        onclick="selTipo(this,'${alias}')">${t.label}</button>`
    ).join("");

    const previewCols = meta.columnas_raw ? Object.values(meta.columnas_raw).slice(0,6) : [];
    const preview = meta.preview || [];

    let previewHTML = "";
    if (previewCols.length) {
      const ths = previewCols.map(c => `<th>${c}</th>`).join("");
      const trs = preview.map(row =>
        "<tr>" + previewCols.map(c => `<td>${row[c]||""}</td>`).join("") + "</tr>"
      ).join("");
      previewHTML = `<div class="preview-wrap"><table class="preview-table"><thead><tr>${ths}</tr></thead><tbody>${trs}</tbody></table></div>`;
    }

    card.innerHTML = `
      <div class="hoja-header" onclick="toggleHoja('${alias}')">
        <div class="hoja-nombre">${nombre}</div>
        <div class="hoja-meta">${meta.filas||0} filas · ${previewCols.length} columnas</div>
        <div class="tipos" onclick="event.stopPropagation()">${tiposBtns}</div>
      </div>
      <div class="hoja-body" id="body-${alias}">
        ${previewHTML || '<p style="color:var(--muted);font-size:.85rem">Hoja sin datos detectables</p>'}
      </div>`;

    cont.appendChild(card);
  }
}

function selTipo(btn, alias) {
  const card = btn.closest(".hoja-card");
  card.querySelectorAll(".tipo-btn").forEach(b => b.classList.remove("activo"));
  btn.classList.add("activo");
}

function toggleHoja(alias) {
  const body = document.getElementById("body-"+alias);
  body.classList.toggle("open");
}

const STEPS = [1, 2, "norm", "consolida", 3];

function irStep(n) {
  document.querySelectorAll(".step").forEach(s => s.classList.remove("active"));
  const el = document.getElementById("step"+n) || document.getElementById("step_"+n);
  if (el) el.classList.add("active");

  // Actualizar dots
  const idx = STEPS.indexOf(n);
  STEPS.forEach((s, i) => {
    const dot = document.getElementById("dot"+s);
    if (!dot) return;
    dot.classList.remove("active","done");
    if (i < idx) dot.classList.add("done");
    if (i === idx) dot.classList.add("active");
  });
  document.getElementById("barra").style.width = ((idx+1)/STEPS.length*100)+"%";

  if (n === "norm") renderPatrones();
  if (n === "consolida") renderSelectsConsolidacion();
  if (n === 3) generarJSON();
}

// ── Normalización ──────────────────────────────────────────────────────────
let diagnosticosPatrones = {};

async function renderPatrones() {
  const cont = document.getElementById("patrones-container");
  if (!archivoNombre || Object.keys(hojasData).length === 0) {
    cont.innerHTML = "<p style='color:var(--muted);font-size:.85rem'>Sube un archivo primero.</p>";
    return;
  }

  // Normalización: usar filename del archivo ya subido (evita doble upload)
  if (archivoNombre) {
    try {
      const r = await fetch("/normalizar?filename=" + encodeURIComponent(archivoNombre), {method:"POST"});
      if (r.ok) {
        const diags = await r.json();
        // Enriquecer hojasData con el diagnóstico de patrones
        for (const [nombre, diag] of Object.entries(diags)) {
          if (hojasData[nombre]) {
            hojasData[nombre]._patron = diag.patron;
            hojasData[nombre]._requiere_humano = diag.requiere_humano;
            hojasData[nombre]._descripcion = diag.descripcion;
          }
        }
      }
    } catch(e) { /* silencioso — normalización es opcional */ }
  }
  const patron_icons = {
    vertical: "✅", horizontal: "↔️", multi_header: "📊",
    formulario: "📝", multi_tabla: "📋", con_totales: "⚠️",
    sparse: "🕳️", vacia: "⬜"
  };
  const patron_labels = {
    vertical: "Tabla estándar", horizontal: "Tabla horizontal/pivot",
    multi_header: "Multi-header (2 filas)", formulario: "Formulario campo:valor",
    multi_tabla: "Múltiples tablas", con_totales: "Incluye totales mezclados",
    sparse: "Datos dispersos", vacia: "Hoja vacía"
  };

  // Usar hojasData para análisis básico (ya tenemos la info)
  cont.innerHTML = "";
  for (const [nombre, meta] of Object.entries(hojasData)) {
    if (meta.vacia) continue;
    const plantilla = meta.plantilla || "heuristica";
    const icon  = plantilla === "heuristica" ? "✅" : "📊";
    const label = `${nombre} — detectado por ${plantilla}`;
    const div = document.createElement("div");
    div.style.cssText = "padding:.6rem .75rem;border:1px solid var(--border);border-radius:8px;margin-bottom:.5rem;display:flex;gap:.75rem;align-items:flex-start";
    div.innerHTML = `
      <span style="font-size:1.2rem">${icon}</span>
      <div style="flex:1">
        <div style="font-weight:500;font-size:.9rem">${nombre}</div>
        <div style="font-size:.8rem;color:var(--muted)">${meta.filas} filas · ${Object.keys(meta.columnas||{}).length} columnas · plantilla: ${plantilla}</div>
      </div>
      <span class="badge ${meta.plantilla && meta.plantilla !== 'heuristica' ? 'badge-green' : 'badge-gray'}" style="font-size:.75rem;align-self:center">
        ${meta.plantilla !== 'heuristica' ? '✓ Plantilla' : 'Heurístico'}
      </span>`;
    cont.appendChild(div);
  }
}

// ── Consolidación ──────────────────────────────────────────────────────────
let gruposConsolidacion = [];

function renderSelectsConsolidacion() {
  const aliases = Object.keys(hojasData).filter(k => !hojasData[k].vacia);
  const makeOpts = () => aliases.map(a => `<option value="${a}">${a}</option>`).join("");
  document.getElementById("sel-alias-1").innerHTML = makeOpts();
  document.getElementById("sel-alias-2").innerHTML = makeOpts();
  if (aliases.length > 1) {
    document.getElementById("sel-alias-2").selectedIndex = 1;
  }
  renderGrupos();
}

function agregarGrupo() {
  const a1   = document.getElementById("sel-alias-1").value;
  const a2   = document.getElementById("sel-alias-2").value;
  const nom  = document.getElementById("nombre-entidad").value.trim() || a1;
  if (a1 === a2) { alert("Selecciona dos hojas distintas"); return; }
  if (!nom)      { alert("Ingresa un nombre para la entidad"); return; }
  // Evitar duplicados
  const yaExiste = gruposConsolidacion.find(g => g.nombre === nom);
  if (yaExiste) {
    if (!yaExiste.aliases.includes(a2)) yaExiste.aliases.push(a2);
  } else {
    gruposConsolidacion.push({ nombre: nom, aliases: [a1, a2] });
  }
  renderGrupos();
  document.getElementById("nombre-entidad").value = "";
}

function renderGrupos() {
  const cont = document.getElementById("grupos-container");
  if (gruposConsolidacion.length === 0) {
    cont.innerHTML = '<div class="badge badge-gray" style="font-size:.8rem">Sin grupos — tablas separadas</div>';
    return;
  }
  cont.innerHTML = gruposConsolidacion.map((g, i) => `
    <div style="padding:.6rem .75rem;border:1px solid var(--border);border-radius:8px;margin-bottom:.5rem;display:flex;align-items:center;gap:.75rem">
      <span style="font-size:.9rem">🔗</span>
      <div style="flex:1">
        <strong>${g.nombre}</strong>
        <span style="font-size:.8rem;color:var(--muted);margin-left:.5rem">← ${g.aliases.join(" + ")}</span>
      </div>
      <button onclick="eliminarGrupo(${i})" style="background:none;border:none;cursor:pointer;color:#dc2626;font-size:.85rem">✕</button>
    </div>`).join("");
}

function eliminarGrupo(idx) {
  gruposConsolidacion.splice(idx, 1);
  renderGrupos();
}

// Sugerencias de IA
async function obtenerSugerenciasIA(descripcion) {
  if (!descripcion || descripcion.length < 20) return;
  const cont = document.getElementById("sugerencias-ia");
  const body = document.getElementById("sugerencias-contenido");
  cont.style.display = "block";
  body.innerHTML = "<p style='color:var(--muted);font-size:.85rem'>⏳ Analizando con Claude...</p>";
  try {
    const r = await fetch("/fase0/analizar", {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({
        empresa: document.getElementById("e_nombre").value || "empresa",
        descripcion,
        hojas_meta: hojasData
      })
    });
    const d = await r.json();
    let html = "";
    if (d.resumen) html += `<p style="font-size:.85rem;margin-bottom:.75rem">${d.resumen}</p>`;
    if (d.consolidaciones?.length) {
      html += `<p style="font-size:.8rem;font-weight:600;margin-bottom:.4rem">Consolidaciones sugeridas:</p>`;
      d.consolidaciones.forEach(c => {
        html += `<div style="font-size:.8rem;margin-bottom:.4rem">
          <strong>${c.nombre_entidad}</strong>: ${c.aliases.join(" + ")}
          <span style="color:var(--muted)"> — ${c.razon}</span>
          <button onclick="aplicarSugerencia('${c.nombre_entidad}',${JSON.stringify(c.aliases)})"
            style="margin-left:.5rem;font-size:.75rem;padding:.2rem .5rem;background:var(--accent);border:none;border-radius:4px;cursor:pointer;color:white">
            Aplicar
          </button>
        </div>`;
      });
    }
    if (d.advertencias?.length) {
      html += `<p style="font-size:.8rem;font-weight:600;margin-top:.75rem;margin-bottom:.4rem">Advertencias:</p>`;
      d.advertencias.forEach(a => {
        html += `<div style="font-size:.8rem;color:#b45309;margin-bottom:.3rem">⚠️ ${a.hoja}: ${a.descripcion}</div>`;
      });
    }
    body.innerHTML = html || "<p style='font-size:.85rem;color:var(--muted)'>Sin sugerencias específicas</p>";
  } catch(e) {
    body.innerHTML = `<p style="font-size:.85rem;color:var(--muted)">Error al consultar la API: ${e.message}</p>`;
  }
}

function aplicarSugerencia(nombre, aliases) {
  const yaExiste = gruposConsolidacion.find(g => g.nombre === nombre);
  if (!yaExiste) {
    gruposConsolidacion.push({ nombre, aliases });
    renderGrupos();
  }
}

function recopilarClasificaciones() {
  const result = {};
  document.querySelectorAll(".hoja-card").forEach(card => {
    const alias  = card.dataset.alias;
    const nombre = card.dataset.nombre;
    const activo = card.querySelector(".tipo-btn.activo");
    result[nombre] = { alias, tipo: activo ? activo.dataset.tipo : "ignorar" };
  });
  return result;
}

function generarJSON() {
  const empresa = {
    nombre:        document.getElementById("e_nombre").value || "Mi Empresa",
    email:         document.getElementById("e_email").value,
    telefono:      document.getElementById("e_telefono").value,
    web:           document.getElementById("e_web").value,
    rut:           document.getElementById("e_rut").value,
    archivo:       archivoNombre,
    color_primary: "1E3A5F",
    color_accent:  "F5A623",
  };

  const clasifs = recopilarClasificaciones();
  const hojasClasif = {};
  for (const [nombre, data] of Object.entries(clasifs)) {
    hojasClasif[nombre] = data.tipo;
  }

  fetch("/generar", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({
      empresa,
      clasificaciones: hojasClasif,
      hojas_meta: hojasData,
      grupos_consolidacion: gruposConsolidacion
    })
  })
  .then(r => r.json())
  .then(data => {
    jsonGenerado = data;
    document.getElementById("json-output").textContent = JSON.stringify(data, null, 2);
    document.getElementById("resultado").style.display = "block";

    const hojasCont = Object.values(hojasClasif).filter(t => !["ignorar","vista"].includes(t)).length;
    document.getElementById("resumen-final").innerHTML = `
      <div style="display:flex;gap:.5rem;flex-wrap:wrap">
        <span class="badge badge-green">${hojasCont} hojas configuradas</span>
        <span class="badge badge-blue">${Object.keys(data.hojas||{}).length} en el JSON</span>
        <span style="font-size:.85rem;color:var(--muted);align-self:center">
          Empresa: <strong>${empresa.nombre}</strong>
        </span>
      </div>`;
  });
}

function descargarJSON() {
  if (!jsonGenerado) return;
  const nombre = document.getElementById("e_nombre").value || "empresa";
  const alias  = nombre.toLowerCase().replace(/[^a-z0-9]/g,"_");
  const blob   = new Blob([JSON.stringify(jsonGenerado, null, 2)], {type:"application/json"});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = `${alias}.json`;
  a.click();
}

function copiarJSON() {
  navigator.clipboard.writeText(JSON.stringify(jsonGenerado, null, 2));
  alert("JSON copiado al portapapeles");
}
</script>
</body>
</html>
"""

# ── Formularios dinámicos ────────────────────────────────────────────────────
FORM_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>KraftDo — {titulo}</title>
<style>
:root{{--primary:#1E3A5F;--accent:#F5A623;--bg:#f8f7f4;--card:#fff;--border:#e5e4dd;--text:#1c1c1a;--muted:#6b6b67}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,sans-serif;background:var(--bg);color:var(--text);padding:1rem}}
header{{background:var(--primary);color:#fff;padding:.75rem 1.5rem;border-radius:8px;margin-bottom:1.5rem;display:flex;justify-content:space-between;align-items:center}}
header h1{{font-size:1rem;font-weight:500}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:1.5rem;margin-bottom:1rem}}
.field{{margin-bottom:1rem}}
label{{display:block;font-size:.8rem;font-weight:600;color:var(--muted);margin-bottom:.3rem;text-transform:uppercase;letter-spacing:.04em}}
input,select,textarea{{width:100%;padding:.6rem .75rem;border:1px solid var(--border);border-radius:6px;font-size:.9rem;background:var(--bg)}}
input:focus,select:focus,textarea:focus{{outline:none;border-color:var(--primary)}}
textarea{{resize:vertical;min-height:80px}}
.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:1rem}}
@media(max-width:560px){{.grid-2{{grid-template-columns:1fr}}}}
.btn{{padding:.6rem 1.2rem;border-radius:8px;border:none;cursor:pointer;font-size:.9rem;font-weight:500}}
.btn-primary{{background:var(--primary);color:#fff}}
.btn-outline{{background:transparent;border:1px solid var(--border);color:var(--text);margin-right:.5rem}}
.actions{{display:flex;gap:.5rem;margin-top:1rem}}
.alert{{padding:.75rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem}}
.alert-ok{{background:#d1fae5;color:#065f46}}
.alert-err{{background:#fee2e2;color:#991b1b}}
.required{{color:#dc2626;margin-left:2px}}
.field-hint{{font-size:.75rem;color:var(--muted);margin-top:.2rem}}
#registros-table{{width:100%;border-collapse:collapse;font-size:.85rem;margin-top:1rem}}
#registros-table th{{background:var(--bg);padding:.5rem;text-align:left;border-bottom:1px solid var(--border);font-weight:600;color:var(--muted)}}
#registros-table td{{padding:.5rem;border-bottom:1px solid #f0efe8}}
#registros-table tr:hover{{background:#fffbf0}}
.btn-edit{{font-size:.75rem;padding:.25rem .6rem;background:var(--primary);color:#fff;border:none;border-radius:4px;cursor:pointer}}
.btn-del{{font-size:.75rem;padding:.25rem .6rem;background:#dc2626;color:#fff;border:none;border-radius:4px;cursor:pointer;margin-left:.25rem}}
.badge{{font-size:.75rem;padding:.2rem .6rem;border-radius:999px;font-weight:500}}
.b-activo{{background:#d1fae5;color:#065f46}}
.b-inactivo{{background:#f3f4f6;color:#6b7280}}
</style>
</head>
<body>
<header>
  <h1>🗂️ {titulo}</h1>
  <span style="font-size:.8rem;opacity:.7">{empresa} · {alias}</span>
</header>
<div id="msg"></div>

<div class="card">
  <h2 style="font-size:.85rem;font-weight:600;color:var(--muted);margin-bottom:1rem;text-transform:uppercase" id="form-titulo">Nuevo registro</h2>
  <form id="form" onsubmit="guardar(event)">
    {campos_html}
    <div class="actions">
      <button type="button" class="btn btn-outline" onclick="cancelar()">Cancelar</button>
      <button type="submit" class="btn btn-primary" id="btn-guardar">Guardar</button>
    </div>
  </form>
</div>

<div class="card">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.75rem">
    <h2 style="font-size:.85rem;font-weight:600;color:var(--muted);text-transform:uppercase">Registros</h2>
    <input type="text" placeholder="Buscar..." id="buscar-input" oninput="buscarLocal(this.value)"
           style="width:200px;font-size:.8rem;padding:.3rem .6rem">
  </div>
  <div style="overflow-x:auto">
    <table id="registros-table">
      <thead><tr id="thead-row"></tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
  <div id="paginacion" style="margin-top:.5rem;font-size:.8rem;color:var(--muted)"></div>
</div>

<script>
const API   = "{api_url}";
const EMP   = "{empresa}";
const ALIAS = "{alias}";
const IDENT = "{ident}";
const COLS  = {cols_json};

let todosRegistros = [];
let editandoId = null;

async function cargar() {{
  try {{
    const r = await fetch(`${{API}}/${{EMP}}/registros/${{ALIAS}}`);
    const d = await r.json();
    todosRegistros = d.registros || [];
    renderTabla(todosRegistros);
  }} catch(e) {{ msg("Error cargando registros: " + e, "err"); }}
}}

function renderTabla(registros) {{
  const thead = document.getElementById("thead-row");
  const tbody = document.getElementById("tbody");
  const cols  = Object.keys(COLS).slice(0,8);

  if (!thead.innerHTML) {{
    thead.innerHTML = cols.map(c => `<th>${{c}}</th>`).join("") + "<th>Acciones</th>";
  }}

  tbody.innerHTML = registros.map(r => {{
    const celdas = cols.map(c => {{
      let v = r[c] ?? "";
      if (c === "estado") {{
        const cls = v.toLowerCase() === "activo" ? "b-activo" : "b-inactivo";
        return `<td><span class="badge ${{cls}}">${{v}}</span></td>`;
      }}
      if (typeof v === "number" && v > 1000) v = "$" + v.toLocaleString("es-CL");
      return `<td>${{v}}</td>`;
    }}).join("");
    const id = r[IDENT] || "";
    return `<tr>${{celdas}}<td>
      <button class="btn-edit" onclick='editar(${{JSON.stringify(r)}})'>✏️</button>
      <button class="btn-del"  onclick="eliminar('${{id}}')">🗑️</button>
    </td></tr>`;
  }}).join("");

  document.getElementById("paginacion").textContent = `${{registros.length}} registros`;
}}

function buscarLocal(texto) {{
  if (!texto) {{ renderTabla(todosRegistros); return; }}
  const t = texto.toLowerCase();
  const filtrados = todosRegistros.filter(r =>
    Object.values(r).some(v => String(v||"").toLowerCase().includes(t))
  );
  renderTabla(filtrados);
}}

async function guardar(e) {{
  e.preventDefault();
  const form = document.getElementById("form");
  const datos = {{}};
  new FormData(form).forEach((v,k) => {{ if(v) datos[k] = v; }});

  const url    = editandoId
    ? `${{API}}/${{EMP}}/registros/${{ALIAS}}/${{editandoId}}`
    : `${{API}}/${{EMP}}/registros/${{ALIAS}}`;
  const method = editandoId ? "PUT" : "POST";

  try {{
    const r = await fetch(url, {{
      method,
      headers: {{"Content-Type":"application/json","X-API-Key": localStorage.getItem("kraftdo_key")||""}},
      body: JSON.stringify(datos)
    }});
    const d = await r.json();
    if (!r.ok) throw new Error(d.detail || JSON.stringify(d));
    msg(editandoId ? "✅ Registro actualizado" : "✅ Registro creado", "ok");
    form.reset(); editandoId = null;
    document.getElementById("form-titulo").textContent = "Nuevo registro";
    document.getElementById("btn-guardar").textContent = "Guardar";
    await cargar();
  }} catch(err) {{ msg("❌ " + err.message, "err"); }}
}}

function editar(registro) {{
  editandoId = registro[IDENT];
  document.getElementById("form-titulo").textContent = "Editando: " + editandoId;
  document.getElementById("btn-guardar").textContent = "Actualizar";
  const form = document.getElementById("form");
  Object.entries(registro).forEach(([k,v]) => {{
    const el = form.elements[k];
    if (el) el.value = v || "";
  }});
  form.scrollIntoView({{behavior:"smooth"}});
}}

async function eliminar(id) {{
  if (!confirm(`¿Eliminar "${{id}}"? Esta acción no se puede deshacer.`)) return;
  try {{
    const r = await fetch(`${{API}}/${{EMP}}/registros/${{ALIAS}}/${{id}}`, {{
      method: "DELETE",
      headers: {{"X-API-Key": localStorage.getItem("kraftdo_key")||""}}
    }});
    const d = await r.json();
    if (!r.ok) throw new Error(d.detail);
    msg("✅ Eliminado", "ok");
    await cargar();
  }} catch(err) {{ msg("❌ " + err.message, "err"); }}
}}

function cancelar() {{
  document.getElementById("form").reset();
  editandoId = null;
  document.getElementById("form-titulo").textContent = "Nuevo registro";
  document.getElementById("btn-guardar").textContent = "Guardar";
}}

function msg(texto, tipo) {{
  const el = document.getElementById("msg");
  el.innerHTML = `<div class="alert alert-${{tipo}}">${{texto}}</div>`;
  setTimeout(() => {{ el.innerHTML = ""; }}, 4000);
}}

cargar();
</script>
</body>
</html>
"""

def _gen_campos_html(schema: dict) -> str:
    """Genera el HTML de campos del formulario desde el schema."""
    campos = schema.get("campos", {})
    ident  = schema.get("identificador")
    html   = []
    pares  = []  # campos que van en grid de 2

    for campo, info in campos.items():
        if campo == ident:
            # Campo ID: oculto al crear, visible al editar
            html.append(f'<input type="hidden" name="{campo}" id="campo_{campo}">')
            continue

        label = campo.replace("_", " ").capitalize()
        req   = "required" if info["requerido"] else ""
        star  = '<span class="required">*</span>' if info["requerido"] else ""

        tipo = info["tipo"]
        if tipo == "select":
            opts = "".join(f'<option value="{o}">{o}</option>' for o in info["opciones"])
            inp  = f'<select name="{campo}" {req}><option value="">— seleccionar —</option>{opts}</select>'
        elif tipo == "textarea":
            inp  = f'<textarea name="{campo}" rows="3" {req}></textarea>'
        elif tipo == "checkbox":
            inp  = f'<input type="checkbox" name="{campo}" value="1" style="width:auto">'
        elif tipo == "date":
            inp  = f'<input type="date" name="{campo}" {req}>'
        elif tipo == "number":
            inp  = f'<input type="number" name="{campo}" step="any" {req}>'
        elif tipo == "email":
            inp  = f'<input type="email" name="{campo}" {req}>'
        elif tipo == "tel":
            inp  = f'<input type="tel" name="{campo}" placeholder="+56 9 1234 5678" {req}>'
        else:
            inp  = f'<input type="text" name="{campo}" {req}>'

        bloque = f'<div class="field"><label>{label}{star}</label>{inp}</div>'
        pares.append(bloque)

    # Agrupar de a 2 en grid
    for i in range(0, len(pares), 2):
        if i + 1 < len(pares):
            html.append(f'<div class="grid-2">{pares[i]}{pares[i+1]}</div>')
        else:
            html.append(pares[i])

    return "\n".join(html)


# ── Historial ────────────────────────────────────────────────────────────────
HISTORIAL_DIR = os.path.join(SCRIPT_DIR, "historial")

def guardar_en_historial(empresa: str, cfg: dict):
    os.makedirs(HISTORIAL_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(HISTORIAL_DIR, f"{empresa}_{ts}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)
    return path

# ── Rutas ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def index():
    return HTML

def _check_classifier_key(request_key: str = None):
    """Verifica API key del classifier si está configurada."""
    key = os.environ.get("CLASSIFIER_KEY", "")
    if key and request_key != key:
        raise HTTPException(401, "API key requerida. Configura CLASSIFIER_KEY en .env")

@app.post("/analizar")
async def analizar(
    file: UploadFile = File(...),
    x_api_key: str = None
):
    _check_classifier_key(x_api_key)
    if not file.filename.endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(400, "Solo archivos Excel (.xlsx, .xlsm)")
    path = os.path.join(UPLOAD_DIR, file.filename)
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    hojas = analizar_excel(path)
    return JSONResponse(hojas)

@app.post("/generar")
def generar(data: GenerarRequest):
    resultado = generar_json(data.empresa, data.clasificaciones, data.hojas_meta)

    # Aplicar consolidaciones si hay grupos definidos
    if data.grupos_consolidacion:
        try:
            from consolidator import Consolidator
            import re as _re

            def _normalizar_alias(s: str) -> str:
                """Convierte nombre visible a slug ASCII válido (quita tildes y emojis)."""
                import unicodedata as _ud
                s = _ud.normalize("NFKD", str(s))
                s = s.encode("ascii", "ignore").decode("ascii")
                s = _re.sub(r"[^a-z0-9_]", "_", s.lower().strip())
                return _re.sub(r"_+", "_", s).strip("_") or "hoja"

            # Mapear nombre visible → clave real en el JSON
            # ej: "Productos NFC" → "__productos_nfc"
            hojas_json = resultado.get("hojas", {})
            mapa_alias = {}
            for clave in hojas_json:
                # Intentar match por slug del nombre de la hoja
                nombre_hoja = hojas_json[clave].get("nombre", "")
                slug_nombre = _normalizar_alias(nombre_hoja)
                mapa_alias[slug_nombre] = clave
                # También mapear por la clave directa (por si coincide)
                mapa_alias[clave] = clave

            c = Consolidator(resultado)
            for grupo in data.grupos_consolidacion:
                nombre = grupo.get("nombre")
                aliases_raw = grupo.get("aliases", [])
                # Resolver cada alias a la clave real del JSON
                aliases_reales = []
                for a in aliases_raw:
                    slug = _normalizar_alias(a)
                    # Buscar coincidencia exacta primero, luego parcial
                    clave_real = (
                        mapa_alias.get(a) or
                        mapa_alias.get(slug) or
                        next((k for k in hojas_json if slug in k or k in slug), None)
                    )
                    if clave_real:
                        aliases_reales.append(clave_real)

                if nombre and len(aliases_reales) >= 2:
                    try:
                        c.agregar_grupo(nombre, aliases_reales)
                    except Exception as e:
                        resultado.setdefault("_consolidacion_warnings", [])
                        resultado["_consolidacion_warnings"].append(str(e))

            if c.grupos:
                resultado = c.generar_json_consolidado()
        except Exception as e:
            resultado["_consolidacion_error"] = str(e)

    # Guardar en historial automáticamente
    nombre = data.empresa.get("nombre", "empresa").lower().replace(" ", "_")
    guardar_en_historial(nombre, resultado)
    return JSONResponse(resultado)

@app.get("/form/{empresa}/{alias}", response_class=HTMLResponse)
async def formulario_dinamico(empresa: str, alias: str, api_url: str = "http://localhost:8000"):
    """
    Formulario dinámico CRUD para cualquier hoja de cualquier empresa.
    Lee el schema del JSON y genera el formulario automáticamente.
    Compatible con Excel local y Google Sheets.
    """
    import sys
    sys.path.insert(0, SCRIPT_DIR)
    from core import Sistema

    try:
        s      = Sistema(empresa)
        schema = s.schema(alias)
    except Exception as e:
        raise HTTPException(400, f"Error cargando schema: {e}")

    campos_html = _gen_campos_html(schema)
    cols_json   = json.dumps({k: v["tipo"] for k, v in schema["campos"].items()})
    ident       = schema.get("identificador", "")
    titulo      = schema.get("nombre", alias).replace("📦","").replace("📋","").replace("👥","").strip()

    return FORM_HTML.format(
        titulo      = titulo,
        empresa     = empresa,
        alias       = alias,
        ident       = ident,
        api_url     = api_url,
        campos_html = campos_html,
        cols_json   = cols_json,
    )


@app.get("/forms/{empresa}", response_class=HTMLResponse)
async def indice_formularios(empresa: str, api_url: str = "http://localhost:8000"):
    """Índice con links a todos los formularios de una empresa."""
    import sys
    sys.path.insert(0, SCRIPT_DIR)
    from core import Sistema

    try:
        s     = Sistema(empresa)
        hojas = s.hojas_disponibles()
    except Exception as e:
        raise HTTPException(400, str(e))

    links = "".join(
        f'<li style="margin:.5rem 0"><a href="/form/{empresa}/{alias}?api_url={api_url}" '
        f'style="color:var(--primary)">'
        f'{info["nombre"]} <span style="font-size:.75rem;color:#9ca3af">({info["tipo"]})</span>'
        f'</a></li>'
        for alias, info in hojas.items()
        if info["tipo"] in ("catalogo", "registros")
    )

    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>KraftDo Forms — {empresa}</title>
<style>:root{{--primary:#1E3A5F}}body{{font-family:system-ui;padding:2rem;max-width:600px;margin:0 auto}}
h1{{color:var(--primary);margin-bottom:1rem}}ul{{list-style:none;padding:0}}
a{{text-decoration:none}}a:hover{{text-decoration:underline}}</style></head>
<body>
<h1>📋 Formularios — {empresa}</h1>
<ul>{links}</ul>
</body></html>"""



# ── Fase 0: Claude API para sugerencias de negocio ───────────────────────────
@app.post("/fase0/analizar")
async def fase0_analizar(datos: DescripcionNegocio):
    """
    Fase 0 — Inteligencia de negocio con Claude API.

    El usuario describe en lenguaje natural cómo funciona su negocio.
    Claude analiza la descripción + las hojas detectadas y sugiere:
      - Qué hojas consolidar en una misma entidad
      - Qué hojas son tablas horizontales
      - Relaciones entre entidades
      - Posibles errores de modelado
    """
    import urllib.request

    descripcion  = datos.descripcion
    hojas_meta   = datos.hojas_meta
    empresa_name = datos.empresa

    # Resumen de las hojas para el contexto
    resumen_hojas = ""
    for nombre, meta in hojas_meta.items():
        if not meta.get("vacia"):
            cols  = list(meta.get("columnas_raw", {}).values())[:6]
            filas = meta.get("filas", 0)
            cols_str = ", ".join(str(c) for c in cols)
            resumen_hojas += f"- {nombre}: {filas} filas, columnas: {cols_str}\n"

    prompt = f"""Eres un experto en normalización de bases de datos y sistemas para PYMEs latinoamericanas.

Un cliente tiene un Excel con este negocio:
"{descripcion}"

Las hojas detectadas en su Excel son:
{resumen_hojas}

Analiza este contexto y responde SOLO con un JSON con esta estructura exacta:
{{
  "resumen": "2-3 oraciones describiendo el negocio y sus flujos principales",
  "consolidaciones": [
    {{
      "nombre_entidad": "nombre_tabla_en_bd",
      "aliases": ["hoja1", "hoja2"],
      "razon": "por qué estas hojas son la misma entidad",
      "discriminador": "campo_tipo"
    }}
  ],
  "advertencias": [
    {{
      "hoja": "nombre_hoja",
      "tipo": "horizontal | multi_tabla | formulario | sparse",
      "descripcion": "qué tiene de raro y cómo resolverlo"
    }}
  ],
  "relaciones": [
    {{
      "desde": "tabla.campo",
      "hacia": "tabla.campo",
      "tipo": "belongsTo | hasMany"
    }}
  ],
  "recomendaciones": ["consejo 1", "consejo 2"]
}}

Responde SOLO con el JSON, sin texto adicional, sin markdown."""

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        # Sin API key — retornar análisis básico local
        return JSONResponse(_analisis_local(hojas_meta))

    try:
        # Usar Instructor si está disponible — garantiza respuesta tipada
        try:
            import instructor
            from pydantic import BaseModel, Field
            from anthropic import Anthropic

            class ConsolidacionSugerida(BaseModel):
                nombre_entidad: str
                aliases:        list[str]
                razon:          str
                discriminador:  str = "tipo"

            class AdvertenciaHoja(BaseModel):
                hoja:        str
                tipo:        str
                descripcion: str

            class RelacionSugerida(BaseModel):
                desde: str
                hacia: str
                tipo:  str = "belongsTo"

            class AnalisisNegocio(BaseModel):
                resumen:         str = Field(description="2-3 oraciones del negocio")
                consolidaciones: list[ConsolidacionSugerida] = []
                advertencias:    list[AdvertenciaHoja] = []
                relaciones:      list[RelacionSugerida] = []
                recomendaciones: list[str] = []

            client = instructor.from_anthropic(Anthropic(api_key=api_key))
            resultado_typed = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1500,
                response_model=AnalisisNegocio,
                messages=[{"role": "user", "content": prompt}]
            )
            return JSONResponse(resultado_typed.model_dump())

        except ImportError:
            # Fallback sin Instructor: llamada directa a la API
            import urllib.request
            payload = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 1500,
                "messages": [{"role": "user", "content": prompt}]
            }).encode()
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "Content-Type":      "application/json",
                    "x-api-key":         api_key,
                    "anthropic-version": "2023-06-01",
                },
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
                texto = data["content"][0]["text"].strip()
                texto = texto.replace("```json", "").replace("```", "").strip()
                return JSONResponse(json.loads(texto))

    except Exception as e:
        resultado = _analisis_local(hojas_meta)
        resultado["_error_api"] = str(e)
        return JSONResponse(resultado)


def _analisis_local(hojas_meta: dict) -> dict:
    """
    Análisis básico local cuando no hay API key de Anthropic.
    Detecta patrones comunes sin IA.
    """
    from normalizer import detectar_patron
    import openpyxl

    consolidaciones = []
    advertencias    = []

    # Detectar grupos de hojas con nombres similares (posibles consolidaciones)
    grupos_detectados = _detectar_grupos_por_nombre(list(hojas_meta.keys()))
    for nombre, aliases in grupos_detectados.items():
        if len(aliases) >= 2:
            consolidaciones.append({
                "nombre_entidad": nombre,
                "aliases":        aliases,
                "razon":          f"Las hojas {aliases} parecen ser variantes de la misma entidad",
                "discriminador":  "tipo",
            })

    return {
        "resumen":          "Análisis automático local (sin API key de Anthropic).",
        "consolidaciones":  consolidaciones,
        "advertencias":     advertencias,
        "relaciones":       [],
        "recomendaciones":  [
            "Configura ANTHROPIC_API_KEY para análisis con IA",
            "Revisa manualmente qué hojas son la misma entidad",
        ],
        "_local": True,
    }


def _detectar_grupos_por_nombre(nombres: list) -> dict:
    """Detecta grupos de hojas con nombres similares."""
    import difflib

    grupos = {}
    usados = set()

    for i, n1 in enumerate(nombres):
        if n1 in usados:
            continue
        similares = [n1]
        for n2 in nombres[i+1:]:
            if n2 in usados:
                continue
            ratio = difflib.SequenceMatcher(None,
                re.sub(r'[^\w]', '', n1.lower()),
                re.sub(r'[^\w]', '', n2.lower())
            ).ratio()
            if ratio > 0.5:
                similares.append(n2)

        if len(similares) >= 2:
            alias_limpio = re.sub(r'[^a-z0-9]', '_',
                min(similares, key=len).lower().strip())
            grupos[alias_limpio] = similares
            usados.update(similares)

    return grupos


# ── Normalización de patrones ─────────────────────────────────────────────────
@app.post("/normalizar")
async def normalizar_excel(
    filename: Optional[str] = None,
    file: Optional[UploadFile] = File(None)
):
    """
    Analiza un Excel y retorna el diagnóstico de patrones de cada hoja.
    Fase 2 del flujo.

    Acepta:
      - filename: nombre de un archivo ya subido con /analizar (evita doble upload)
      - file: archivo nuevo a analizar directamente
    """
    if filename:
        # Reusar archivo ya subido — evita doble upload
        path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(path):
            raise HTTPException(404, f"Archivo '{filename}' no encontrado. Sube el archivo primero con /analizar")
    elif file:
        path = os.path.join(UPLOAD_DIR, file.filename)
        content_bytes = await file.read()
        with open(path, "wb") as f:
            f.write(content_bytes)
    else:
        raise HTTPException(400, "Proporciona 'filename' (archivo ya subido) o 'file' (archivo nuevo)")

    from normalizer import analizar_excel_completo
    try:
        diagnosticos = analizar_excel_completo(path)
        resultado = {}
        for hoja, diag in diagnosticos.items():
            resultado[hoja] = {
                k: v for k, v in diag.items()
                if isinstance(v, (str, int, float, bool, list, dict, type(None)))
            }
        return JSONResponse(resultado)
    except Exception as e:
        raise HTTPException(400, f"Error analizando Excel: {e}")


# ── Consolidación ─────────────────────────────────────────────────────────────
@app.post("/consolidar/analizar")
async def analizar_consolidacion(datos: GrupoConsolidacion):
    """
    Analiza si un grupo de hojas es consolidable y retorna el reporte.
    """
    from consolidator import Consolidator
    cfg = datos.empresa_cfg or {}
    if not cfg.get("hojas"):
        raise HTTPException(400, "empresa_cfg.hojas es requerido")

    c = Consolidator(cfg)
    try:
        c.agregar_grupo(datos.nombre_entidad, datos.aliases)
        reporte = c.analizar_grupo(datos.nombre_entidad)
        return JSONResponse(reporte)
    except Exception as e:
        raise HTTPException(400, str(e))


@app.post("/consolidar/generar")
async def generar_consolidado(datos: GrupoConsolidacion):
    """
    Genera el JSON consolidado para un grupo de hojas.
    Retorna el JSON modificado con la entidad consolidada.
    """
    from consolidator import Consolidator
    cfg = datos.empresa_cfg or {}
    if not cfg.get("hojas"):
        raise HTTPException(400, "empresa_cfg.hojas es requerido")

    c = Consolidator(cfg)
    try:
        c.agregar_grupo(datos.nombre_entidad, datos.aliases)
        cfg_nuevo = c.generar_json_consolidado()
        return JSONResponse(cfg_nuevo)
    except Exception as e:
        raise HTTPException(400, str(e))


@app.get("/historial/{empresa}")
def historial(empresa: str):
    """Lista versiones anteriores del JSON de una empresa."""
    if not os.path.exists(HISTORIAL_DIR):
        return JSONResponse([])
    archivos = sorted([
        f for f in os.listdir(HISTORIAL_DIR)
        if f.startswith(empresa.lower()) and f.endswith(".json")
    ], reverse=True)
    return JSONResponse([{
        "archivo": f,
        "fecha":   f.replace(empresa.lower()+"_","").replace(".json",""),
        "url":     f"/historial/{empresa}/{f}"
    } for f in archivos[:10]])

@app.get("/historial/{empresa}/{archivo}")
def historial_archivo(empresa: str, archivo: str):
    """Descarga una versión anterior del JSON."""
    path = os.path.join(HISTORIAL_DIR, archivo)
    if not os.path.exists(path):
        raise HTTPException(404, "Archivo no encontrado")
    from fastapi.responses import FileResponse
    return FileResponse(path, media_type="application/json", filename=archivo)


# ── Main ─────────────────────────────────────────────────────────────────────

# ── Onboarding automático ──────────────────────────────────────────────────────
@app.post("/onboarding")
async def onboarding_endpoint(request: Request):
    """
    Activa el onboarding completo de una empresa nueva.
    Se llama desde el Classifier después de aprobar el mapeo.
    Body: { "empresa": "slug", "cfg": {...json completo...} }
    """
    import sys, threading
    sys.path.insert(0, str(SCRIPT_DIR))

    try:
        body    = await request.json()
        empresa = body.get("empresa", "").strip().lower().replace(" ", "_")
        cfg     = body.get("cfg", {})

        if not empresa or not cfg:
            raise HTTPException(400, "Se requieren 'empresa' y 'cfg'")

        if not re.match(r'^[a-z0-9_]+$', empresa):
            raise HTTPException(400, "El slug solo puede tener letras minúsculas, números y guiones bajos")

        # Ejecutar en background para no bloquear la respuesta
        from onboarding import onboarding
        threading.Thread(
            target=onboarding,
            args=(empresa, cfg),
            daemon=True
        ).start()

        return JSONResponse({
            "status":  "iniciado",
            "empresa": empresa,
            "mensaje": f"Onboarding de '{empresa}' iniciado. Recibirás un email cuando termine.",
        })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error iniciando onboarding: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8001)
    parser.add_argument("--host", default="127.0.0.1")
    args = parser.parse_args()
    import uvicorn
    try:
        from jobs.sentry_config import init_sentry, capturar_error
    except Exception:
        pass
    print(f"""
╔══════════════════════════════════════════════════════╗
║       KraftDo Classifier — UI de Clasificación      ║
╚══════════════════════════════════════════════════════╝

🌐 Abre en tu navegador:
   http://{args.host}:{args.port}

Ctrl+C para detener.
""")
    uvicorn.run("classifier:app", host=args.host, port=args.port, reload=False)
